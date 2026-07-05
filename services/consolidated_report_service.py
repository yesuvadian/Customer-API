"""
Consolidated Test Report Service
================================
Generates a single combined PDF spanning MULTIPLE equipment and MULTIPLE test
types, filtered by:
  • department subtree (user's hierarchy)
  • equipment type
  • selected test types (multi-select)
  • tested-at date range

Layout (per the KPTCL combined report style):
  TITLE
  └─ for each equipment (grouped):
       Equipment header (identity auto-pulled from Equipment register)
       └─ for each matching TestResult (deduplicated):
            Section — <test name> [status]
            (reuses ReportService._build_result_section — full tables,
             cross-session deviations, etc.)

Test-type agnostic: any test type (tan-delta, IDAX, oil BDV, DFR, future) is
rendered via the same generic section builders — no per-test code.
"""

import io
from datetime import datetime, date
from typing import Optional
from uuid import UUID

from sqlalchemy.orm import Session, joinedload
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak, Table, TableStyle

from models import TestResult, TestingRequest, Equipment, OrgDepartment, CategoryDetails
from services.report_service import ReportService
from utils.common_service import get_dept_subtree_ids


class ConsolidatedReportService:
    def __init__(self, db: Session):
        self.db = db
        self._rs = ReportService(db)  # reuse section builders + styles

    # ─── Public entry ────────────────────────────────────────────────────────
    def _query_grouped(
        self,
        department_id: UUID,
        equipment_type_id: int,
        test_type_ids: list[int],
        date_from: Optional[date],
        date_to: Optional[date],
        equipment_ids: Optional[list[UUID]] = None,
    ) -> dict:
        """Run the filtered query and return {eq_id: {request, results}} grouped + deduped."""
        dept_ids = get_dept_subtree_ids(self.db, department_id)

        # Only include terminal states — testing is done and reviewer has approved.
        # outcome_active = approved + follow-up action dispatched (most common)
        # closed         = approved + no follow-up action needed
        # completed      = approved + equipment replacement finance-approved
        # commissioned   = TAQC inspection approved + equipment created
        from models import TestingRequestStatus as TRS
        _terminal = [
            TRS.outcome_active,
            TRS.closed,
            TRS.completed,
            TRS.commissioned,
        ]
        q = (
            self.db.query(TestResult)
            .join(TestingRequest, TestResult.testing_request_id == TestingRequest.id)
            .options(joinedload(TestResult.test_session))
            .filter(
                TestingRequest.department_id.in_(dept_ids),
                TestingRequest.equipment_type_id == equipment_type_id,
                TestingRequest.status.in_(_terminal),
            )
        )
        if test_type_ids:
            q = q.filter(TestingRequest.test_type_id.in_(test_type_ids))
        if equipment_ids:
            q = q.filter(TestingRequest.equipment_id.in_(equipment_ids))
        if date_from:
            q = q.filter(TestResult.tested_at >= datetime.combine(date_from, datetime.min.time()))
        if date_to:
            q = q.filter(TestResult.tested_at <= datetime.combine(date_to, datetime.max.time()))

        results = q.order_by(TestResult.tested_at.asc()).all()

        grouped: dict = {}
        for r in results:
            req = self.db.query(TestingRequest).filter(
                TestingRequest.id == r.testing_request_id
            ).first()
            eq_id = str(req.equipment_id) if req and req.equipment_id else "unassigned"
            grouped.setdefault(eq_id, {"request": req, "results": []})["results"].append(r)

        for g in grouped.values():
            g["results"] = self._dedup_results(g["results"])
        return grouped

    def generate(
        self,
        department_id: UUID,
        equipment_type_id: int,
        test_type_ids: list[int],
        date_from: Optional[date],
        date_to: Optional[date],
        equipment_ids: Optional[list[UUID]] = None,
    ) -> bytes:
        """Build the consolidated PDF and return raw bytes."""
        grouped = self._query_grouped(
            department_id, equipment_type_id, test_type_ids,
            date_from, date_to, equipment_ids,
        )

        # ── Build PDF ──────────────────────────────────────────────────────────
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=A4,
            topMargin=15 * mm, bottomMargin=15 * mm,
            leftMargin=15 * mm, rightMargin=15 * mm,
        )
        styles = self._rs._build_styles()
        story = []

        story.extend(self._rs._build_header(styles, None))
        story.append(Paragraph(
            self._title_line(equipment_type_id, date_from, date_to),
            styles["ReportSubTitle"],
        ))
        story.append(Spacer(1, 4 * mm))

        if not grouped:
            story.append(Paragraph("No test results found for the selected criteria.", styles["Normal"]))
            doc.build(story)
            out = buffer.getvalue()
            buffer.close()
            return out

        first = True
        for eq_id, g in grouped.items():
            if not first:
                story.append(PageBreak())
            first = False

            # Equipment identity header (auto-pulled from register)
            story.extend(self._build_equipment_header(styles, eq_id, g["request"]))
            story.append(Spacer(1, 3 * mm))

            # Each test result rendered via the shared generic section builder
            for idx, result in enumerate(g["results"]):
                story.extend(self._rs._build_result_section(styles, result, idx))
                story.append(Spacer(1, 4 * mm))

        doc.build(story)
        out = buffer.getvalue()
        buffer.close()
        return out

    def generate_excel(
        self,
        department_id: UUID,
        equipment_type_id: int,
        test_type_ids: list[int],
        date_from: Optional[date],
        date_to: Optional[date],
        equipment_ids: Optional[list[UUID]] = None,
    ) -> bytes:
        """
        Pivot layout: one sheet per equipment.
        Rows = parameters, Columns = test dates.
        Header block shows equipment identity (station, capacity, serial, etc.).
        """
        import re
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        grouped = self._query_grouped(
            department_id, equipment_type_id, test_type_ids,
            date_from, date_to, equipment_ids,
        )

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # ── Styles ──────────────────────────────────────────────────────────────
        hdr_fill   = PatternFill("solid", fgColor="1565C0")
        date_fill  = PatternFill("solid", fgColor="0D47A1")
        alt_fill   = PatternFill("solid", fgColor="E3F2FD")
        ok_fill    = PatternFill("solid", fgColor="C8E6C9")
        alert_fill = PatternFill("solid", fgColor="FFCDD2")
        pass_fill  = PatternFill("solid", fgColor="A5D6A7")
        fail_fill  = PatternFill("solid", fgColor="EF9A9A")

        hdr_font = Font(bold=True, color="FFFFFF", size=9)
        bold9    = Font(bold=True, size=9)
        norm9    = Font(size=9)
        thin = Side(style="thin", color="CCCCCC")
        bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
        center_a = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_a   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
        right_a  = Alignment(horizontal="right",  vertical="center")

        # ── Helpers ─────────────────────────────────────────────────────────────
        _tpl_cache: dict = {}

        def _template(tkey):
            if tkey not in _tpl_cache:
                from services.evaluation_service import EvaluationService
                _tpl_cache[tkey] = EvaluationService.get_template_data(tkey, self.db) or {}
            return _tpl_cache[tkey]

        def _limit_str(fld: dict) -> str:
            unit  = fld.get("unit", "")
            rules = fld.get("rules") or fld.get("evaluation_rules") or []
            parts = []
            for rule in rules:
                if rule.get("type") != "THRESHOLD":
                    continue
                lo, hi = rule.get("min"), rule.get("max")
                if hi is not None and lo is None:
                    parts.append(f"≤ {hi} {unit} (Max)".strip())
                elif lo is not None and hi is None:
                    parts.append(f"≥ {lo} {unit} (Min)".strip())
                elif lo is not None and hi is not None:
                    parts.append(f"{lo} – {hi} {unit}".strip())
            return "  /  ".join(parts)

        def _eval_map(result) -> dict:
            evmap = {}
            er = result.evaluation_result
            if isinstance(er, dict):
                for ef in er.get("fields", []):
                    evmap[ef.get("key", "")] = ef.get("status", "")
            return evmap

        _used: set = set()

        def _safe_title(raw: str) -> str:
            base = re.sub(r'[\\/*?:\[\]]', '-', raw or "Eq")[:28] or "Eq"
            t, n = base, 1
            while t in _used:
                t = f"{base[:26]}_{n}"; n += 1
            _used.add(t)
            return t

        def _eq_header_pairs(eq) -> list[tuple[str, str]]:
            """
            Build fixed (label, value) pairs from the Equipment register.
            Walks the department hierarchy for Zone and SSMD; reads nameplate_data
            for Capacity; uses Equipment model columns for the rest.
            Returns 9 pairs in display order (3 per row):
              Name Of Station | Capacity      | Serial Number
              SSMD            | Voltage Class | Date Of Commission
              Zone            | Make          | YOM
            """
            def _v(*vals):
                for v in vals:
                    if v not in (None, "", 0):
                        return str(v)
                return "-"

            nd = (eq.nameplate_data or {}) if eq else {}

            # Walk department ancestry: [leaf, parent, grandparent, ...]
            ancestors: list[str] = []
            current_id = eq.department_id if eq else None
            seen: set = set()
            while current_id and current_id not in seen:
                seen.add(current_id)
                dept = self.db.query(OrgDepartment).filter(
                    OrgDepartment.id == current_id).first()
                if not dept:
                    break
                ancestors.append(dept.name)
                current_id = dept.parent_department_id

            station = ancestors[0] if ancestors else "-"
            ssmd    = ancestors[1] if len(ancestors) >= 2 else "-"
            zone    = ancestors[-1] if len(ancestors) >= 3 else (ancestors[-1] if ancestors else "-")

            capacity = _v(
                nd.get("rated_mva"), nd.get("rated_mva_onan"),
                nd.get("capacity_mva"), nd.get("rated_mva_onan_mva"),
            )
            if capacity != "-":
                capacity = f"{capacity} MVA"

            doc_date = "-"
            if eq and eq.commissioned_date:
                doc_date = eq.commissioned_date.strftime("%d-%b-%y")

            return [
                ("Name Of Station", station),
                ("Capacity",        capacity),
                ("Serial Number",   _v(eq.factory_serial_number if eq else None,
                                       nd.get("factory_serial_number"), nd.get("serial_number"))),
                ("SSMD",            ssmd),
                ("Voltage Class",   _v(eq.voltage_class if eq else None,
                                       nd.get("voltage_class"), nd.get("rated_voltage"))),
                ("Date Of Commission", doc_date),
                ("Zone",            zone),
                ("Make",            _v(eq.manufacturer if eq else None,
                                       nd.get("manufacturer"), nd.get("make"))),
                ("YOM",             _v(eq.year_of_manufacture if eq else None,
                                       nd.get("yom"), nd.get("year_of_manufacture"))),
            ]

        # ── No results ───────────────────────────────────────────────────────────
        if not grouped:
            ws = wb.create_sheet("No Data")
            ws["A1"] = "No test results found for the selected criteria."
            buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

        # ── One sheet per equipment ──────────────────────────────────────────────
        for eq_id, g in grouped.items():
            req     = g["request"]
            results = g["results"]

            eq   = (self.db.query(Equipment).filter(Equipment.id == req.equipment_id).first()
                    if req and req.equipment_id else None)
            ueic = (eq.ueic if eq else None) or str(eq_id)[:12]
            ws   = wb.create_sheet(_safe_title(ueic))

            # Group results by template key so each test type gets its own block
            tpl_groups: dict[str, list] = {}
            for res in results:
                tpl_groups.setdefault(res.template_key or "unknown", []).append(res)

            max_dates = max(len(v) for v in tpl_groups.values()) if tpl_groups else 1
            total_cols = max(6, 2 + max_dates)

            # Column widths — A=param, B=limit, C+=dates
            ws.column_dimensions["A"].width = 30
            ws.column_dimensions["B"].width = 34
            for i in range(max_dates):
                ws.column_dimensions[get_column_letter(3 + i)].width = 14

            # ── Equipment header — from Equipment register (not template-driven) ──
            # Title row
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
            tc = ws.cell(row=1, column=1,
                         value=f"Consolidated Test Report  —  {ueic}")
            tc.font = Font(bold=True, size=13, color="1565C0")
            tc.alignment = center_a
            ws.row_dimensions[1].height = 22

            # Build 9-field header from Equipment model + department hierarchy:
            #   Name Of Station | Capacity      | Serial Number
            #   SSMD            | Voltage Class | Date Of Commission
            #   Zone            | Make          | YOM
            hdr_pairs = _eq_header_pairs(eq)

            cur_row = 2
            # Lay out 3 pairs per row: label | value | label | value | label | value
            for i in range(0, len(hdr_pairs), 3):
                batch = hdr_pairs[i:i + 3]
                for j, (label, val) in enumerate(batch):
                    col_l, col_v = 1 + j * 2, 2 + j * 2
                    lc = ws.cell(row=cur_row, column=col_l, value=f"{label} :")
                    lc.font = bold9; lc.alignment = right_a
                    vc = ws.cell(row=cur_row, column=col_v, value=val)
                    vc.font = norm9; vc.alignment = left_a
                ws.row_dimensions[cur_row].height = 16
                cur_row += 1

            cur = cur_row + 1  # one blank spacer row before the data table

            # ── One pivot block per template ─────────────────────────────────────
            for tkey, tresults in tpl_groups.items():
                tpl = _template(tkey)
                n   = len(tresults)

                # Template / test-type name header
                ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=2 + n)
                hc = ws.cell(row=cur, column=1,
                             value=tpl.get("name") or tkey.replace("_", " ").title())
                hc.font = hdr_font; hc.fill = hdr_fill; hc.alignment = center_a
                for ci in range(2, 3 + n):
                    ws.cell(row=cur, column=ci).fill = hdr_fill
                cur += 1

                # Column headers: Parameter | Permissible Limit | date1 | date2 | …
                for ci, label in enumerate(["Parameter", "Permissible Limit"], 1):
                    ch = ws.cell(row=cur, column=ci, value=label)
                    ch.font = hdr_font; ch.fill = hdr_fill
                    ch.alignment = center_a; ch.border = bdr
                for i, res in enumerate(tresults):
                    dt   = res.tested_at or res.cts
                    dstr = dt.strftime("%d-%b-%Y") if dt else f"Test {i + 1}"
                    dc = ws.cell(row=cur, column=3 + i, value=dstr)
                    dc.font = hdr_font; dc.fill = date_fill
                    dc.alignment = center_a; dc.border = bdr
                ws.row_dimensions[cur].height = 28
                cur += 1

                emaps = [_eval_map(res) for res in tresults]

                # ── Helpers for building pivot rows ──────────────────────────────
                def _cond_to_status(cond_val: str) -> str:
                    cv = (cond_val or "").strip().lower()
                    if cv in ("good", "normal"):          return "NORMAL"
                    if cv in ("fair", "alert", "poor",
                              "abnormal", "abnormal / critical"): return "ALERT"
                    return ""

                def _fmt_thresholds(param_thresholds: dict, unit: str) -> str:
                    """Format Good/Fair/Poor bands from thresholds dict into a readable string."""
                    if not param_thresholds:
                        return unit
                    # Use the first voltage-class entry as representative
                    bands = next(iter(param_thresholds.values()), {})
                    parts = []
                    for cond, rng in bands.items():
                        lo, hi = (rng[0] if isinstance(rng, (list, tuple)) and len(rng) > 0 else None,
                                  rng[1] if isinstance(rng, (list, tuple)) and len(rng) > 1 else None)
                        if lo is not None and hi is not None:
                            parts.append(f"{cond}: {lo}–{hi}")
                        elif hi is not None:
                            parts.append(f"{cond}: ≤{hi}")
                        elif lo is not None:
                            parts.append(f"{cond}: ≥{lo}")
                    text = ", ".join(parts)
                    return f"{text}  {unit}".strip() if unit else text

                def _build_row_keyed_rows(fld: dict) -> list[tuple]:
                    """
                    Row-keyed table: default_rows defines the parameter list.
                    A readonly column carries the parameter name; number column(s)
                    carry the measurements.
                    """
                    fkey    = fld["key"]
                    cols    = fld.get("columns", [])
                    def_rows = fld.get("default_rows", [])

                    # Readonly column = parameter identifier
                    id_col  = next((c for c in cols if c.get("type") == "readonly"), None)
                    id_key  = id_col["key"] if id_col else None

                    # Number columns = measurements (exclude hidden / calculated)
                    val_cols = [c for c in cols
                                if c.get("type") == "number" and not c.get("hidden", False)]

                    # Calculated column with thresholds → for limit extraction
                    calc_col = next((c for c in cols if c.get("type") == "calculated"), None)
                    thresholds_map: dict = {}
                    if calc_col:
                        thresholds_map = (calc_col.get("rule") or {}).get(
                            "config", {}).get("thresholds", {})

                    rows_out = []
                    for dr in def_rows:
                        param_name = str(dr.get(id_key, "")) if id_key else ""
                        param_unit = str(dr.get("unit", "")) if "unit" in dr else ""
                        param_thresholds = thresholds_map.get(param_name, {})
                        limit_str = _fmt_thresholds(param_thresholds, param_unit)

                        for vcol in val_cols:
                            vkey   = vcol["key"]
                            vlabel = vcol.get("label", vkey.replace("_", " ").title())
                            row_label = (f"{param_name} ({vlabel})"
                                         if len(val_cols) > 1 else param_name)

                            vals, sts = [], []
                            for ri, res in enumerate(tresults):
                                td_list = (res.test_data or {}).get(fkey) or []
                                val, st = "", ""
                                for row_d in (td_list if isinstance(td_list, list) else []):
                                    if id_key and str(row_d.get(id_key, "")) == param_name:
                                        raw = row_d.get(vkey)
                                        if raw not in (None, ""):
                                            val = str(raw)
                                        # Status from calculated condition column
                                        if calc_col:
                                            cond_val = row_d.get(calc_col["key"], "")
                                            st = _cond_to_status(str(cond_val))
                                        break
                                # No fallback to whole-table status — that would
                                # incorrectly colour every row when any cell is ALERT.
                                vals.append(val); sts.append(st)

                            rows_out.append((row_label, limit_str, vals, sts))

                    return rows_out

                def _build_col_keyed_rows(fld: dict) -> list[tuple]:
                    """
                    Column-keyed table: each non-hidden template column becomes
                    a row, values taken from the first matching dict in the list.
                    """
                    fkey  = fld["key"]
                    cols  = [c for c in fld.get("columns", [])
                             if not c.get("hidden", False)
                             and c.get("type") not in ("readonly", "calculated")]
                    if not cols:
                        for res in tresults:
                            td_l = (res.test_data or {}).get(fkey)
                            if isinstance(td_l, list) and td_l and isinstance(td_l[0], dict):
                                cols = [{"key": k, "label": k.replace("_", " ").title()}
                                        for k in td_l[0].keys()]
                                break

                    rows_out = []
                    for col in cols:
                        ckey   = col["key"]
                        clabel = col.get("label", ckey.replace("_", " ").title())
                        col_limit = col.get("limit") or _limit_str(fld)
                        vals, sts = [], []
                        for ri, res in enumerate(tresults):
                            td_l = (res.test_data or {}).get(fkey)
                            val = ""
                            if isinstance(td_l, list):
                                for row_d in td_l:
                                    if ckey in row_d and row_d[ckey] not in (None, ""):
                                        val = str(row_d[ckey]); break
                            vals.append(val)
                            sts.append(emaps[ri].get(fkey, ""))
                        rows_out.append((clabel, col_limit, vals, sts))
                    return rows_out

                # ── Build pivot rows: TABLE fields only ───────────────────────────
                param_rows: list[tuple] = []
                table_fields = [f for s in tpl.get("sections", [])
                                for f in s.get("fields", []) if f.get("type") == "table"]
                multi_table = len(table_fields) > 1

                for fld in table_fields:
                    # Section header row when multiple tables exist in the template
                    if multi_table:
                        section_label = fld.get("label") or fld["key"].replace("_", " ").title()
                        param_rows.append(("__SECTION__", section_label, [], []))

                    # Row-keyed (default_rows + readonly id column) → row-as-parameter
                    if fld.get("default_rows") and any(
                            c.get("type") == "readonly" for c in fld.get("columns", [])):
                        param_rows.extend(_build_row_keyed_rows(fld))
                    else:
                        param_rows.extend(_build_col_keyed_rows(fld))

                # Fallback when template has no table fields at all
                if not param_rows:
                    seen_k: set = set()
                    for res in tresults:
                        for k, v in (res.test_data or {}).items():
                            if k in seen_k or k == "overall_result":
                                continue
                            if isinstance(v, list) and v and isinstance(v[0], dict):
                                first_row = v[0]
                                for ck in first_row.keys():
                                    cvals = []
                                    for r in tresults:
                                        rows_list = (r.test_data or {}).get(k, [])
                                        cv = ""
                                        if isinstance(rows_list, list):
                                            for rd in rows_list:
                                                if ck in rd and rd[ck] not in (None, ""):
                                                    cv = str(rd[ck]); break
                                        cvals.append(cv)
                                    param_rows.append((ck.replace("_", " ").title(), "",
                                                       cvals, [""] * n))
                                seen_k.add(k)

                # Write parameter rows
                data_rows = [r for r in param_rows if r[0] != "__SECTION__"]
                data_idx  = 0  # separate counter for alternating row colour
                for plabel, limit, vals, sts in param_rows:
                    # Section sub-header (between two table blocks)
                    if plabel == "__SECTION__":
                        data_idx = 0  # reset so each table block starts on the same alternating colour
                        sub_fill = PatternFill("solid", fgColor="1A3A5C")
                        ws.merge_cells(start_row=cur, start_column=1,
                                       end_row=cur, end_column=2 + n)
                        sc = ws.cell(row=cur, column=1, value=limit)  # limit holds label here
                        sc.font = hdr_font; sc.fill = sub_fill; sc.alignment = left_a
                        for ci in range(2, 3 + n):
                            ws.cell(row=cur, column=ci).fill = sub_fill
                        cur += 1
                        continue

                    bg = alt_fill if data_idx % 2 == 0 else None
                    data_idx += 1

                    lc = ws.cell(row=cur, column=1, value=plabel)
                    lc.font = bold9; lc.border = bdr; lc.alignment = left_a
                    if bg: lc.fill = bg

                    limc = ws.cell(row=cur, column=2, value=limit)
                    limc.font = norm9; limc.border = bdr; limc.alignment = left_a
                    if bg: limc.fill = bg

                    for i, (val, st) in enumerate(zip(vals, sts)):
                        vc = ws.cell(row=cur, column=3 + i, value=val)
                        vc.font = norm9; vc.border = bdr; vc.alignment = center_a
                        if st == "ALERT":
                            vc.fill = alert_fill
                        elif st == "NORMAL":
                            vc.fill = ok_fill
                        elif bg:
                            vc.fill = bg
                    cur += 1

                # Overall result row
                or_c = ws.cell(row=cur, column=1, value="Overall Result")
                or_c.font = hdr_font; or_c.fill = hdr_fill
                or_c.border = bdr; or_c.alignment = left_a
                ws.cell(row=cur, column=2).fill = hdr_fill
                ws.cell(row=cur, column=2).border = bdr
                for i, res in enumerate(tresults):
                    ov = (res.overall_result or "-").upper()
                    oc = ws.cell(row=cur, column=3 + i, value=ov)
                    oc.font = Font(bold=True, size=9)
                    oc.fill = (pass_fill if ov in ("PASS", "NORMAL", "OK")
                               else fail_fill if ov in ("FAIL", "ALERT")
                               else hdr_fill)
                    oc.border = bdr; oc.alignment = center_a
                cur += 3  # spacer before next template block

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ─── Helpers ───────────────────────────────────────────────────────────────
    @staticmethod
    def _dedup_results(results: list) -> list:
        """Deduplicate within the same testing request: keep the latest result
        per (request_id, template_key, voltage_level). This collapses combined
        and individual voltage-level results from the same request session, but
        preserves all results across different testing requests so the date
        range filter shows every session within the selected period.
        """
        latest: dict = {}
        for r in results:
            vlevel = ""
            if isinstance(r.test_data, dict):
                vlevel = str(r.test_data.get("voltage_level", ""))
            key = (str(r.testing_request_id), r.template_key or r.test_name or "", vlevel)
            cur = latest.get(key)
            if cur is None or (r.tested_at or r.cts) > (cur.tested_at or cur.cts):
                latest[key] = r
        # Preserve chronological order
        return sorted(latest.values(), key=lambda x: (x.tested_at or x.cts))

    def _title_line(self, equipment_type_id, date_from, date_to) -> str:
        cd = self.db.query(CategoryDetails).filter(CategoryDetails.id == equipment_type_id).first()
        et = cd.name if cd else "Equipment"
        rng = ""
        if date_from or date_to:
            rng = f"  |  {date_from or '…'} to {date_to or '…'}"
        return f"Consolidated Test Report — {et}{rng}"

    def _build_equipment_header(self, styles, eq_id: str, request):
        """Equipment identity block — pulled from the Equipment register."""
        elements = [Paragraph("Equipment Details", styles["SectionHeader"])]

        if eq_id == "unassigned" or not request or not request.equipment_id:
            elements.append(Paragraph("Unlinked equipment", styles["CellValue"]))
            return elements

        eq = self.db.query(Equipment).filter(Equipment.id == request.equipment_id).first()
        nd = (eq.nameplate_data or {}) if eq else {}

        def _v(*vals):
            for v in vals:
                if v not in (None, "", "-"):
                    return str(v)
            return "-"

        ueic = eq.ueic if eq else "-"
        substation = request.department.name if request and request.department else "-"
        capacity = _v(nd.get("rated_mva_onan"))
        capacity = f"{capacity} MVA" if capacity != "-" else "-"
        v_ratio = "-"
        if nd.get("hv_voltage"):
            v_ratio = "/".join(str(x) for x in [nd.get("hv_voltage"), nd.get("mv_voltage"), nd.get("lv_voltage")] if x) + " kV"
        doc_date = "-"
        if eq and eq.commissioned_date:
            doc_date = eq.commissioned_date.strftime("%d.%m.%Y")

        data = [
            ["UEIC", ueic, "Substation", substation],
            ["Make", _v(eq.manufacturer if eq else None, nd.get("manufacturer_name")),
             "Capacity", capacity],
            ["Voltage Ratio", v_ratio,
             "Vector Group", _v(eq.vector_group if eq else None, nd.get("vector_group"))],
            ["Serial Number", _v(eq.factory_serial_number if eq else None, nd.get("factory_serial_number")),
             "Year of Mfg", _v(eq.year_of_manufacture if eq else None, nd.get("year_of_manufacture"))],
            ["DOC", doc_date, "", ""],
        ]
        elements.append(self._rs._detail_table(data, styles))

        # Bushing details (identity) — from the equipment nameplate register
        bushings = nd.get("bushing_details")
        if isinstance(bushings, list) and bushings:
            elements.append(Spacer(1, 2 * mm))
            elements.append(Paragraph("Bushing Details", styles.get("SmallNote", styles["CellLabel"])))
            elements.extend(self._rs._build_list_table(
                "Bushing Details",
                bushings,
                [
                    {"key": "winding",   "label": "Winding / Level"},
                    {"key": "phase",     "label": "Phase"},
                    {"key": "make",      "label": "Make"},
                    {"key": "serial_no", "label": "Sl. No."},
                    {"key": "yo_mfg",    "label": "Y.O. Mfg."},
                ],
            ))
        return elements
