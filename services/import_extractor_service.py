"""
Import Extractor Service — wraps CLI seed scripts as importable API functions.

Supports:
  PDF:   oil_test    → seed_oil_tests_from_pdf.extract_with_easyocr
         tan_delta   → seed_tandelta_from_pdf.extract_with_easyocr
  Excel: oil_test    → seed_oil_tests_from_excel.parse_excel
"""

from __future__ import annotations

import io
import tempfile
from pathlib import Path
from typing import Optional
from uuid import UUID

from sqlalchemy.orm import Session

from models import CategoryDetails, Equipment
from category_labels import RequestCategoryLabels
from test_templates import normalize_row_id


# ── UI category → DB category_type mapping ───────────────────────────────────
# Each key is what the Flutter UI sends; db_category_type is the value stored
# in CategoryDetails.category_type.  request_category is the TestingRequest
# enum value used when creating the TR.

CATEGORY_TYPE_MAP: dict[str, dict] = {
    "condition_monitoring": {
        "label": "Condition Monitoring",
        "db_category_type": RequestCategoryLabels.TEST.key,
        "request_category": RequestCategoryLabels.TEST.key,
    },
    "test": {
        "label": RequestCategoryLabels.TEST.value,
        "db_category_type": RequestCategoryLabels.TEST.key,
        "request_category": RequestCategoryLabels.TEST.key,
    },
    "inspection": {
        "label": RequestCategoryLabels.INSPECTION.value,
        "db_category_type": RequestCategoryLabels.INSPECTION.key,
        "request_category": RequestCategoryLabels.INSPECTION.key,
    },
    "maintenance": {
        "label": RequestCategoryLabels.MAINTENANCE.value,
        "db_category_type": RequestCategoryLabels.MAINTENANCE.key,
        "request_category": RequestCategoryLabels.MAINTENANCE.key,
    },
    "repair": {
        "label": RequestCategoryLabels.REPAIR_LIFECYCLE.value,
        "db_category_type": None,          # no category_type for repair in DB
        "request_category": RequestCategoryLabels.REPAIR_LIFECYCLE.key,
    },
}

# Keep IMPORT_CATEGORIES as an alias so the router import doesn't break.
IMPORT_CATEGORIES = CATEGORY_TYPE_MAP


# ── Extractor registry ────────────────────────────────────────────────────────
# Maps CategoryDetails.name → supported file formats and extractor type string.
# This is a code-level registry — the DB has no concept of which test types
# have PDF/Excel extractors.  Add new entries here when new seed scripts exist.

EXTRACTABLE_TEST_TYPES: dict[str, dict] = {
    # Repair — generic row-by-row Excel import; no PDF OCR
    "Repair":                                            {"excel": "repair"},
    "Transformer Oil Test":                              {"pdf": "oil_test",  "excel": "oil_test"},
    "Insulating Oil Test":                               {"pdf": "oil_test",  "excel": "oil_test"},
    "Oil BDV Test":                                      {"pdf": "oil_test",  "excel": "oil_test"},
    "Transformer Dissolved Gas Analysis (DGA)":          {"pdf": "oil_test", "excel": "oil_test"},
    "Dissolved Gas Analysis":                            {"pdf": "oil_test", "excel": "oil_test"},
    "DGA Test":                                          {"pdf": "oil_test", "excel": "oil_test"},
    "Capacitance & Tan Delta Test (Transformer)":        {"pdf": "tan_delta", "excel": "tan_delta"},
    "Tan-Delta, Capacitance & Insulation Diagnostics":   {"pdf": "tan_delta", "excel": "tan_delta"},
    "Winding Resistance Measurement":              {"pdf": "tan_delta", "excel": "tan_delta"},
    "220kV Bushing Tan-Delta Test":                      {"pdf": "tan_delta", "excel": "tan_delta"},
    "66kV Bushing Tan-Delta Test":                       {"pdf": "tan_delta", "excel": "tan_delta"},
}



def get_test_types_for_category(
    category_key: str,
    db: Session,
) -> list[dict]:
    """
    Return active CategoryDetails rows for the given import category.

    Filters by CategoryDetails.category_type and is_active. The toggle endpoint
    in org_test_templates syncs CategoryDetails.is_active when a template is
    disabled, so this filter automatically reflects Template Designer state.
    """
    cfg = CATEGORY_TYPE_MAP.get(category_key)
    if not cfg:
        return []

    db_category_type: str | None = cfg["db_category_type"]
    request_category: str = cfg["request_category"]

    query = db.query(CategoryDetails).filter(CategoryDetails.is_active.is_(True))
    if db_category_type:
        query = query.filter(CategoryDetails.category_type == db_category_type)

    seen_names: set = set()
    results = []
    for ct in query.order_by(CategoryDetails.name).all():
        name = ct.name or ""
        if not name or name in seen_names:
            continue
        seen_names.add(name)
        extractor = EXTRACTABLE_TEST_TYPES.get(name, {})
        from test_templates import get_template_for_test_type as _get_tpl
        has_template = _get_tpl(name) is not None
        results.append({
            "id": ct.id,
            "name": name,
            "description": ct.description,
            "has_pdf": "pdf" in extractor,
            "has_excel": "excel" in extractor or has_template,  # flat-schema upload supported for all
            "has_template": has_template,
            "extractor_type": extractor.get("pdf") or extractor.get("excel"),
            "request_category": request_category,
        })
    return results


# ── Extractor dispatch ────────────────────────────────────────────────────────

def _extractor_type_for(test_type_name: str, file_ext: str) -> str | None:
    """Return extractor_type string or None if not supported."""
    extractor = EXTRACTABLE_TEST_TYPES.get(test_type_name, {})
    ext = file_ext.lower().lstrip(".")
    if ext in ("xlsx", "xls"):
        ext = "excel"
    return extractor.get(ext)


def extract_records(
    file_bytes: bytes,
    filename: str,
    test_type_name: str,
) -> tuple[list[dict], list[str]]:
    """
    Extract structured report dicts from an uploaded PDF or Excel file.
    Returns (records, warnings).
    """
    ext = Path(filename).suffix.lower().lstrip(".")
    extractor_type = _extractor_type_for(test_type_name, ext)

    if ext == "pdf":
        if extractor_type is None:
            return [], [f"No PDF extractor available for '{test_type_name}'."]
        return _extract_pdf(file_bytes, extractor_type)
    elif ext in ("xlsx", "xls"):
        # Always attempt Excel extraction — flat-schema format works for any test type
        return _extract_excel(file_bytes, extractor_type or "flat_schema", test_type_name)
    else:
        return [], [f"Unsupported file format: .{ext}"]


def _extract_pdf(file_bytes: bytes, extractor_type: str) -> tuple[list[dict], list[str]]:
    warnings: list[str] = []
    try:
        import fitz
        from PIL import Image

        doc = fitz.open(stream=file_bytes, filetype="pdf")
        images = []
        for page in doc:
            mat = fitz.Matrix(2.0, 2.0)
            pix = page.get_pixmap(matrix=mat)
            img = Image.open(io.BytesIO(pix.tobytes("png")))
            images.append(img)
        doc.close()

        if not images:
            return [], ["PDF has no pages."]

        if extractor_type == "oil_test":
            from seed_oil_tests_from_pdf import extract_with_easyocr
            records = extract_with_easyocr(images, workers=1)
        elif extractor_type == "tan_delta":
            from seed_tandelta_from_pdf import extract_with_easyocr
            records = extract_with_easyocr(images, workers=1)
        else:
            return [], [f"Unknown extractor type: {extractor_type}"]

        if not records:
            warnings.append("No records could be extracted from the PDF. Check the file format.")

        return records, warnings

    except ImportError as e:
        return [], [f"OCR dependency missing: {e}. Install fitz (PyMuPDF) and easyocr."]
    except Exception as e:
        return [], [f"PDF extraction failed: {e}"]


def _is_flat_schema_format(rows: list) -> bool:
    """
    Detect the flat import schema format:
    Row 1 = visible labels, Row 2 = machine field keys, Row 3 = hints, Row 4+ = data.
    We check that row 2 (index 1) is non-empty and its cells look like snake_case keys
    rather than human labels.
    """
    if len(rows) < 2:
        return False
    key_row = rows[1]
    # Exclude hidden marker cells (start with __) before checking format
    non_empty = [v for v in key_row if v is not None and str(v).strip() and not str(v).startswith("__")]
    if not non_empty:
        return False
    # snake_case keys: lowercase letters, digits, underscores — no spaces
    import re
    return all(re.match(r"^[a-z][a-z0-9_]*$", str(v).strip()) for v in non_empty[:5])


def _parse_flat_schema(rows: list, test_type_name: str) -> list[dict]:
    """
    Parse a flat-schema Excel (generated by /data-import/schema).
    Row 1 = labels (skip), Row 2 = field keys, Row 3 = hints (skip), Row 4+ = data.
    Returns a list of report dicts compatible with build_form_data().
    """
    if len(rows) < 4:
        return []

    import datetime as _dt

    def _cell_str(v):
        if v is None:
            return None
        if isinstance(v, _dt.datetime):
            return v.strftime("%Y-%m-%d")
        if isinstance(v, _dt.date):
            return v.isoformat()
        s = str(v).strip()
        return s if s else None

    keys = [str(v).strip() if v is not None else f"col_{i}"
            for i, v in enumerate(rows[1])]

    records = []
    for row in rows[3:]:  # data starts at row index 3 (Excel row 4)
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        rec: dict = {}
        for k, v in zip(keys, row):
            if k.startswith("__"):   # skip hidden marker cells
                continue
            rec[k] = _cell_str(v)
        records.append(rec)
    return records


def _read_table_sheets(wb, main_sheet_title: str) -> dict[str, list[dict]]:
    """Read extra worksheets generated by the schema builder (one per table field)."""
    table_data: dict[str, list[dict]] = {}
    for sheet_name in wb.sheetnames:
        if sheet_name == main_sheet_title:
            continue
        tws = wb[sheet_name]
        trows = list(tws.iter_rows(values_only=True))
        if len(trows) < 2:
            continue
        t_key_row = [str(v).strip() if v is not None else "" for v in trows[1]]
        # Find the hidden marker __table_key__=<field_key>
        tbl_field_key = None
        for cell_val in t_key_row:
            if cell_val.startswith("__table_key__="):
                tbl_field_key = cell_val.split("=", 1)[1]
                break
        if not tbl_field_key:
            tbl_field_key = sheet_name.lower().replace(" ", "_")
        t_label_row = [str(v).strip() if v is not None else f"col_{i}" for i, v in enumerate(trows[0])]
        t_headers = []
        for i in range(len(t_label_row)):
            k = t_key_row[i] if i < len(t_key_row) else ""
            lbl = t_label_row[i]
            if k.startswith("__") or lbl.startswith("col_"):
                t_headers.append(None)  # sentinel: skip this column
            else:
                t_headers.append(k if k else lbl)
        t_data_start = 3 if len(trows) > 3 else 1
        tbl_rows = []
        for trow in trows[t_data_start:]:
            if all(v is None for v in trow):
                continue
            tbl_row: dict = {}
            for h, v in zip(t_headers, trow):
                if h:  # None sentinel means skip
                    cell = str(v) if v is not None else ""
                    tbl_row[h] = normalize_row_id(cell) if h == "test_configuration" else cell
            if any(v for v in tbl_row.values()):
                tbl_rows.append(tbl_row)
        if tbl_rows:
            table_data[tbl_field_key] = tbl_rows
    return table_data


def _extract_excel(file_bytes: bytes, extractor_type: str, test_type_name: str = "") -> tuple[list[dict], list[str]]:
    warnings: list[str] = []
    try:
        # ── Always attempt flat-schema detection first (works for all test types) ──
        import openpyxl as _opxl
        import tempfile as _tmp
        _tf = _tmp.NamedTemporaryFile(suffix=".xlsx", delete=False)
        _tf.write(file_bytes)
        _tf.flush()
        _tf_path = Path(_tf.name)
        _tf.close()
        try:
            _wb = _opxl.load_workbook(_tf_path, data_only=True)
            # The scalar/main sheet is always named "Import Data" by
            # build_import_schema_workbook (line ~535) — look it up by name
            # rather than trusting wb.active, which reflects whichever tab
            # was last selected in Excel (e.g. a "Bushing Details" tab left
            # active after the user finished filling it in last) and would
            # otherwise cause that table sheet to be misread as the main sheet.
            _ws = _wb["Import Data"] if "Import Data" in _wb.sheetnames else _wb.active
            _rows = list(_ws.iter_rows(values_only=True))
        finally:
            _tf_path.unlink(missing_ok=True)

        if _is_flat_schema_format(_rows):
            records = _parse_flat_schema(_rows, test_type_name)
            if not records:
                warnings.append("No data rows found in the schema template. Fill in rows from row 4 onwards.")
            else:
                table_data = _read_table_sheets(_wb, _ws.title)
                if table_data:
                    for rec in records:
                        rec.update(table_data)
                for rec in records:
                    rec["_flat_schema"] = True          # NEW
            return records, warnings

        if extractor_type == "oil_test":
            from seed_oil_tests_from_excel import parse_excel

            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp.write(file_bytes)
                tmp_path = Path(tmp.name)

            try:
                records = parse_excel(tmp_path)
                # Legacy parser only reads its own known blocks — merge in any
                # generic __table_key__ sheets (e.g. Bushing Details) so they
                # aren't silently dropped when this fallback path is taken.
                _table_data = _read_table_sheets(_wb, _ws.title)
            finally:
                tmp_path.unlink(missing_ok=True)

            if _table_data:
                for rec in records:
                    rec.update(_table_data)
            if not records:
                warnings.append("No records extracted from Excel. Check sheet layout.")
            return records, warnings

        elif extractor_type == "tan_delta":
            from seed_tandelta_from_excel import parse_excel

            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp.write(file_bytes)
                tmp_path = Path(tmp.name)

            try:
                records = parse_excel(tmp_path)
                _table_data = _read_table_sheets(_wb, _ws.title)
            finally:
                tmp_path.unlink(missing_ok=True)

            if _table_data:
                for rec in records:
                    rec.update(_table_data)
            if not records:
                warnings.append("No records extracted from Excel. Check sheet/block layout.")
            return records, warnings

        elif extractor_type == "repair":
            # Generic row-by-row Excel reader for repair records.
            # Treats each non-header row as one record; all columns become form_data fields.
            import openpyxl

            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp.write(file_bytes)
                tmp_path = Path(tmp.name)

            try:
                wb = openpyxl.load_workbook(tmp_path, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
            finally:
                tmp_path.unlink(missing_ok=True)

            if not rows:
                return [], ["Excel file is empty."]

            # Row 1 = visible labels, Row 2 = machine keys, Row 3 = hints, Row 4+ = data
            # Prefer machine keys (row 2) when available; fall back to label row (row 1).
            label_row = [str(v).strip() if v is not None else f"col_{i}" for i, v in enumerate(rows[0])]
            key_row   = [str(v).strip() if v is not None else "" for v in rows[1]] if len(rows) > 1 else []
            headers   = [
                (key_row[i] if i < len(key_row) and key_row[i] and not key_row[i].startswith("__") else label_row[i])
                for i in range(len(label_row))
            ]
            # Data starts at row 4 (index 3) when hint row is present, else row 2 (index 1)
            data_start = 3 if len(rows) > 3 else 1

            records = []
            for row in rows[data_start:]:
                if all(v is None for v in row):
                    continue
                rec: dict = {}
                for h, v in zip(headers, row):
                    rec[h] = str(v) if v is not None else ""
                # Map common column names to standard fields
                rec.setdefault("serial_number", rec.get("serial_number", rec.get("Serial Number", rec.get("serial", ""))))
                rec.setdefault("test_date",     rec.get("test_date",     rec.get("Date of Test", rec.get("Date", ""))))
                rec.setdefault("sub_station",   rec.get("sub_station",   rec.get("Sub Station", rec.get("Substation", ""))))
                records.append(rec)

            # ── Read table sheets and merge into records ───────────────────────
            table_data = _read_table_sheets(wb, ws.title)
            for rec in records:
                rec.update(table_data)

            if not records:
                warnings.append("No data rows found in Excel. Check the file layout.")
            return records, warnings

        else:
            return [], [f"Excel extraction not supported for type: {extractor_type}"]

    except ImportError as e:
        return [], [f"Excel dependency missing: {e}. Install openpyxl."]
    except Exception as e:
        return [], [f"Excel extraction failed: {e}"]


# ── Form data builder ─────────────────────────────────────────────────────────

def build_form_data(report, test_type_name, equipment=None):
    extractor = EXTRACTABLE_TEST_TYPES.get(test_type_name, {})
    extractor_type = extractor.get("pdf") or extractor.get("excel")

    if report.get("_flat_schema"):
        clean = {k: v for k, v in report.items() if k != "_flat_schema"}
        return clean

    if extractor_type == "oil_test":
        from seed_oil_tests_from_pdf import _oil_test_data
        return _oil_test_data(report, equipment)
    elif extractor_type == "tan_delta":
        from seed_tandelta_from_pdf import _tandelta_test_data
        return _tandelta_test_data(report, equipment)
    else:
        return report


# ── Equipment resolver ────────────────────────────────────────────────────────

def resolve_equipment(serial_number: str | None, equipment_id: UUID | None, db: Session):
    """Resolve equipment from serial_number or equipment_id."""
    if equipment_id:
        return db.query(Equipment).filter(Equipment.id == equipment_id).first()
    if serial_number:
        return (
            db.query(Equipment)
            .filter(Equipment.factory_serial_number == serial_number)
            .first()
        )
    return None


# ── Template key lookup ───────────────────────────────────────────────────────

def get_template_key(test_type_name: str) -> str | None:
    from test_templates import TEST_TYPE_TO_TEMPLATE
    return TEST_TYPE_TO_TEMPLATE.get(test_type_name)


# ── Blank import-template (schema) workbook builder ───────────────────────────
# Shared by:
#   - GET /data-import/schema        (generic, no equipment context - always
#     includes every section, visibility_data=None)
#   - GET /testing/results/{id}/import-schema (scoped to one equipment - only
#     includes sections/fields applicable to its voltage_ratio)
#
# Sheet 1 ("Import Data") holds scalar fields: row 1 = visible label, row 2 =
# hidden machine key, row 3 = hint, row 4 = where the value goes. One sheet
# per table field: row 1 = label, row 2 = hidden key (+ a "__table_key__=..."
# marker cell), row 3 = hint, rows 4+ = data (pre-filled from default_rows).
# parse_structured_import_workbook() reads this exact shape back.

def build_import_schema_workbook(tpl: dict | None, visibility_data: dict | None = None):
    """Build a blank Excel import template workbook from a template
    definition dict (as returned by test_templates.get_template_for_test_type).

    visibility_data: when given (e.g. {"voltage_ratio": "400"}), sections/
    fields whose visibility_rule evaluates to False against it are omitted
    entirely - used for a single-equipment-scoped download. Pass None for
    the generic bulk-import download, which has no equipment context yet
    and so must include every section unconditionally.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from services.visibility_rule import is_template_field_visible, evaluate_visibility_rule

    def _visible(section: dict, field: dict) -> bool:
        if visibility_data is None:
            return True
        return is_template_field_visible(section, field, visibility_data)

    def _row_visible(drow: dict) -> bool:
        if visibility_data is None:
            return True
        # Same "unevaluable -> visible" fallback as _visible/is_template_field_visible:
        # only drop a row when its own rule definitely evaluates to False.
        return evaluate_visibility_rule(drow.get("visibility_rule"), visibility_data) is not False

    # Fixed identity columns always present (needed for equipment matching)
    fixed_cols = [
        ("sub_station",   "Sub Station",   "Name of substation / location"),
        ("serial_number", "Serial Number", "Equipment serial number (used to match equipment)"),
        ("test_date",     "Date of Test",  "YYYY-MM-DD format"),
    ]

    # Dynamic scalar columns + collect table field definitions
    dyn_cols: list[tuple[str, str, str]] = []
    table_fields: list[dict] = []
    if tpl:
        for sec in tpl.get("sections", []):
            sec_title = sec.get("title", "")
            for f in sec.get("fields", []):
                if not _visible(sec, f):
                    continue
                ftype = f.get("type", "text")
                if ftype == "table":
                    table_fields.append(f)
                    continue
                if ftype in ("calculated", "readonly"):
                    continue
                if f.get("import_skip"):
                    continue
                key   = f.get("key", "")
                label = f.get("label", key)
                hint  = f.get("unit", "") or f.get("hint", "") or sec_title
                if key:
                    dyn_cols.append((key, label, hint))

    all_cols = fixed_cols + dyn_cols

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Import Data"

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    hint_fill   = PatternFill("solid", fgColor="EFF6FF")
    table_fill  = PatternFill("solid", fgColor="1E3A5F")
    row_fill    = PatternFill("solid", fgColor="F8FAFF")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    hint_font   = Font(italic=True, color="6B7280", size=9)
    key_font    = Font(color="FFFFFF", size=1)

    for ci, (key, label, hint) in enumerate(all_cols, start=1):
        # Row 1: visible header label
        h = ws.cell(row=1, column=ci, value=label)
        h.font      = header_font
        h.fill      = header_fill
        h.alignment = Alignment(horizontal="center", wrap_text=True)

        # Row 2: machine key (hidden — white on white)
        k = ws.cell(row=2, column=ci, value=key)
        k.font = key_font

        # Row 3: hint
        hnt = ws.cell(row=3, column=ci, value=hint)
        hnt.font      = hint_font
        hnt.fill      = hint_fill
        hnt.alignment = Alignment(wrap_text=True)

        # Column width
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(18, len(label) + 4)

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 0   # hide key row
    ws.row_dimensions[3].height = 20
    ws.freeze_panes = "A4"

    # ── One sheet per table field ──────────────────────────────────────────────
    for tf in table_fields:
        tbl_key   = tf.get("key", "table")
        tbl_label = tf.get("label", tbl_key)

        # Only editable, non-calculated columns
        editable_cols = [
            c for c in tf.get("columns", [])
            if c.get("type") not in ("calculated",) and not c.get("read_only")
        ]
        if not editable_cols:
            continue

        # Sheet name: strip Excel-invalid chars, truncate to 31 chars
        import re as _re
        sheet_name = _re.sub(r'[:\\/?*\[\]]', '', tbl_label)[:31]
        tws = wb.create_sheet(title=sheet_name)

        # Row 1: visible label  Row 2: machine key (hidden)  Row 3: unit hint
        for ci, col in enumerate(editable_cols, start=1):
            col_key   = col.get("key", f"col{ci}")
            col_label = col.get("label", col_key)
            col_hint  = col.get("unit", "") or col.get("placeholder", "")

            h = tws.cell(row=1, column=ci, value=col_label)
            h.font      = header_font
            h.fill      = table_fill
            h.alignment = Alignment(horizontal="center", wrap_text=True)

            k = tws.cell(row=2, column=ci, value=col_key)
            k.font = key_font

            hnt = tws.cell(row=3, column=ci, value=col_hint)
            hnt.font      = hint_font
            hnt.fill      = hint_fill
            hnt.alignment = Alignment(wrap_text=True)

            tws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(18, len(col_label) + 4)

        # Row 1 in col 0 (A): store table field key as a hidden marker
        meta_cell = tws.cell(row=2, column=len(editable_cols) + 1, value=f"__table_key__={tbl_key}")
        meta_cell.font = key_font

        tws.row_dimensions[1].height = 28
        tws.row_dimensions[2].height = 0
        tws.row_dimensions[3].height = 18
        tws.freeze_panes = "A4"

        # Pre-fill default rows (starting at row 4), scoped to this
        # equipment's voltage class same as section/field visibility above -
        # e.g. a 400kV transformer only gets (HV+LV)-GND/(HV+LV)-TV/TV-GND,
        # not the 220/66kV-only HV-LV/HV-GND/LV-TV/LV-GND/TV-HV rows too.
        default_rows = [r for r in tf.get("default_rows", []) if _row_visible(r)]
        for ri, drow in enumerate(default_rows, start=4):
            for ci, col in enumerate(editable_cols, start=1):
                val = drow.get(col.get("key", ""), "")
                cell = tws.cell(row=ri, column=ci, value=val if val != "" else None)
                if ri % 2 == 0:
                    cell.fill = row_fill
        # Add 5 blank rows after defaults for extra data
        start_blank = max(4, 4 + len(default_rows))
        for ri in range(start_blank, start_blank + 5):
            for ci in range(1, len(editable_cols) + 1):
                cell = tws.cell(row=ri, column=ci, value=None)
                if ri % 2 == 0:
                    cell.fill = row_fill

    return wb


def parse_structured_import_workbook(file_bytes: bytes, tpl: dict | None = None) -> dict:
    """Parse an Excel file matching the exact structure
    build_import_schema_workbook() produces - i.e. the user's filled-in
    downloaded template - back into a form_data dict ready to merge into a
    live TestResultForm's state.

    Reads by the hidden machine key embedded in row 2 of every sheet, not
    by fuzzy label matching - safe to do because we control both ends of
    this round trip (unlike the messy-real-world-document OCR parsers in
    seed_tandelta_from_pdf.py / seed_tandelta_from_excel.py, which this
    function does NOT share code with).

    tpl, when given, is the live template for the test this upload is
    for - used to reject a structurally-valid import file (real
    "__table_key__=" markers, real "Import Data" sheet) that was filled
    in for a DIFFERENT test type. Every field/table key convention is
    shared across templates (build_import_schema_workbook() is generic),
    so e.g. a capacitance_tandelta_transformer export uploaded onto a
    transformer_oil_test request parses "successfully" as far as sheet
    shape goes, but every key it produces (winding_itc_factor,
    winding_test_results, ...) is foreign to the oil-test template and
    silently matches nothing in the live form. Comparing the parsed keys
    against the template's own field keys catches that case tpl is None
    for the generic /data-import/schema bulk-import flow, which has no
    single test type to validate against - only the per-request
    /testing/requests/{id}/import-upload path (which always has one)
    passes it.
    """
    import openpyxl

    _BAD_FORMAT_MSG = "This file doesn't match this test's import template. please download the correct template and fill it in."
    _WRONG_TEMPLATE_MSG = "This file looks like it's for a different test type. Please download the correct template for this test."

    # Present on essentially every template regardless of test type -
    # sub_station/serial_number/test_date are the fixed identity columns
    # build_import_schema_workbook() adds unconditionally, and
    # overall_result/recommendation come from the "Overall Assessment"
    # section every template carries. Excluded from the wrong-template
    # comparison below since they'd otherwise "match" no matter which two
    # (different) test types are being compared.
    _GENERIC_KEYS = {
        "sub_station", "serial_number", "test_date",
        "overall_result", "recommendation",
    }

    def _template_field_keys(t: dict | None) -> set:
        keys = set()
        for sec in (t or {}).get("sections", []):
            for f in sec.get("fields", []):
                k = f.get("key")
                if k:
                    keys.add(k)
        return keys - _GENERIC_KEYS

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception:
        raise ValueError("Not a valid Excel (.xlsx) file.")

    form_data: dict = {}
    recognised_any_sheet = False

    def _row_has_data(row) -> bool:
        return any(v not in (None, "") for v in row)

    # ── Sheet 1: scalar fields ("Import Data") ──────────────────────────────
    if "Import Data" in wb.sheetnames:
        ws = wb["Import Data"]
        key_row = next(ws.iter_rows(min_row=2, max_row=2), None)
        if key_row is not None:
            recognised_any_sheet = True
            keys = [c.value for c in key_row]
            for row in ws.iter_rows(min_row=4, values_only=True):
                if not _row_has_data(row):
                    continue
                for key, val in zip(keys, row):
                    if key and val not in (None, ""):
                        form_data[str(key)] = val
                break  # exactly one data row expected for a single test result

    # ── Remaining sheets: one per table field ───────────────────────────────
    for sheet_name in wb.sheetnames:
        if sheet_name == "Import Data":
            continue
        ws = wb[sheet_name]
        key_row = next(ws.iter_rows(min_row=2, max_row=2), None)
        if key_row is None:
            continue  # sheet too short to carry the hidden key row - skip quietly

        table_key = None
        marker_idx = None
        for idx, cell in enumerate(key_row):
            if isinstance(cell.value, str) and cell.value.startswith("__table_key__="):
                table_key = cell.value.split("=", 1)[1]
                marker_idx = idx
                break
        if table_key is None:
            continue  # not a sheet this format recognises - skip quietly
        recognised_any_sheet = True

        col_keys = [c.value for c in key_row[:marker_idx]]
        n_cols = len(col_keys)

        rows_out: list[dict] = []
        for row in ws.iter_rows(min_row=4, max_col=n_cols, values_only=True):
            if not _row_has_data(row):
                continue
            row_dict = {}
            for key, val in zip(col_keys, row):
                if key:
                    row_dict[str(key)] = val if val is not None else ""
            rows_out.append(row_dict)

        if rows_out:
            form_data[table_key] = rows_out

    # A file with none of the expected "Import Data" sheet or any
    # __table_key__-marked sheet isn't this test's template at all (e.g. the
    # wrong test's template, or an unrelated spreadsheet) - surface that
    # clearly instead of silently returning an empty form_data, which the
    # caller would otherwise report as a misleading "0 fields imported"
    # success.
    if not recognised_any_sheet:
        raise ValueError(_BAD_FORMAT_MSG)

    # The file parsed as a structurally valid template, but every key it
    # produced may still belong to a DIFFERENT test's template (see the
    # docstring). Only checkable when tpl was actually supplied and has a
    # real field list; a wholly-empty tpl (e.g. an unconfigured request)
    # can't distinguish "wrong template" from "legitimately no fields yet",
    # so skip the check rather than reject every upload for that request.
    template_keys = _template_field_keys(tpl)
    parsed_keys = set(form_data.keys()) - _GENERIC_KEYS
    if tpl is not None and template_keys and parsed_keys and not (parsed_keys & template_keys):
        raise ValueError(_WRONG_TEMPLATE_MSG)

    return form_data


# ── Calculated-column evaluation (server-side authoritative pass) ───────────
#
# The Flutter import review screen (import_data_page.dart) recomputes
# "calculated" table columns (FORMULA / THRESHOLD rules, e.g. "% D.F @ 20C"
# and "Condition") live as the user edits a record, then saves the result
# into that record's form_data. But a record the user never opened in the
# review panel -- e.g. a bulk "Submit All" pass -- is submitted with
# whatever form_data the extractor produced, which never had those derived
# columns computed at all. This mirrors lib/common/rule_engine.dart's
# FORMULA/THRESHOLD handling so every submitted record gets them regardless
# of whether it was individually reviewed client-side.

def _rt_num(v) -> Optional[float]:
    try:
        if v is None or (isinstance(v, str) and not v.strip()):
            return None
        return float(v)
    except (TypeError, ValueError):
        return None


def _rt_has_value(row: dict, key: Optional[str]) -> bool:
    if not key:
        return False
    v = row.get(key)
    return v is not None and str(v).strip() != ""


def _rt_formula(config: dict, row: dict, form_data: dict) -> Optional[float]:
    formula = config.get("formula")
    inputs = config.get("inputs") or {}
    precision = int(config.get("precision", 3) or 3)

    def resolve(k: str) -> Optional[float]:
        ref = inputs.get(k)
        if not isinstance(ref, str):
            return None
        if ref.startswith("$form."):
            return _rt_num(form_data.get(ref[len("$form."):]))
        raw = row.get(ref)
        if raw is None:
            raw = form_data.get(ref)
        return _rt_num(raw)

    if formula == "PRODUCT":
        a, b = resolve("a"), resolve("b")
        if a is None or b is None:
            return None
        return round(a * b, precision)

    # Only PRODUCT is used by the calculated table columns that exist today
    # (df_corrected_20c = df_measured x itc_factor). Extend here if a future
    # template adds a different table-column FORMULA.
    return None


def _rt_threshold(config: dict, row: dict, form_data: dict) -> Optional[str]:
    input_field = config.get("input_field")
    raw = row.get(input_field) if input_field else None
    if raw is None and input_field:
        raw = form_data.get(input_field)
    value = _rt_num(raw)
    if value is None:
        return None

    current = config.get("thresholds")
    for lf in (config.get("lookup_fields") or []):
        if not isinstance(current, dict):
            return None
        if isinstance(lf, dict):
            field_ref = str(lf.get("field", ""))
            mapping = lf.get("mapping") or {}
            if field_ref.startswith("$form."):
                raw_key = form_data.get(field_ref[len("$form."):])
            else:
                raw_key = row.get(field_ref, form_data.get(field_ref))
            key = mapping.get(str(raw_key)) if raw_key is not None else None
        else:
            s = str(lf)
            if s.startswith("$form."):
                key = form_data.get(s[len("$form."):])
            else:
                key = row.get(s, form_data.get(s))
        if key is None:
            return None
        current = current.get(key)

    if not isinstance(current, dict):
        return None

    # Outer descriptive-wrapper auto-unwrap (mirrors rule_engine.dart's
    # _threshold): a single outer key like "220 kV Bushing % D.F @ 20C
    # (IEC OIP)" wraps the real { band: [lo, hi] } leaf with no matching
    # lookup_fields entry to navigate into it.
    if current and all(isinstance(v, dict) for v in current.values()):
        leaf: dict = {}
        for v in current.values():
            leaf.update(v)
    else:
        leaf = current

    bands = []
    for name, rng in leaf.items():
        if not isinstance(rng, list) or len(rng) < 2:
            continue
        bands.append((name, _rt_num(rng[0]), _rt_num(rng[1])))
    if not bands:
        return None
    bands.sort(key=lambda b: (b[1] if b[1] is not None else float("-inf")))

    eps = 1e-9
    n = len(bands)
    for i, (name, lo, hi) in enumerate(bands):
        is_first = i == 0
        is_last = i == n - 1
        last_exclusive = is_last and n > 2
        min_ok = lo is None or (
            (value > lo and abs(value - lo) > eps) if last_exclusive
            else (value > lo or abs(value - lo) <= eps)
        )
        max_ok = hi is None or (
            (value < hi and abs(value - hi) > eps) if is_first
            else (value < hi or abs(value - hi) <= eps)
        )
        if min_ok and max_ok:
            return name

    for name, lo, hi in bands:
        if (lo is None or value >= lo) and (hi is None or value < hi):
            return name
    return None


_RANK = {"NORMAL": 0, "ALERT": 1, "CRITICAL": 2}


def _rt_worst_of_bands(config: dict, row: dict) -> Optional[str]:
    """Mirrors rule_engine.dart's _worstOfBands: evaluates several
    row-scoped fields against their own two-cutoff band (higher-is-better),
    reports the label for the WORST severity found. A field with no value
    yet is skipped rather than treated as a failure; returns None only if
    none of the configured fields have a value."""
    fields = config.get("fields") or {}
    labels = config.get("labels") or {"NORMAL": "Normal", "ALERT": "Alert", "CRITICAL": "Critical"}
    eps = 1e-9

    worst = None
    for field_key, band in fields.items():
        value = _rt_num(row.get(field_key))
        if value is None:
            continue
        normal_min = _rt_num((band or {}).get("normal_min"))
        alert_min = _rt_num((band or {}).get("alert_min"))

        if normal_min is not None and (value > normal_min or abs(value - normal_min) <= eps):
            status = "NORMAL"
        elif alert_min is not None and (value > alert_min or abs(value - alert_min) <= eps):
            status = "ALERT"
        else:
            status = "CRITICAL"

        if worst is None or _RANK[status] > _RANK[worst]:
            worst = status

    if worst is None:
        return None
    return labels.get(worst, worst)


def _rt_row_inputs_present(rule: dict, row: dict) -> bool:
    """True when every ROW-scoped input the rule needs (the row's own
    measurement) is present. $form.-prefixed inputs are shared, section-wide
    values (e.g. an ITC correction factor) whose absence doesn't mean
    "nothing was measured" for this row."""
    rtype = rule.get("type")
    config = rule.get("config") or {}
    if rtype == "THRESHOLD":
        return _rt_has_value(row, config.get("input_field"))
    if rtype == "FORMULA":
        for ref in (config.get("inputs") or {}).values():
            if isinstance(ref, str) and ref.startswith("$form."):
                continue
            if not _rt_has_value(row, ref if isinstance(ref, str) else None):
                return False
        return True
    if rtype == "WORST_OF_BANDS":
        # At least one of the several fields needs a value — unlike THRESHOLD/
        # FORMULA's single input, a partially-filled row (e.g. only CC-LF
        # measured so far) still has something worth classifying.
        return any(_rt_has_value(row, k) for k in (config.get("fields") or {}))
    return True


def apply_calculated_columns(template_key: str, form_data: dict) -> dict:
    """Derives every "calculated" table column (FORMULA / THRESHOLD) from the
    template and writes the result back into form_data, in place. Safe to
    call on any form_data shape -- silently no-ops on unknown template keys
    or fields it can't find."""
    from test_templates import get_template_by_key

    tpl = get_template_by_key(template_key)
    if not tpl:
        return form_data

    for section in tpl.get("sections", []):
        for field in section.get("fields", []):
            if field.get("type") != "table":
                continue
            rows = form_data.get(field.get("key"))
            if not isinstance(rows, list):
                continue
            calc_cols = [
                c for c in field.get("columns", [])
                if c.get("type") == "calculated" and c.get("rule")
            ]
            if not calc_cols:
                continue
            for row in rows:
                if not isinstance(row, dict):
                    continue
                for col in calc_cols:
                    rule = col["rule"]
                    rtype = rule.get("type")
                    config = rule.get("config") or {}
                    if rtype == "FORMULA":
                        computed = _rt_formula(config, row, form_data)
                    elif rtype == "THRESHOLD":
                        computed = _rt_threshold(config, row, form_data)
                    elif rtype == "WORST_OF_BANDS":
                        computed = _rt_worst_of_bands(config, row)
                    else:
                        continue
                    col_key = col.get("key")
                    if not col_key:
                        continue
                    if computed is not None:
                        row[col_key] = computed
                    elif not _rt_row_inputs_present(rule, row):
                        row[col_key] = None
                    # else: a shared/form-level input (e.g. ITC factor) is
                    # missing but this row's own reading exists -- preserve
                    # whatever value the import already carried for this
                    # column rather than blanking out real archived data.
    return form_data
