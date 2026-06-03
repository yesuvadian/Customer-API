"""
Reporting Service
=================
Generic report engine driven entirely from the database.

Flow
----
  1.  Router calls ReportingService.generate(definition_id, params, format, user_id)
  2.  Service fetches the ReportDefinition row -> reads query_key
  3.  _run_query() loads sql_template + org_alias from report_query_keys (5-min
      cache), injects {org_clause}, binds :named params, executes -> list[dict]
  4.  _render_excel() or _render_pdf() converts to bytes (in-memory, no disk I/O)
  5.  Router returns Response(content=bytes, media_type=...)
  6.  ReportLog row is written for audit trail

Adding a new report type requires only a new row in report_query_keys —
no Python code change or deployment.

All queries are org-scoped when org_id is provided.
"""

from __future__ import annotations

import io
import logging
import re
import time
from datetime import datetime, timezone, date
from typing import Optional, Dict, Any
from uuid import UUID

from sqlalchemy.orm import Session
from sqlalchemy import text

from models import ReportDefinition, ReportLog

logger = logging.getLogger(__name__)

# ── Optional rendering deps ────────────────────────────────────────────────

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

try:
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    _HAS_REPORTLAB = True
except ImportError:
    _HAS_REPORTLAB = False


# ── Query SQL cache (loaded from report_query_keys.sql_template) ──────────

_sql_cache:    Optional[Dict[str, Dict[str, Any]]] = None
_sql_cache_ts: float = 0.0
_SQL_CACHE_TTL = 300  # seconds — refresh every 5 minutes


def _load_sql_cache(db: Session) -> Dict[str, Dict[str, Any]]:
    """
    Return {query_key: {sql_template, org_alias}} from DB, refreshed every
    5 minutes.  On DB error the stale cache is served so in-flight reports
    are not disrupted.
    """
    global _sql_cache, _sql_cache_ts
    now_ts = time.monotonic()
    if _sql_cache is not None and (now_ts - _sql_cache_ts) < _SQL_CACHE_TTL:
        return _sql_cache
    try:
        from models import ReportQueryKey
        rows = (
            db.query(ReportQueryKey)
            .filter(
                ReportQueryKey.is_active.is_(True),
                ReportQueryKey.sql_template.isnot(None),
            )
            .all()
        )
        _sql_cache = {
            r.key: {
                "sql_template": r.sql_template,
                "org_alias":    r.org_alias or "",
            }
            for r in rows
        }
        _sql_cache_ts = now_ts
    except Exception as exc:
        import logging
        logging.getLogger(__name__).warning(
            f"[ReportingService] SQL cache refresh failed: {exc}"
        )
    return _sql_cache or {}


# ── Helpers ────────────────────────────────────────────────────────────────

def _p(params: dict, key: str, default=None):
    v = params.get(key, default)
    return v if v not in (None, "", "null") else default


# SQLAlchemy's text() regex uses a negative lookahead (?![:\w]) so it does NOT
# recognise :param when immediately followed by ::type (PostgreSQL cast syntax).
# The raw :param reaches psycopg2 which rejects the colon syntax entirely.
# Fix: rewrite  :param::type  →  CAST(:param AS type)  at execution time.
_CAST_PARAM_RE = re.compile(r':([A-Za-z_][A-Za-z0-9_]*)::([A-Za-z0-9]+)')

def _fix_cast_params(sql: str) -> str:
    """Rewrite :param::pgtype → CAST(:param AS pgtype) so SQLAlchemy binds them."""
    return _CAST_PARAM_RE.sub(r'CAST(:\1 AS \2)', sql)


# ══════════════════════════════════════════════════════════════════════════
# ReportingService
# ══════════════════════════════════════════════════════════════════════════

class ReportingService:

    def __init__(self, db: Session, org_id: Optional[UUID] = None):
        self.db = db
        self.org_id = org_id

    # ── Public ─────────────────────────────────────────────────────────────

    def list_definitions(self, active_only: bool = True) -> list[ReportDefinition]:
        q = self.db.query(ReportDefinition)
        if active_only:
            q = q.filter(ReportDefinition.is_active.is_(True))
        if self.org_id:
            q = q.filter(
                (ReportDefinition.organization_id == self.org_id)
                | ReportDefinition.organization_id.is_(None)
            )
        return q.order_by(ReportDefinition.name).all()

    def get_definition(self, definition_id: UUID) -> Optional[ReportDefinition]:
        return self.db.query(ReportDefinition).filter_by(id=definition_id).first()

    def list_logs(
        self,
        definition_id: Optional[UUID] = None,
        limit: int = 50,
    ) -> list[ReportLog]:
        q = self.db.query(ReportLog)
        if definition_id:
            q = q.filter(ReportLog.definition_id == definition_id)
        if self.org_id:
            q = q.filter(ReportLog.organization_id == self.org_id)
        return q.order_by(ReportLog.cts.desc()).limit(limit).all()

    def create_definition(self, data: dict, user_id: Optional[UUID] = None) -> ReportDefinition:
        defn = ReportDefinition(
            organization_id = self.org_id,
            name            = data["name"],
            description     = data.get("description"),
            query_key       = data["query_key"],
            parameters      = data.get("parameters", {}),
            output_format   = data.get("output_format", "excel"),
            frequency       = data.get("frequency", "on_demand"),
            recipient_roles = data.get("recipient_roles", []),
            is_active       = True,
            is_system       = False,
            created_by      = user_id,
        )
        self.db.add(defn)
        self.db.commit()
        self.db.refresh(defn)
        return defn

    def update_definition(self, definition_id: UUID, data: dict,
                          user_id: Optional[UUID] = None) -> ReportDefinition:
        defn = self.get_definition(definition_id)
        if not defn:
            raise ValueError("Not found")
        for field in ("name", "description", "parameters", "output_format",
                      "frequency", "recipient_roles", "is_active"):
            if field in data:
                setattr(defn, field, data[field])
        defn.modified_by = user_id
        self.db.commit()
        self.db.refresh(defn)
        return defn

    def generate(
        self,
        definition_id: UUID,
        parameters: dict,
        output_format: str,          # "excel" | "pdf"
        user_id: Optional[UUID] = None,
    ) -> tuple[bytes, str, str]:
        """Returns (raw_bytes, filename, content_type)."""
        defn = self.get_definition(definition_id)
        if not defn:
            raise ValueError(f"ReportDefinition {definition_id} not found")

        log = ReportLog(
            definition_id   = definition_id,
            organization_id = self.org_id,
            generated_by    = user_id,
            parameters_used = parameters,
            output_format   = output_format,
            status          = "generating",
            started_at      = datetime.now(timezone.utc),
        )
        self.db.add(log)
        self.db.commit()
        self.db.refresh(log)

        try:
            rows = self._run_query(defn.query_key, parameters)
            ts        = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
            safe_name = defn.name.replace(" ", "_").replace("/", "-")

            if output_format == "pdf":
                raw, content_type = self._render_pdf(rows, defn)
                filename = f"{safe_name}_{ts}.pdf"
            else:
                raw, content_type = self._render_excel(rows, defn)
                filename = f"{safe_name}_{ts}.xlsx"

            log.status       = "completed"
            log.file_name    = filename
            log.file_size    = len(raw)
            log.row_count    = len(rows)
            log.completed_at = datetime.now(timezone.utc)
            defn.last_generated_at = datetime.now(timezone.utc)

            # ── In-app notification for the requesting user ────────────────────
            # Only for on-demand runs (user_id is set).  Scheduled runs already
            # fire notifications via run_scheduled_reports → NotificationService.
            if user_id:
                try:
                    from models import UserNotification
                    fmt_label = output_format.upper()
                    self.db.add(UserNotification(
                        user_id         = user_id,
                        organization_id = self.org_id,
                        event_type      = "report_generated",
                        title           = f"Report ready: {defn.name}",
                        body            = f"{len(rows)} rows · {fmt_label}",
                        severity        = "info",
                        source_id       = log.id,
                        source_type     = "report_log",
                        extra_data      = {
                            "report_name": defn.name,
                            "format":      output_format,
                            "row_count":   len(rows),
                            "file_name":   filename,
                        },
                    ))
                except Exception as notif_exc:
                    logger.warning(f"[Reports] Bell notification failed: {notif_exc}")

            self.db.commit()

        except Exception as exc:
            # The failed query may have left the DB connection in an aborted
            # transaction.  Roll back first so the error-log commit succeeds.
            self.db.rollback()
            log.status        = "failed"
            log.error_message = str(exc)
            log.completed_at  = datetime.now(timezone.utc)
            try:
                self.db.commit()
            except Exception:
                pass   # best-effort — don't mask the original error
            raise

        return raw, filename, content_type

    # ── Query dispatcher ───────────────────────────────────────────────────

    def _run_query(self, query_key: str, params: dict) -> list[dict]:
        """
        Load SQL from report_query_keys (DB-cached, 5-min TTL), inject org
        clause, bind user parameters, and execute.

        To add a new report type: insert a row into report_query_keys with
        sql_template + org_alias.  No Python code change required.
        """
        cache = _load_sql_cache(self.db)
        entry = cache.get(query_key)
        if not entry:
            raise ValueError(f"Unknown query_key: '{query_key}'")

        sql_str   = entry["sql_template"]
        org_alias = entry["org_alias"]

        # Inject org-scoping clause
        if self.org_id and org_alias:
            org_clause = f"AND {org_alias}.organization_id = :org_id"
        else:
            org_clause = ""
        sql_str = sql_str.replace("{org_clause}", org_clause)
        sql_str = _fix_cast_params(sql_str)   # :param::pgtype → CAST(:param AS pgtype)

        # Build bound-param dict — every supported parameter defaults to None
        # so SQL NULL-guards (:p IS NULL OR ...) pass through silently when
        # a caller doesn't supply that filter.
        bound: Dict[str, Any] = {
            "org_id":         str(self.org_id) if self.org_id else None,
            # date range
            "date_from":      None,
            "date_to":        None,
            # integer scalars
            "year":           None,
            "month":          None,
            "months":         None,
            "quarter":        None,
            "period_days":    None,
            # string filters
            "status":         None,
            "category":       None,
            "severity":       None,
            "outcome":        None,
            "equipment_type": None,
            "make":           None,
            "voltage_class":  None,
            # uuid filters
            "workflow_id":    None,
            "department_id":  None,
        }
        # Overlay with caller-supplied values (skip null-ish strings)
        for k, v in params.items():
            if v not in (None, "", "null"):
                bound[k] = v

        return self._exec(text(sql_str), bound)

    # ── Executor ───────────────────────────────────────────────────────────

    def _exec(self, sql, params: dict = None) -> list[dict]:
        result = self.db.execute(sql, params or {})
        keys   = list(result.keys())
        return [dict(zip(keys, row)) for row in result.fetchall()]

    # ── Excel renderer ─────────────────────────────────────────────────────

    def _render_excel(self, rows: list[dict],
                      defn: ReportDefinition) -> tuple[bytes, str]:
        if not _HAS_OPENPYXL:
            raise RuntimeError("openpyxl not installed. Run: pip install openpyxl")

        wb = openpyxl.Workbook()
        ws = wb.active
        # Excel sheet titles cannot contain \ / * ? : [ ] and are max 31 chars.
        _safe_title = re.sub(r'[\\/*?:\[\]]', '-', defn.name or "Report")
        ws.title = _safe_title[:31] or "Report"

        if not rows:
            ws["A1"] = "No data found for the selected parameters."
            buf = io.BytesIO()
            wb.save(buf)
            return (buf.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        # Styles
        hdr_fill  = PatternFill("solid", fgColor="1565C0")
        hdr_font  = Font(bold=True, color="FFFFFF", size=10)
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin      = Side(style="thin", color="CCCCCC")
        bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)
        alt_fill  = PatternFill("solid", fgColor="EBF2FB")

        # Row 1: report title
        cols = len(rows[0])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
        c = ws.cell(row=1, column=1, value=defn.name)
        c.font      = Font(bold=True, size=13, color="1565C0")
        c.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 28

        # Row 2: generated timestamp + description
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
        ts = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
        desc_part = f"  |  {defn.description}" if defn.description else ""
        ws.cell(row=2, column=1, value=f"Generated: {ts}{desc_part}")
        ws.row_dimensions[2].height = 16

        # Row 3: headers
        headers = list(rows[0].keys())
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=3, column=ci,
                        value=h.replace("_", " ").title())
            c.fill      = hdr_fill
            c.font      = hdr_font
            c.alignment = hdr_align
            c.border    = bdr
        ws.row_dimensions[3].height = 20

        # Data rows
        for ri, row in enumerate(rows, 4):
            fill = alt_fill if ri % 2 == 0 else None
            for ci, key in enumerate(headers, 1):
                val = row.get(key)
                if isinstance(val, datetime):
                    val = val.replace(tzinfo=None)
                cell = ws.cell(row=ri, column=ci, value=val)
                if fill:
                    cell.fill = fill
                cell.border    = bdr
                cell.alignment = Alignment(wrap_text=False)

        # Auto-width
        for ci, h in enumerate(headers, 1):
            col_vals = [str(row.get(h) or "") for row in rows]
            max_len  = max(len(h.replace("_", " ").title()),
                           max((len(v) for v in col_vals), default=0))
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 50)

        ws.freeze_panes = "A4"

        buf = io.BytesIO()
        wb.save(buf)
        return (buf.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # ── PDF renderer (ReportLab) ─────────────────────────────────────────────

    def _render_pdf(self, rows: list[dict],
                    defn: ReportDefinition) -> tuple[bytes, str]:
        if not _HAS_REPORTLAB:
            raise RuntimeError(
                "ReportLab is not installed. Run: pip install reportlab"
            )

        buffer = io.BytesIO()
        # Use landscape orientation for better table fit
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), 
                                topMargin=0.4*inch, bottomMargin=0.4*inch,
                                leftMargin=0.4*inch, rightMargin=0.4*inch)
        story = []

        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1565C0'),
            alignment=TA_CENTER,
            spaceAfter=12,
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=11,
            textColor=colors.HexColor('#1565C0'),
            spaceAfter=8,
        )

        # Title
        ts = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
        title = f"{defn.name} — Generated {ts}"
        story.append(Paragraph(title, title_style))
        
        # Description if present
        if defn.description:
            story.append(Paragraph(f"<i>{defn.description}</i>", 
                                  ParagraphStyle('Italic', parent=styles['Normal'], 
                                                fontSize=9, textColor=colors.grey)))
        
        story.append(Spacer(1, 0.15*inch))

        if not rows:
            story.append(Paragraph("No data found for the selected parameters.", 
                                  styles['Normal']))
            doc.build(story)
            buffer.seek(0)
            return buffer.getvalue(), "application/pdf"

        # Prepare table data with headers
        headers = list(rows[0].keys())
        table_data = [
            [h.replace("_", " ").title() for h in headers]  # Header row
        ]
        
        # Add data rows with formatting
        for row in rows:
            row_data = []
            for h in headers:
                val = row.get(h)
                # Format dates
                if isinstance(val, datetime):
                    val = val.replace(tzinfo=None).strftime("%Y-%m-%d %H:%M")
                # Format booleans and None
                elif val is None:
                    val = ""
                else:
                    val = str(val)
                row_data.append(val)
            table_data.append(row_data)

        # Create table with dynamic column widths
        # Calculate rough column widths based on content
        col_count = len(headers)
        page_width = landscape(letter)[0] - 0.8*inch  # Account for margins
        col_width = page_width / col_count

        table = Table(table_data, colWidths=[col_width] * col_count)
        
        # Style the table
        table.setStyle(TableStyle([
            # Header row
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1565C0')),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, 0), 6),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            
            # Data rows
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#333333')),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 1), (-1, -1), 'TOP'),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            
            # Alternate row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#EBF2FB')]),
            
            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ]))
        
        story.append(table)
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue(), "application/pdf"


# ── Scheduled runner (called from APScheduler) ─────────────────────────────

def run_scheduled_reports(db_factory) -> int:
    """Check all scheduled ReportDefinitions and generate those that are due."""
    db = db_factory()
    count = 0
    try:
        now  = datetime.now(timezone.utc)
        defs = db.query(ReportDefinition).filter(
            ReportDefinition.is_active.is_(True),
            ReportDefinition.frequency != "on_demand",
        ).all()
        for defn in defs:
            if not _is_due(defn, now):
                continue
            try:
                svc = ReportingService(db, defn.organization_id)

                # "both" → generate Excel then PDF; otherwise normalise to excel
                formats = (
                    ["excel", "pdf"] if defn.output_format == "both"
                    else [defn.output_format
                          if defn.output_format in ("excel", "pdf")
                          else "excel"]
                )

                # Collect all generated files so one email can carry all
                # attachments (avoids two separate emails for "both" format).
                generated_files: list[dict] = []   # [{content, filename, fmt, rel_path}]

                for fmt in formats:
                    raw, filename, _ = svc.generate(defn.id, {}, fmt)

                    if defn.recipient_roles:
                        rel_path = _save_report_file(filename, raw)
                        generated_files.append(
                            {"content": raw, "filename": filename,
                             "fmt": fmt, "rel_path": rel_path}
                        )

                        # ── Email/SMS via full template system (optional) ───
                        if defn.notification_event:
                            try:
                                from services.notification_service import NotificationService
                                NotificationService(db).fire(
                                    event_type    = defn.notification_event,
                                    context       = {
                                        "report_name":   defn.name,
                                        "report_period": _period_label(defn.frequency, now),
                                        "download_url":  f"/reports/download/{rel_path}",
                                        "format":        fmt.upper(),
                                    },
                                    organization_id = defn.organization_id,
                                    source_id       = defn.id,
                                    source_type     = "report_definition",
                                    severity        = "info",
                                    roles_override  = defn.recipient_roles,
                                )
                            except Exception as notif_exc:
                                logger.warning(
                                    f"[Reports] Email/SMS for '{defn.name}': {notif_exc}"
                                )

                        # ── In-app bell notification (one per format) ──────
                        try:
                            _notify_recipients_inapp(
                                db, defn, filename, fmt,
                                rel_path, now,
                            )
                        except Exception as notif_exc:
                            logger.warning(
                                f"[Reports] In-app notif for '{defn.name}': {notif_exc}"
                            )

                # ── ONE email with all attachments ─────────────────────────
                # Sent after all formats are generated so "both" arrives as a
                # single message with Excel + PDF attached.
                if generated_files:
                    try:
                        _email_report_to_recipients(db, defn, generated_files, now)
                    except Exception as email_exc:
                        logger.warning(
                            f"[Reports] Email for '{defn.name}': {email_exc}"
                        )

                count += 1

            except Exception as exc:
                logger.error(f"[Reports] Scheduled '{defn.name}' failed: {exc}")
    finally:
        db.close()
    return count


def _save_report_file(filename: str, raw: bytes) -> str:
    """Save generated report to uploads/reports/ and return the filename."""
    import os
    folder = "uploads/reports"
    os.makedirs(folder, exist_ok=True)
    path = os.path.join(folder, filename)
    with open(path, "wb") as fh:
        fh.write(raw)
    return filename


_MIME = {
    "excel": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "pdf":   "application/pdf",
}


def _email_report_to_recipients(
    db, defn: "ReportDefinition",
    generated_files: "list[dict]",   # [{content, filename, fmt, rel_path}, ...]
    now: datetime,
) -> None:
    """
    Send all generated report files as attachments in ONE email to every active
    user that holds one of the definition's recipient_roles.

    ``generated_files`` is the list built in run_scheduled_reports — each entry
    contains ``content`` (bytes), ``filename``, ``fmt`` ("excel"|"pdf"), and
    ``rel_path``.  Passing multiple entries lets "both" format deliver a single
    message with Excel + PDF attached rather than two separate emails.

    Uses BCC so recipients cannot see each other's addresses.
    Requires no notification_event template — works for any scheduled definition
    that has recipient_roles set.
    """
    from models import OrgRole, OrgUserRole, User
    from utils.email_service import EmailService

    if not defn.recipient_roles or not generated_files:
        return

    # ── Resolve email addresses ────────────────────────────────────────────
    emails: list[str] = []
    seen_users: set = set()

    for role_ref in defn.recipient_roles:
        role_q = db.query(OrgRole).filter(
            OrgRole.organization_id == defn.organization_id
        )
        try:
            import uuid as _uuid
            role = role_q.filter(OrgRole.id == _uuid.UUID(str(role_ref))).first()
        except (ValueError, AttributeError):
            role = role_q.filter(OrgRole.name == str(role_ref)).first()

        if not role:
            continue

        rows = (
            db.query(OrgUserRole)
            .filter(
                OrgUserRole.org_role_id == role.id,
                OrgUserRole.is_active.is_(True),
            )
            .all()
        )
        for ur in rows:
            if ur.user_id in seen_users:
                continue
            seen_users.add(ur.user_id)
            user = db.query(User).filter(User.id == ur.user_id).first()
            if user and user.email:
                emails.append(user.email)

    if not emails:
        logger.info(f"[Reports] No email recipients found for '{defn.name}'")
        return

    # ── Build attachment list ──────────────────────────────────────────────
    attachments = [
        {
            "content":   gf["content"],
            "filename":  gf["filename"],
            "mime_type": _MIME.get(gf["fmt"], "application/octet-stream"),
        }
        for gf in generated_files
    ]

    # ── Build email body ───────────────────────────────────────────────────
    period     = _period_label(defn.frequency, now)
    fmt_labels = " + ".join(sorted({gf["fmt"].upper() for gf in generated_files}))
    file_list  = ", ".join(gf["filename"] for gf in generated_files)
    subject    = f"[Report] {defn.name} — {period}"
    body_html  = (
        f"<h3 style='color:#1565C0'>{defn.name}</h3>"
        f"<p>Your scheduled report for <strong>{period}</strong> is ready.</p>"
        f"<p><strong>Format:</strong> {fmt_labels} &nbsp;|&nbsp; "
        f"<strong>Attached:</strong> {file_list}</p>"
        f"{'<p>' + defn.description + '</p>' if defn.description else ''}"
        f"<hr style='border:none;border-top:1px solid #eee'/>"
        f"<p style='color:#888;font-size:12px'>"
        f"Generated automatically by CogniWatt SEACMS · {now.strftime('%d %b %Y %H:%M UTC')}"
        f"</p>"
    )

    try:
        EmailService().send_multi_attachment_email_starttls_bcc(
            bcc_emails  = emails,
            subject     = subject,
            body_html   = body_html,
            attachments = attachments,
        )
        logger.info(
            f"[Reports] Emailed '{defn.name}' ({fmt_labels}) "
            f"to {len(emails)} recipient(s) with {len(attachments)} attachment(s)"
        )
    except Exception as exc:
        logger.warning(f"[Reports] Email delivery failed for '{defn.name}': {exc}")


def _notify_recipients_inapp(
    db, defn: "ReportDefinition",
    filename: str, fmt: str,
    rel_path: str, now: datetime,
) -> None:
    """
    Write a UserNotification row for every active user who holds one of the
    definition's recipient_roles.  Works without a notification_event template
    — always fires for scheduled reports that have recipient_roles set.

    recipient_roles is a JSON list of OrgRole UUIDs or role name strings.
    """
    from models import UserNotification, OrgRole, OrgUserRole

    if not defn.recipient_roles:
        return

    period = _period_label(defn.frequency, now)
    notified_users: set = set()

    for role_ref in defn.recipient_roles:
        # Accept both UUID strings and role name strings
        role_q = db.query(OrgRole).filter(
            OrgRole.organization_id == defn.organization_id
        )
        try:
            import uuid as _uuid
            role = role_q.filter(OrgRole.id == _uuid.UUID(str(role_ref))).first()
        except (ValueError, AttributeError):
            role = role_q.filter(OrgRole.name == str(role_ref)).first()

        if not role:
            continue

        user_role_rows = (
            db.query(OrgUserRole)
            .filter(
                OrgUserRole.org_role_id == role.id,
                OrgUserRole.is_active.is_(True),
            )
            .all()
        )
        for ur in user_role_rows:
            if ur.user_id in notified_users:
                continue   # deduplicate when a user holds multiple matching roles
            notified_users.add(ur.user_id)
            db.add(UserNotification(
                user_id         = ur.user_id,
                organization_id = defn.organization_id,
                event_type      = "scheduled_report_ready",
                title           = f"Scheduled report ready: {defn.name}",
                body            = f"{period}  ·  {fmt.upper()}",
                severity        = "info",
                source_id       = defn.id,
                source_type     = "report_definition",
                extra_data      = {
                    "report_name":  defn.name,
                    "format":       fmt,
                    "download_url": f"/reports/download/{rel_path}",
                    "period":       period,
                    "file_name":    filename,
                },
            ))

    if notified_users:
        db.commit()


def _period_label(frequency: str, now: datetime) -> str:
    """Human-readable period label for the notification body."""
    if frequency == "monthly":
        return now.strftime("%B %Y")
    if frequency == "quarterly":
        q = (now.month - 1) // 3 + 1
        return f"Q{q} {now.year}"
    if frequency == "annual":
        return str(now.year - 1)
    if frequency == "weekly":
        return f"Week ending {now.strftime('%d %b %Y')}"
    return now.strftime("%d %b %Y")


def _is_due(defn: ReportDefinition, now: datetime) -> bool:
    if defn.last_generated_at is None:
        return True
    delta_days = (now - defn.last_generated_at).days
    if defn.frequency == "daily":
        return delta_days >= 1
    if defn.frequency == "weekly":
        return delta_days >= 7
    if defn.frequency == "monthly":
        return delta_days >= 28
    if defn.frequency == "quarterly":
        return delta_days >= 89
    if defn.frequency == "annual":
        return delta_days >= 364
    return False
