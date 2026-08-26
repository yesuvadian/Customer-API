"""
KPTCL Capacitance & Tan Delta Excel -> DB seed.

Usage:
    # Seed directly from Excel (also saves JSON)
    python seed_tandelta_from_excel.py --from-excel KPTCL_TanDelta_Devanahalli.xlsx

    # Extract to JSON only -- review before seeding
    python seed_tandelta_from_excel.py --from-excel KPTCL_TanDelta_Devanahalli.xlsx --emit-only

    # Seed from a previously saved / edited JSON
    python seed_tandelta_from_excel.py --from-json data/excel_reports.json

    # Dry run -- show what would be seeded, no DB writes
    python seed_tandelta_from_excel.py --from-excel KPTCL_TanDelta_Devanahalli.xlsx --dry-run

Excel layout (one sheet per equipment; unlike the oil workbook this is NOT
one-column-per-date -- it's one vertical BLOCK per source PDF/report, laid
out left-to-right across the sheet, in the same shape parse_ocr_text()
already produces from OCR):

    Row 0        : "SOURCE: <pdf>.pdf | Test Date: <dd.mm.yyyy>"   <- marks a block's start column
    ...          : "TRANSFORMER DETAILS" / "TRANSFORMER / ICT DETAILS" header,
                    then label/value rows (Sub Station, Serial No., Make, ...)
    ...          : "TEST CONDITIONS" header, then more label/value rows
                    (Weather, Kit Used, Correction Factor, ...)
    ...          : "WINDING TEST RESULTS" header (optionally preceded by a
                    free-text "Note: ..." row) + a fixed 6-column table:
                    Test Configuration | Test Voltage | Capacitance Measured |
                    % D.F Measured | % D.F @ 20C (Corrected) | Previous Test Result
    ...          : "<kV> BUSHING DETAILS" header + 4-col nameplate table
                    (Phase | Make | Sl. No. | Y.O. Mfg.) -- informational only,
                    not needed by the template, so it is parsed but discarded.
    ...          : "<kV> BUSHING TEST RESULTS" header + the same 6-column
                    shape as winding, repeated per voltage class present
                    (400 / 220 / 66 / 33 / 11 kV).
    ...          : "IDAX TEST RESULTS (INSULATION DIAGNOSTICS)" header + a
                    6-column table: Test Configuration | % Moisture |
                    Tr. Analysis (Moisture) | Oil Conductivity (pS/m) |
                    Tr. Analysis (Oil Cond.) | Previous % Moisture
    ...          : "OBSERVATIONS / REMARKS" header + free text.

Any section can instead contain a single "Not present / not tested / not
reported ..." row -- treated as an empty table for that section.

Sheets/blocks whose nameplate section reads "CT / STATION DETAILS" or
"STATION / BAY DETAILS" (Current Transformer test data, not Power
Transformer) are skipped -- there's no CT template to map them onto yet.
"""

import argparse
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl


# ── Column schemas (fixed position, not header-text driven -- header labels
#    are inconsistent across blocks, e.g. a trailing "Previous X" column is
#    sometimes unlabeled even though data is present in it) ────────────────

_WINDING_COLS = ("test_configuration", "voltage_raw", "capacitance_raw", "df_measured_raw", "df_corrected_raw", "previous_raw")
_BUSHING_TEST_COLS = ("bushing", "voltage_raw", "capacitance_raw", "df_measured_raw", "df_corrected_raw", "previous_raw")
_IDAX_COLS = ("test_configuration", "moisture_raw", "tr_analysis_moisture", "oil_conductivity_raw", "tr_analysis_oil", "previous_raw")

_BUSHING_KVS = ("400", "220", "66", "33", "11")

# Nameplate / test-conditions label -> report field, keyed by the label with
# all non-alphanumeric characters stripped and lowercased (so "D.O.C." and
# "Sub Station" normalise to stable keys regardless of punctuation/spacing).
_LABEL_MAP = {
    "substation":            "sub_station",
    "testdate":               "test_date",
    "previoustestdate":       "previous_test_date",
    "make":                   "make",
    "capacity":               "capacity_mva",
    "voltageratio":           "voltage_ratio",
    "serialno":               "serial_number",
    "yearofmfg":              "year_of_manufacture",
    "vectorgroup":            "vector_group",
    "oiltemp":                "oil_temp_c",
    "doc":                    "date_of_commission",
    "weather":                "weather_condition",
    "ambienttemperature":     "ambient_temp_c",
    "conditionofbushings":    "bushing_condition",
    "kitused":                "instrument_make",
    "correctionfactor":       "_correction_factor_raw",
    "itccorrectionfactor":    "_correction_factor_raw",
    "testvoltage":            "test_voltage_kv_raw",
}

# Section-header regexes, matched against the normalised (alnum-only, lower)
# first-cell text of a row whose other cells in the block are empty.
_SECTION_PATTERNS = [
    (re.compile(r"^transformer(ict)?details$"), "nameplate"),
    (re.compile(r"^testconditions$"),            "nameplate"),
    (re.compile(r"^windingtestresults$"),        "winding_pending"),
    (re.compile(r"^idaxtestresults.*$"),         "idax_pending"),
    (re.compile(r"^observationsremarks$"),       "remarks"),
    (re.compile(r"^sweepfrequencyresponseanalysis$"), "idle"),
]
_BUSHING_DETAILS_PAT = re.compile(r"^(\d+)kvbushingdetails$")
_BUSHING_TEST_PAT    = re.compile(r"^(\d+)kvbushingtestresults$")
_CT_SHEET_PAT         = re.compile(r"^(ctstationdetails|ctdetails|cttestresults|station.?baydetails)$")

# Matches a *type-qualified* ITC/Correction Factor label -- "ITC Factor
# (Winding)", "ITC Factor (220kV Bushing)", "ITC Factor (Bushings)" -- as
# opposed to the plain "Correction Factor" / "ITC Correction Factor" labels
# already in _LABEL_MAP. The qualifier lives in the label text itself, not
# the value, so it has to be pulled out of the row-0 cell text directly
# rather than via the flat _LABEL_MAP lookup.
_ITC_ROW_PAT = re.compile(r"^(?:itc|correction)\s*factor\s*\(([^)]*)\)\s*$", re.I)


def _classify_itc_qualifier(qualifier: str) -> str:
    """'Winding' -> 'winding'; '220kV Bushing' -> '220'; 'Bushings' (no kV
    given) -> 'generic', applied to every bushing kV tested in the block
    that has no more specific value of its own."""
    q = qualifier.lower()
    if "wind" in q:
        return "winding"
    m = re.search(r"(\d+)\s*kv", q)
    if m:
        return m.group(1)
    return "generic"


# Matches the "Test Date: DD.MM.YYYY" fragment inside a block's own row-0
# header cell ("SOURCE: <pdf> | Test Date: <date>") -- this is the only place
# the date lives; it never appears in the nameplate label/value rows, so
# without this the block-splitter has no way to tell blocks apart by date.
# A test that spanned multiple days is sometimes written as a day range
# ("19-20.02.2020") -- the optional leading "<day>-" is skipped so the
# captured date is the range's end day, used as the effective test date.
_BLOCK_HEADER_DATE_PAT = re.compile(
    r"test\s*date\s*:\s*(?:[\d]{1,2}\s*[-–]\s*)?([\d]{1,2}[./\-][\d]{1,2}[./\-][\d]{2,4})",
    re.I,
)

_SKIP_ROW_PAT = re.compile(r"^(not present|not tested|not conducted|not reported|note)", re.I)


# ── Small parsing helpers ──────────────────────────────────────────────────

def _norm(s) -> str:
    return re.sub(r"[^a-z0-9]", "", str(s or "").lower())


def _clean(v) -> str:
    return str(v).strip() if v is not None else ""


def _num(v) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = re.sub(r"[^\d.\-]", "", str(v).replace(",", ""))
    try:
        return float(s) if s and s not in ("-", ".") else None
    except ValueError:
        return None


def _parse_date_str(v) -> str | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y", "%d-%b-%Y", "%d-%b-%y", "%d.%m.%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def _normalize_serial(serial: str) -> str:
    if not serial:
        return ""
    s = re.sub(r"\s+", "", serial.upper())
    s = re.sub(r"^(HT|KT)(\d)", r"\1-\2", s)
    return s


def _parse_capacitance_pf(v) -> float | None:
    """'10,682.48 nF' / '554.31 pF' / '7.10 nF' -> pF float (nF converted x1000)."""
    if v is None:
        return None
    s = str(v).replace(",", "").strip()
    m = re.search(r"([\d.]+)\s*(nf|pf)\b", s, re.I)
    if not m:
        return _num(s)
    val = float(m.group(1))
    return val * 1000.0 if m.group(2).lower() == "nf" else val


def _parse_voltage_kv(v, fallback_kv: float | None = None) -> float | None:
    """'10 kV' -> 10.0; '8083 V' -> 8.083 (raw volts, converted to kV);
    'N/A' / '-' / None -> fallback overall test voltage, if any."""
    if v is None:
        return fallback_kv
    s = str(v).strip()
    if s.upper() in ("N/A", "-", ""):
        return fallback_kv
    m = re.search(r"([\d.]+)\s*k\.?v\.?\b", s, re.I)
    if m:
        return float(m.group(1))
    m = re.search(r"([\d.]+)\s*v\b", s, re.I)
    if m:
        return round(float(m.group(1)) / 1000.0, 4)
    return _num(s) or fallback_kv


def _parse_previous(v) -> tuple[float | None, str | None]:
    """'0.25 (20.01.2023)' -> (0.25, '2023-01-20')
    '0.090 (for 5kV, as on 19.05.2017 - Eltel Kit)' -> (0.09, '2017-05-19')
    'N/A' / '- (not available)' / None -> (None, None)"""
    if v is None:
        return None, None
    s = str(v).strip()
    if not s or s.upper().startswith("N/A") or s.startswith("-"):
        return None, None
    m = re.match(r"([\d.]+)", s)
    val = float(m.group(1)) if m else None
    dm = re.search(r"(\d{1,2}[.\-/]\d{1,2}[.\-/]\d{2,4})", s)
    date = _parse_date_str(dm.group(1)) if dm else None
    return val, date


def _parse_idax_previous(v) -> tuple[float | None, float | None]:
    """Returns (previous_moisture, previous_oil_conductivity). Some sheets put
    a previous OIL CONDUCTIVITY value (not moisture) in this column, flagged
    by an explicit '(Oil cond. as on ...: X)' annotation -- the template has
    no field for that, so it's kept separate rather than mis-filed as
    moisture_percent_previous."""
    if v is None:
        return None, None
    s = str(v).strip()
    m = re.search(r"oil\s*cond\w*[^:]*:\s*([\d.]+)", s, re.I)
    if m:
        return None, float(m.group(1))
    val, _ = _parse_previous(s)
    return val, None


def _parse_correction_factor(v) -> dict[str, float]:
    """'0.76 (IEC)' / '1.226 (ITC)' -> {'winding': <val>} (single unqualified
    factor is assumed to apply to the winding measurement, matching the
    convention seen whenever a report *does* spell out multiple factors).
    '1.201 / 1.127 / 1.144 (ITC, winding/400kV/33kV bushings resp.)' ->
    {'winding': 1.201, '400': 1.127, '33': 1.144}, matched positionally
    against the hint list in the trailing parenthetical."""
    if not v:
        return {}
    s = str(v)
    pre = re.split(r"\(", s, maxsplit=1)[0]
    floats = [float(x) for x in re.findall(r"[\d.]+", pre)]
    if not floats:
        return {}
    if len(floats) == 1:
        return {"winding": floats[0]}
    hint_m = re.search(r"\(([^)]*)\)", s)
    hints: list[str] = []
    if hint_m:
        for tok in re.split(r"[,/&]", hint_m.group(1)):
            tok_n = tok.strip().lower()
            if "wind" in tok_n:
                hints.append("winding")
            else:
                kvm = re.search(r"(\d+)\s*kv", tok_n)
                if kvm:
                    hints.append(kvm.group(1))
    result: dict[str, float] = {}
    for f, h in zip(floats, hints):
        result[h] = f
    if not result:
        result["winding"] = floats[0]
    return result


def _row_blank(row: tuple, start: int, width: int) -> bool:
    return all(row[start + i] is None for i in range(width) if start + i < len(row))


# ── Block parser ────────────────────────────────────────────────────────────

def _find_block_starts(rows: list[tuple]) -> list[int]:
    if not rows:
        return []
    row0 = rows[0]
    starts = [i for i, v in enumerate(row0) if v and "source:" in str(v).lower()]
    return starts or [0]


def _parse_block(rows: list[tuple], b_start: int, b_end: int, sheet_name: str) -> dict | None:
    nameplate: dict[str, str] = {}

    # The block's test date lives only in its row-0 "SOURCE: ... | Test Date:
    # ..." header cell, never in the nameplate label/value rows below it —
    # pull it out up front so _build_report()'s _LABEL_MAP-driven lookup
    # (np.get("test_date")) has something to find.
    if rows and b_start < len(rows[0]):
        header_cell = rows[0][b_start]
        if header_cell:
            m = _BLOCK_HEADER_DATE_PAT.search(str(header_cell))
            if m:
                nameplate["test_date"] = m.group(1)

    winding_rows: list[dict] = []
    bushing_rows: dict[str, list[dict]] = {kv: [] for kv in _BUSHING_KVS}
    bushing_details_rows: dict[str, list[dict]] = {kv: [] for kv in _BUSHING_KVS}
    idax_rows: list[dict] = []
    remarks_parts: list[str] = []
    section_notes: dict[str, list[str]] = {}

    state = None                 # 'nameplate' | 'winding_pending' | 'bushingdetails_pending' |
                                  # 'bushingtest_pending' | 'idax_pending' | 'in_table' | 'remarks' | None
    table_type = None            # 'winding' | 'bushing_details' | 'bushing_test' | 'idax'
    current_kv = None
    width = b_end - b_start

    def note(section: str, text: str) -> None:
        section_notes.setdefault(section, []).append(text)

    i = 0
    while i < len(rows):
        row = rows[i]
        cellA = row[b_start] if b_start < len(row) else None

        if cellA is None:
            if state == "in_table" and _row_blank(row, b_start, width):
                state, table_type = None, None
            i += 1
            continue

        text = _clean(cellA)
        norm = _norm(text)

        # ── CT-only block: bail out entirely, no template to map onto ──────
        if _CT_SHEET_PAT.match(norm):
            return None

        matched_section = False
        for pat, new_state in _SECTION_PATTERNS:
            if pat.match(norm):
                state, table_type = new_state, None
                matched_section = True
                break
        if matched_section:
            i += 1
            continue

        bm = _BUSHING_DETAILS_PAT.match(norm)
        if bm:
            current_kv, state, table_type = bm.group(1), "bushingdetails_pending", None
            i += 1
            continue
        bm = _BUSHING_TEST_PAT.match(norm)
        if bm:
            current_kv, state, table_type = bm.group(1), "bushingtest_pending", None
            i += 1
            continue

        # ── Generic fallback: an all-caps, single-cell row we don't
        #    recognise (future section types) resets to idle rather than
        #    being mis-read as nameplate/table data. ───────────────────────
        _TABLE_STATES = {
            "winding_pending", "bushingdetails_pending", "bushingtest_pending",
            "idax_pending", "in_table",
        }

        cellB = row[b_start + 1] if b_start + 1 < len(row) else None
        if (
                cellB is None
                and text
                and re.sub(r"[^A-Za-z]", "", text) == re.sub(r"[^A-Za-z]", "", text).upper()
                and len(text) > 3
                and state not in _TABLE_STATES
                and state != "remarks"
            ):
                state, table_type = None, None
                i += 1
                continue

        if state == "nameplate":
            itc_m = _ITC_ROW_PAT.match(text)
            if itc_m:
                val = _num(cellB)
                if val is not None:
                    nameplate.setdefault("_itc_map", {})[_classify_itc_qualifier(itc_m.group(1))] = val
                i += 1
                continue
            key = _LABEL_MAP.get(norm)
            if key:
                nameplate[key] = _clean(cellB)
            i += 1
            continue

        if state in ("winding_pending", "bushingdetails_pending", "bushingtest_pending", "idax_pending"):
            if _SKIP_ROW_PAT.match(text):
                sect = {"winding_pending": "winding", "bushingdetails_pending": f"bushing_{current_kv}",
                        "bushingtest_pending": f"bushing_{current_kv}", "idax_pending": "idax"}[state]
                note(sect, text)
                if not text.lower().startswith("note"):
                    state, table_type = None, None
                i += 1
                continue
            table_type = {
                "winding_pending": "winding",
                "bushingdetails_pending": "bushing_details",
                "bushingtest_pending": "bushing_test",
                "idax_pending": "idax",
            }[state]
            state = "in_table"
            i += 1
            continue

        if state == "in_table":
            vals = [row[b_start + j] if b_start + j < len(row) else None for j in range(6)]
            if table_type == "winding":
                prev_val, prev_date = _parse_previous(vals[5])
                winding_rows.append({
                    "test_configuration": _clean(vals[0]),
                    "voltage_kv": _parse_voltage_kv(vals[1]),
                    "capacitance_pf": _parse_capacitance_pf(vals[2]),
                    "df_measured": _num(vals[3]),
                    "df_corrected_20c": _num(vals[4]),
                    "df_previous_corrected": prev_val,
                    "_prev_date": prev_date,
                })
            elif table_type == "bushing_test":
                prev_val, prev_date = _parse_previous(vals[5])
                bushing_rows[current_kv].append({
                    "bushing": _clean(vals[0]),
                    "voltage_kv": _parse_voltage_kv(vals[1]),
                    "capacitance_pf": _parse_capacitance_pf(vals[2]),
                    "df_measured": _num(vals[3]),
                    "df_corrected_20c": _num(vals[4]),
                    "df_previous_corrected": prev_val,
                    "_prev_date": prev_date,
                })
            elif table_type == "idax":
                prev_moisture, prev_oilcond = _parse_idax_previous(vals[5])
                idax_rows.append({
                    "test_configuration": _clean(vals[0]),
                    "moisture_percent": _num(vals[1]),
                    "moisture_percent_previous": prev_moisture,
                    "tr_analysis_moisture": _clean(vals[2]) or None,
                    "oil_conductivity_psm": _num(vals[3]),
                    "tr_analysis_oil": _clean(vals[4]) or None,
                    "_prev_oilcond": prev_oilcond,
                })
            elif table_type == "bushing_details":
                # Phase | Make | Sl. No. | Y.O. Mfg. -- nameplate row for one
                # phase of this bushing. Collected raw here; pivoted into the
                # template's detail/r_phase/y_phase/b_phase shape in
                # _build_report() via _pivot_bushing_details_rows().
                bushing_details_rows[current_kv].append({
                    "phase":  _clean(vals[0]),
                    "make":   _clean(vals[1]),
                    "sl_no":  _clean(vals[2]),
                    "yo_mfg": _clean(vals[3]),
                })
            i += 1
            continue

        if state == "remarks":
            remarks_parts.append(text)
            i += 1
            continue

        i += 1

    if not nameplate.get("serial_number"):
        print(f"    [SKIP] Sheet '{sheet_name}' block @col{b_start} -- no serial number (not a transformer block).")
        return None

    return {
        "nameplate": nameplate,
        "winding": winding_rows,
        "bushing": bushing_rows,
        "bushing_details": bushing_details_rows,
        "idax": idax_rows,
        "remarks": " ".join(remarks_parts).strip(),
        "notes": section_notes,
    }


# ── Report assembly (mirrors report_skeleton / _tandelta_test_data shapes) ──

_OVERALL_ALERT_PAT = re.compile(
    r"suggest|arrest|leakage|clarify|variation|higher side|conduct ir test|flagged|service aged|poor|abnormal",
    re.I,
)


def _infer_overall_result(remarks: str) -> str:
    if remarks and _OVERALL_ALERT_PAT.search(remarks):
        return "CONDITIONAL"
    return "PASS"


_PHASE_COL_MAP = {"r": "r_phase", "y": "y_phase", "b": "b_phase"}


def _phase_col(phase_text: str) -> str | None:
    """Map a raw 'Phase' cell (e.g. \"'R' Phase\", \"R Phase\", \"R\") to the
    template's r_phase/y_phase/b_phase column key."""
    m = re.match(r"^[^A-Za-z]*([RYB])", phase_text or "", re.I)
    return _PHASE_COL_MAP.get(m.group(1).lower()) if m else None


def _pivot_bushing_details_rows(rows: list[dict]) -> list[dict]:
    """Pivot raw Phase/Make/Sl.No/Y.O.Mfg rows (one row per phase) into the
    template's bushing_{kv}kv_details shape: one row per detail type
    (Make / Sl. No. / Y.O. Mfg.), with r_phase/y_phase/b_phase columns."""
    pivoted = {
        "Make":     {"detail": "Make"},
        "Sl. No.":  {"detail": "Sl. No."},
        "Y.O. Mfg.": {"detail": "Y.O. Mfg."},
    }
    classified = [(row, _phase_col(row.get("phase", ""))) for row in rows]

    # Some equipment's bushing-details rows are never labelled 'R'/'Y'/'B'
    # Phase at all -- e.g. a single combined "400kV" row, or "33kV-a"/"33kV-b"
    # for a two-unit tertiary bushing -- instead of the usual per-phase
    # triplet (seen throughout e.g. the Hoody R/s workbook). When NONE of a
    # kV group's rows carry a recognisable phase, place them into the
    # r_phase/y_phase/b_phase columns positionally, in source order, rather
    # than silently dropping every row -- which otherwise left the "<kV>
    # Bushing Details" table entirely blank for that voltage class.
    if rows and not any(col for _, col in classified):
        for row, col in zip(rows, ("r_phase", "y_phase", "b_phase")):
            pivoted["Make"][col]      = row.get("make", "")
            pivoted["Sl. No."][col]   = row.get("sl_no", "")
            pivoted["Y.O. Mfg."][col] = row.get("yo_mfg", "")
        return list(pivoted.values())

    for row, col in classified:
        if not col:
            continue
        pivoted["Make"][col]      = row.get("make", "")
        pivoted["Sl. No."][col]   = row.get("sl_no", "")
        pivoted["Y.O. Mfg."][col] = row.get("yo_mfg", "")
    return list(pivoted.values())


def _build_report(block: dict) -> dict:
    np = block["nameplate"]
    winding = block["winding"]
    bushing = block["bushing"]
    bushing_details = block["bushing_details"]
    idax = block["idax"]

    # Base: legacy single-string "1.2 / 1.1 (ITC, winding/220kV resp.)" form,
    # still seen in some blocks. Overlaid with the per-row qualified-label
    # form ("ITC Factor (Winding)" / "ITC Factor (220kV Bushing)" / "ITC
    # Factor (Bushings)", one value per row) collected into _itc_map, which
    # is the shape actually used across this workbook's blocks.
    corr = _parse_correction_factor(np.get("_correction_factor_raw"))
    corr.update(np.get("_itc_map", {}))
    generic_bushing_itc = corr.pop("generic", None)
    overall_test_kv = _parse_voltage_kv(np.get("test_voltage_kv_raw"))

    prev_date_default = _parse_date_str(np.get("previous_test_date"))

    def strip_row(rows: list[dict]) -> list[dict]:
        out = []
        for r in rows:
            r = dict(r)
            r.pop("_prev_date", None)
            r.pop("_prev_oilcond", None)
            out.append(r)
        return out

    def table_prev_date(rows: list[dict]) -> str | None:
        for r in rows:
            if r.get("_prev_date"):
                return r["_prev_date"]
        return prev_date_default

    voltage_ratio = np.get("voltage_ratio", "")
    transformer_type = "400/220/33 kV" if "400" in voltage_ratio else "220/66/11 kV"

    capacity_raw = np.get("capacity_mva", "")
    m = re.search(r"([\d.]+)", capacity_raw)
    capacity_mva = m.group(1) if m else ""

    remarks = block["remarks"]

    report: dict = {
        "sub_station": np.get("sub_station", ""),
        "test_date": _parse_date_str(np.get("test_date")) or "",
        "make": np.get("make", ""),
        "capacity_mva": capacity_mva,
        "voltage_ratio": voltage_ratio,
        "serial_number": _normalize_serial(np.get("serial_number", "")),
        "year_of_manufacture": np.get("year_of_manufacture", ""),
        "vector_group": np.get("vector_group", ""),
        "date_of_commission": _parse_date_str(np.get("date_of_commission")),
        "transformer_type": transformer_type,

        "test_voltage_kv": overall_test_kv,
        "frequency_hz": 50,
        "ambient_temp_c": _num(np.get("ambient_temp_c")),
        "oil_temp_c": _num(np.get("oil_temp_c")),
        "weather_condition": np.get("weather_condition", ""),
        "bushing_condition": np.get("bushing_condition", ""),
        "instrument_make": np.get("instrument_make", ""),

        "winding_itc_factor": corr.get("winding"),
        "winding": strip_row(winding),
        "winding_previous_test_date": table_prev_date(winding) if winding else None,
        "winding_observations": " ".join(block["notes"].get("winding", [])) or "",

        "idax_testing_kit": np.get("instrument_make", ""),
        "idax": strip_row(idax),
        "idax_previous_test_date": table_prev_date(idax) if idax else None,
        "idax_observation": " ".join(block["notes"].get("idax", [])) or "",

        "overall_result": _infer_overall_result(remarks),
        "recommendation": remarks or "Seeded from historical KPTCL Excel record.",
    }

    for kv in _BUSHING_KVS:
        rows = bushing.get(kv, [])
        report[f"bushing_{kv}kv"] = strip_row(rows)
        itc_val = corr.get(kv)
        if itc_val is None and rows:
            itc_val = generic_bushing_itc
        report[f"bushing_{kv}kv_itc_factor"] = itc_val
        report[f"bushing_{kv}kv_previous_test_date"] = table_prev_date(rows) if rows else None
        report[f"bushing_{kv}kv_observations"] = " ".join(block["notes"].get(f"bushing_{kv}", [])) or ""
        report[f"bushing_{kv}kv_details"] = _pivot_bushing_details_rows(bushing_details.get(kv, []))

    return report


# ── Public entry point ──────────────────────────────────────────────────────

def parse_excel(path: Path) -> list[dict]:
    """Parse all sheets of a KPTCL Tan Delta Excel workbook into report dicts,
    one per source-PDF block (matching seed_tandelta_from_pdf's report shape
    so it flows through the existing _tandelta_test_data() builder unchanged)."""
    wb = openpyxl.load_workbook(str(path), data_only=True)
    all_reports: list[dict] = []

    for sheet_name in wb.sheetnames:
        if _norm(sheet_name) == "index":
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        block_starts = _find_block_starts(rows)
        print(f"\n-- Sheet: {sheet_name} - {len(block_starts)} block(s) --")

        for b_idx, b_start in enumerate(block_starts):
            b_end = block_starts[b_idx + 1] if b_idx + 1 < len(block_starts) else max(len(r) for r in rows)
            block = _parse_block(rows, b_start, b_end, sheet_name)
            if block is None:
                continue
            report = _build_report(block)
            all_reports.append(report)
            n_bush = sum(len(report[f"bushing_{kv}kv"]) for kv in _BUSHING_KVS)
            print(f"  [{b_idx + 1}] serial={report['serial_number']}  date={report['test_date']}  "
                  f"winding={len(report['winding'])}  bushing_rows={n_bush}  idax={len(report['idax'])}")

    print(f"\n[PARSED] {len(all_reports)} report(s) across all sheets.")
    return all_reports


# ── Reuse DB seed from PDF seeder ────────────────────────────────────────────

def seed_reports(session, reports: list[dict], dry_run: bool = False):
    """Identical pipeline to seed_tandelta_from_pdf -- reused directly."""
    from seed_tandelta_from_pdf import seed_reports as _seed
    return _seed(session, reports, dry_run=dry_run)


# ── JSON helpers ─────────────────────────────────────────────────────────────

def _save_json(reports: list[dict], output: str | None, source_hint: str) -> Path:
    if output:
        out_path = Path(output)
    else:
        stem = Path(source_hint).stem if source_hint else "tandelta_excel_reports"
        ts = datetime.now().strftime("%d%m%Y_%H%M%S")
        out_path = Path("data") / f"{stem}_{ts}.json"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(reports, f, indent=2, default=str, ensure_ascii=False)
    return out_path


# ── Entry point ──────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Seed Capacitance & Tan Delta test results from KPTCL Excel report.",
        epilog=(
            "Examples:\n"
            "  python seed_tandelta_from_excel.py --from-excel KPTCL_TanDelta_Devanahalli.xlsx\n"
            "  python seed_tandelta_from_excel.py --from-excel KPTCL_TanDelta_Devanahalli.xlsx --emit-only\n"
            "  python seed_tandelta_from_excel.py --from-json data/tandelta_excel_reports.json\n"
            "  python seed_tandelta_from_excel.py --from-excel KPTCL_TanDelta_Devanahalli.xlsx --dry-run\n"
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--from-excel", metavar="PATH", nargs="+", help="Path to .xlsx file(s) or folder(s)")
    group.add_argument("--from-json",  metavar="PATH", nargs="+", help="JSON file(s) or folder(s)")
    parser.add_argument("--output",    metavar="FILE", help="Save extracted JSON to this path")
    parser.add_argument("--emit-only", action="store_true", help="Extract JSON only, skip DB seed")
    parser.add_argument("--dry-run",   action="store_true", help="Parse without writing to DB")
    args = parser.parse_args()

    all_reports: list[dict] = []

    if args.from_excel:
        xl_files: list[Path] = []
        for inp in args.from_excel:
            p = Path(inp)
            if p.is_dir():
                found = sorted(p.rglob("*.xlsx")) + sorted(p.rglob("*.xls"))
                print(f"[FOLDER] {p}: {len(found)} Excel file(s) found.")
                xl_files.extend(found)
            elif p.is_file() and p.suffix.lower() in (".xlsx", ".xls"):
                xl_files.append(p)
            else:
                print(f"[WARN] Skipping {inp} -- not an Excel file or folder.")

        if not xl_files:
            print("[ERROR] No Excel files found.")
            sys.exit(1)

        for xl_path in xl_files:
            print(f"\n[1/3] Parsing Excel: {xl_path.name}")
            all_reports.extend(parse_excel(xl_path))

        print(f"\n[TOTAL] {len(all_reports)} record(s) across {len(xl_files)} file(s).")
        if not all_reports:
            print("[DONE] Nothing extracted.")
            return

        out_path = _save_json(all_reports, args.output, args.from_excel[0])
        print(f"\n[SAVED] {out_path}")
        print("        Review/edit, then run:")
        print(f'        python seed_tandelta_from_excel.py --from-json "{out_path}"')

        if args.emit_only:
            print("\n[--emit-only] JSON saved. Skipping DB seed.")
            return

    else:
        json_files: list[Path] = []
        for inp in args.from_json:
            p = Path(inp)
            if not p.exists():
                print(f"[WARN] Path not found: {p}")
                continue
            if p.is_dir():
                found = sorted(p.rglob("*.json"))
                print(f"[FOLDER] {p}: {len(found)} JSON file(s) found.")
                json_files.extend(found)
            elif p.is_file() and p.suffix.lower() == ".json":
                json_files.append(p)
            else:
                print(f"[WARN] Skipping {p} -- not a JSON file or folder.")

        if not json_files:
            print("[ERROR] No JSON files found.")
            sys.exit(1)

        for jf in json_files:
            print(f"[LOAD] {jf}")
            with open(jf, encoding="utf-8") as f:
                data = json.load(f)
            all_reports.extend(data if isinstance(data, list) else [data])

        print(f"[OK] Loaded {len(all_reports)} report(s) from {len(json_files)} JSON file(s)")

    if not all_reports:
        print("[DONE] Nothing to seed.")
        return

    if args.dry_run:
        print("\n[DRY RUN] Report data:\n")
        print(json.dumps(all_reports, indent=2, default=str))
        return

    print("\n[3/3] Seeding to database...")
    from database import VendorSessionLocal as _SessionLocal
    with _SessionLocal() as session:
        seed_reports(session, all_reports, dry_run=False)


if __name__ == "__main__":
    main()