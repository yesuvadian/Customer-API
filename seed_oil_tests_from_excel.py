"""
KPTCL Transformer Oil Test Excel → DB seed.

Usage:
    # Seed directly from Excel (also saves JSON)
    python seed_oil_tests_from_excel.py --from-excel TRANSFORMER.xlsx

    # Extract to JSON only — review before seeding
    python seed_oil_tests_from_excel.py --from-excel TRANSFORMER.xlsx --emit-only

    # Seed from a previously saved / edited JSON
    python seed_oil_tests_from_excel.py --from-json data/excel_reports.json

    # Dry run — show what would be seeded, no DB writes
    python seed_oil_tests_from_excel.py --from-excel TRANSFORMER.xlsx --dry-run

Excel layout (one sheet per substation):
    Row 1 : Name Of Station  |  Capacity  |  Serial Number   (repeated per transformer block)
    Row 2 : SSMD             |  Voltage Class  |  Date Of Commission
    Row 3 : Zone             |  Make           |  YOM
    Row 4 : Permissible limits header
    Row 5 : Last Filtration Date
    Row 6 : Sampling Date
    Row 7 : Date of Test     |  date1  |  date2  ...        (datetime cells mark data columns)
    Row 8+ : Acidity / Resistivity / Tan-Delta / BDV(T) / BDV(B) / IFT / F P / ppm / Remarks
"""

import argparse
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl


# ── Excel parser ──────────────────────────────────────────────────────────────

# Parameter row label → oil_test key
_PARAM_MAP = {
    "acidity":       "acidity",
    "resistivity":   "resistivity",
    "tan":           "tan_delta",
    "bdv (t)":       "bdv_top",
    "bdv (b)":       "bdv_bottom",
    "ift":           "interfacial_tension",
    "f p":           "flash_point",
    "flash":         "flash_point",
    "ppm":           "water_content",
}


def _extract_kv(cell_text: str, key: str) -> str:
    """Extract value after 'key : value' pattern in a cell string."""
    if not cell_text:
        return ""
    m = re.search(re.escape(key) + r"\s*[:\-]\s*(.+?)(?:\s*$)", cell_text, re.I)
    return m.group(1).strip() if m else ""


def _parse_date(v) -> str | None:
    """Parse a cell value to YYYY-MM-DD string."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d-%b-%Y", "%d-%b-%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def _is_date_cell(v) -> bool:
    """True if the cell value looks like a test date."""
    if isinstance(v, datetime):
        return True
    if isinstance(v, str):
        return bool(re.match(r"\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4}", v.strip()))
    return False


def _param_key(label: str) -> str | None:
    """Map a parameter row label to an oil_test key.
    Collapses internal spaces so 'Aci dity' matches 'acidity'."""
    if not label:
        return None
    # collapse all whitespace so "Aci dity" -> "acidity", "Resis tivity" -> "resistivity"
    l = re.sub(r"\s+", "", str(label)).lower()
    for pat, key in _PARAM_MAP.items():
        if re.sub(r"\s+", "", pat) in l:
            return key
    return None


def _parse_voltage(v: str) -> str:
    """Normalise voltage class string to the primary kV number (e.g. '66/11 KV' → '66')."""
    if not v:
        return ""
    m = re.search(r"(\d+)", v)
    return m.group(1) if m else v.strip()


def parse_excel(path: Path) -> list[dict]:
    """Parse all sheets of the KPTCL oil test Excel into a list of report dicts."""
    wb = openpyxl.load_workbook(str(path), data_only=True)
    all_reports: list[dict] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Detect transformer block start columns by scanning for "Name Of Station"
        # — may be on row 0 or row 1 depending on the sheet layout.
        header_row_idx = None
        for ri in range(min(3, len(rows))):
            if any(v and "name of station" in str(v).lower() for v in rows[ri]):
                header_row_idx = ri
                break

        if header_row_idx is None:
            print(f"  [SKIP] Sheet '{sheet_name}' -- no transformer blocks found.")
            continue

        header_row = rows[header_row_idx]
        block_starts = [
            i for i, v in enumerate(header_row)
            if v and "name of station" in str(v).lower()
        ]

        print(f"\n-- Sheet: {sheet_name} - {len(block_starts)} transformer(s) --")

        for b_idx, b_start in enumerate(block_starts):
            b_end = block_starts[b_idx + 1] if b_idx + 1 < len(block_starts) else len(header_row)

            def cell(row_idx, col_offset):
                """Get cell value at (header_row_idx + row_idx, b_start + col_offset), bounds-safe."""
                ri = header_row_idx + row_idx
                if ri >= len(rows):
                    return None
                row = rows[ri]
                idx = b_start + col_offset
                return row[idx] if idx < len(row) else None

            def find_in_block(row_offset, keyword):
                """Find cell in block row (relative to header) containing keyword."""
                ri = header_row_idx + row_offset
                if ri >= len(rows):
                    return ""
                row = rows[ri]
                for i in range(b_start, min(b_end, len(row))):
                    v = row[i]
                    if v and keyword.lower() in str(v).lower():
                        return _extract_kv(str(v), keyword)
                return ""

            # ── Transformer metadata ──────────────────────────────────────────
            station      = find_in_block(0, "Name Of Station")
            capacity     = find_in_block(0, "Capacity")
            serial       = find_in_block(0, "Serial Number")
            voltage_raw  = find_in_block(1, "Voltage Class")
            doc_raw      = find_in_block(1, "Date Of Commission")
            make         = find_in_block(2, "Make")

            voltage_class = _parse_voltage(voltage_raw)
            doc = _parse_date(doc_raw)

            if not serial:
                print(f"  [SKIP] Block {b_idx + 1} in '{sheet_name}' — no serial number.")
                continue

            print(f"  [{b_idx + 1}] Serial={serial}  Voltage={voltage_raw}  Make={make}")

            # ── Locate "Date of Test" row and find date columns ───────────────
            date_of_test_row_idx = None
            for ri in range(header_row_idx, len(rows)):
                row = rows[ri]
                v = row[b_start] if b_start < len(row) else None
                if v and "date of test" in str(v).lower():
                    date_of_test_row_idx = ri
                    break

            if date_of_test_row_idx is None:
                print(f"    [WARN] 'Date of Test' row not found in block {b_idx + 1}.")
                continue

            dot_row = rows[date_of_test_row_idx]
            # Collect (abs_col, date_str) for all date cells in this block
            date_cols: list[tuple[int, str]] = []
            for abs_col in range(b_start, min(b_end, len(dot_row))):
                v = dot_row[abs_col]
                if _is_date_cell(v):
                    d = _parse_date(v)
                    if d:
                        date_cols.append((abs_col, d))

            if not date_cols:
                print(f"    [WARN] No test dates found in block {b_idx + 1}.")
                continue

            print(f"    Dates: {[d for _, d in date_cols]}")

            # ── Parameter rows ────────────────────────────────────────────────
            # Build a map: param_key → {abs_col: value}
            param_values: dict[str, dict[int, float | None]] = {}

            for ri in range(date_of_test_row_idx + 1, len(rows)):
                row = rows[ri]
                label_cell = row[b_start] if b_start < len(row) else None
                if not label_cell:
                    continue
                pk = _param_key(str(label_cell))
                if not pk:
                    continue
                col_vals: dict[int, float | None] = {}
                for abs_col, _ in date_cols:
                    v = row[abs_col] if abs_col < len(row) else None
                    try:
                        col_vals[abs_col] = float(v) if v is not None and v != "" else None
                    except (ValueError, TypeError):
                        col_vals[abs_col] = None
                # Don't overwrite if already have this key (e.g. BDV(T) before BDV(B))
                if pk not in param_values:
                    param_values[pk] = col_vals

            # ── Build one report dict per test date ───────────────────────────
            for abs_col, date_str in date_cols:
                oil = {
                    "acidity":             param_values.get("acidity",             {}).get(abs_col),
                    "resistivity":         param_values.get("resistivity",         {}).get(abs_col),
                    "tan_delta":           param_values.get("tan_delta",           {}).get(abs_col),
                    "bdv_top":             param_values.get("bdv_top",             {}).get(abs_col),
                    "bdv_bottom":          param_values.get("bdv_bottom",          {}).get(abs_col),
                    "interfacial_tension": param_values.get("interfacial_tension", {}).get(abs_col),
                    "flash_point":         param_values.get("flash_point",         {}).get(abs_col),
                    "water_content":       param_values.get("water_content",       {}).get(abs_col),
                    "remarks": {},
                }

                # Skip if no meaningful data at all
                has_data = any(
                    v is not None for v in [
                        oil["acidity"], oil["resistivity"], oil["tan_delta"],
                        oil["bdv_bottom"], oil["interfacial_tension"],
                        oil["flash_point"], oil["water_content"],
                    ]
                )
                if not has_data:
                    continue

                report = {
                    "sub_station":        station,
                    "sample_no":          "",
                    "test_date":          date_str,
                    "capacity_mva":       capacity,
                    "voltage_class":      voltage_class,
                    "make":               make,
                    "serial_number":      serial,
                    "date_of_commission": doc,
                    "standard":           "IS 10593:2017",
                    "oil_test":           oil,
                    "dga":                None,   # no DGA in this Excel
                }
                all_reports.append(report)
                print(f"    + {date_str}  acidity={oil['acidity']}  bdv_b={oil['bdv_bottom']}")

    print(f"\n[PARSED] {len(all_reports)} test record(s) across all sheets.")
    return all_reports


# ── Reuse DB seed from PDF seeder ─────────────────────────────────────────────

def seed_reports(session, reports: list[dict], dry_run: bool = False):
    """Identical pipeline to seed_oil_tests_from_pdf — reused directly."""
    from seed_oil_tests_from_pdf import seed_reports as _seed
    return _seed(session, reports, dry_run=dry_run)


# ── JSON helpers ──────────────────────────────────────────────────────────────

def _save_json(reports: list[dict], output: str | None, source_hint: str) -> Path:
    if output:
        out_path = Path(output)
    else:
        stem = Path(source_hint).stem if source_hint else "excel_reports"
        ts = datetime.now().strftime("%d%m%Y_%H%M%S")
        out_path = Path("data") / f"{stem}_{ts}.json"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(reports, f, indent=2, default=str, ensure_ascii=False)
    return out_path


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Seed transformer oil test results from KPTCL Excel report.",
        epilog=(
            "Examples:\n"
            "  python seed_oil_tests_from_excel.py --from-excel TRANSFORMER.xlsx\n"
            "  python seed_oil_tests_from_excel.py --from-excel TRANSFORMER.xlsx --emit-only\n"
            "  python seed_oil_tests_from_excel.py --from-json data/excel_reports.json\n"
            "  python seed_oil_tests_from_excel.py --from-excel TRANSFORMER.xlsx --dry-run\n"
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--from-excel", metavar="PATH", nargs="+", help="Path to .xlsx file(s) or folder(s)")
    parser.add_argument("--output",    metavar="FILE", help="Save extracted JSON to this path")
   
    parser.add_argument("--emit-only", action="store_true", help="Extract JSON only, skip DB seed")
    parser.add_argument("--dry-run",   action="store_true", help="Parse without writing to DB")
    group.add_argument(
    "--from-json",
    metavar="PATH",
    nargs="+",
    help="JSON file(s) or folder(s)"
)
    args = parser.parse_args()

    if args.from_excel:
        # Expand files and folders into a flat list of .xlsx paths
        xl_files: list[Path] = []
        for inp in args.from_excel:
            p = Path(inp)
            if p.is_dir():
                found = sorted(p.rglob("*.xlsx")) + sorted(p.rglob("*.xls"))
                print(f"[FOLDER] {p}: {len(found)} Excel file(s) found.")
                xl_files.extend(found)
            elif p.is_file() and p.suffix.lower() in (".xlsx", ".xls"):
                xl_files.append(p)
            else:
                print(f"[WARN] Skipping {inp} — not an Excel file or folder.")

        if not xl_files:
            print("[ERROR] No Excel files found.")
            sys.exit(1)

        all_reports: list[dict] = []
        for xl_path in xl_files:
            print(f"\n[1/3] Parsing Excel: {xl_path.name}")
            reports = parse_excel(xl_path)
            all_reports.extend(reports)

        print(f"\n[TOTAL] {len(all_reports)} record(s) across {len(xl_files)} file(s).")

        if not all_reports:
            print("[DONE] Nothing extracted.")
            return

        source_hint = args.from_excel[0]
        out_path = _save_json(all_reports, args.output, source_hint)
        print(f"\n[SAVED] {out_path}")
        print(f"        Review/edit, then run:")
        print(f"        python seed_oil_tests_from_excel.py --from-json \"{out_path}\"")

        if args.emit_only:
            print("\n[--emit-only] JSON saved. Skipping DB seed.")
            return

    else:
        json_files: list[Path] = []

        for inp in args.from_json:
            p = Path(inp)

            if not p.exists():
                print(f"[WARN] Path not found: {p}")
                continue

            if p.is_dir():
                found = sorted(p.rglob("*.json"))
                print(f"[FOLDER] {p}: {len(found)} JSON file(s) found.")
                json_files.extend(found)

            elif p.is_file() and p.suffix.lower() == ".json":
                json_files.append(p)

            else:
                print(f"[WARN] Skipping {p} — not a JSON file or folder.")

        if not json_files:
            print("[ERROR] No JSON files found.")
            sys.exit(1)

        all_reports = []

        for json_file in json_files:
            print(f"[LOAD] {json_file}")

            with open(json_file, encoding="utf-8") as f:
                data = json.load(f)

            if isinstance(data, list):
                all_reports.extend(data)
            else:
                all_reports.append(data)

        print(
            f"[OK] Loaded {len(all_reports)} report(s) "
            f"from {len(json_files)} JSON file(s)"
        )

    if not all_reports:
        print("[DONE] Nothing to seed.")
        return

    if args.dry_run:
        print("\n[DRY RUN] Report data:\n")
        print(json.dumps(all_reports, indent=2, default=str))
        return

    print("\n[3/3] Seeding to database...")
    from database import VendorSessionLocal as _SessionLocal
    with _SessionLocal() as session:
        seed_reports(session, all_reports, dry_run=False)


if __name__ == "__main__":
    main()
