"""
Organization user management endpoints.
Provides CRUD operations for users within organizations.
"""

import io
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from uuid import UUID
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from database import get_db
from middleware.org_auth import require_org_member, require_org_admin
from models import User, OrgRole, OrgUserRole, OrgDepartment
from schemas import (
    OrgUserCreate,
    OrgUserUpdate,
    User as UserOut,
    RoleAssignment,
    OrgUserRoleOut,
    OrgUserWithRoles,
    BulkUserImportRow,
    BulkUserImportResponse
)
from services.org_user_service import OrgUserService


router = APIRouter(
    prefix="/organizations/{org_id}/users",
    tags=["org-users"]
)


@router.get("/bulk-import/schema")
def download_bulk_import_schema(
    org_id: UUID,
    department_name: Optional[str] = None,
    export_existing: bool = False,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_org_member)
):
    """
    Download an Excel template for bulk user import.
    Pass ?export_existing=true to pre-fill with all current users (email, name,
    phone, employee_id, department, role) so you can review or extend them.
    Sheet 2 'Available Roles' always lists current org roles.
    """
    wb = openpyxl.Workbook()

    HEADERS = ["email", "firstname", "lastname", "phone_number",
               "employee_id", "department_name", "role_name"]

    header_fill   = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font   = Font(bold=True, color="FFFFFF")
    existing_fill = PatternFill(start_color="F0FFF4", end_color="F0FFF4", fill_type="solid")
    example_font  = Font(color="6B7280")

    # ── Sheet 1: Users ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Users"

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 26

    if export_existing:
        # Build lookup maps
        dept_map  = {d.id: d.name for d in db.query(OrgDepartment).filter_by(
            organization_id=org_id, is_active=True).all()}
        role_map  = {r.id: r.name for r in db.query(OrgRole).filter_by(
            organization_id=org_id, is_active=True).all()}

        # Fetch every user in the org with their primary role assignment
        users = db.query(User).filter_by(
            organization_id=org_id, isactive=True).order_by(User.email).all()

        if users:
            # Build user→(dept, role) from OrgUserRole — take the first active one per user
            user_role_rows = (
                db.query(OrgUserRole)
                .filter(OrgUserRole.org_role_id.in_(list(role_map.keys())),
                        OrgUserRole.is_active == True)
                .all()
            )
            user_to_dept  = {}
            user_to_role  = {}
            for ur in user_role_rows:
                uid = str(ur.user_id)
                if uid not in user_to_role:
                    user_to_role[uid] = role_map.get(ur.org_role_id, "")
                    user_to_dept[uid] = dept_map.get(ur.department_id, "") if ur.department_id else ""

            for row_num, u in enumerate(users, start=2):
                uid = str(u.id)
                row_data = [
                    u.email or "",
                    u.firstname or "",
                    u.lastname or "",
                    u.phone_number or "",
                    u.employee_id or "",
                    user_to_dept.get(uid, ""),
                    user_to_role.get(uid, ""),
                ]
                for col_idx, val in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_num, column=col_idx, value=val)
                    cell.fill = existing_fill
                    cell.font = Font(name="Arial", size=10)
        else:
            # No users yet — show example row
            _write_example_row(ws, department_name, example_font)
    else:
        _write_example_row(ws, department_name, example_font)

    # ── Sheet 2: Available Roles ───────────────────────────────────────────
    ws_roles = wb.create_sheet("Available Roles")
    ws_roles.cell(row=1, column=1, value="Role Name (use exactly as shown)").fill = header_fill
    ws_roles.cell(row=1, column=1).font = header_font
    ws_roles.column_dimensions["A"].width = 38

    roles = db.query(OrgRole).filter(
        OrgRole.organization_id == org_id, OrgRole.is_active == True
    ).order_by(OrgRole.name).all()

    for row_idx, role in enumerate(roles, start=2):
        ws_roles.cell(row=row_idx, column=1, value=role.name)
    if not roles:
        ws_roles.cell(row=2, column=1, value="(No roles created yet)")

    # ── Stream ─────────────────────────────────────────────────────────────
    filename = "users_export.xlsx" if export_existing else "bulk_users_import_schema.xlsx"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def _write_example_row(ws, department_name, font):
    example = ["user@example.com", "John", "Doe", "+919876543210",
               "EMP001", department_name or "Bagalkot Zone", "Field Engineer"]
    for col_idx, value in enumerate(example, start=1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.font = font


@router.post("/bulk-import", response_model=BulkUserImportResponse)
def bulk_import_users(
    org_id: UUID,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_org_admin)
):
    """
    Bulk import users from an uploaded Excel file.
    Parses the 'Users' sheet, creates or updates users by email.
    Department matched by name; role matched by name or created.
    Passwords not updated for existing users; new users get 'Welcome@123'.
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only .xlsx or .xls files are accepted"
        )

    content = file.file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Could not read the uploaded file. Ensure it is a valid Excel file."
        )

    ws = wb.active  # reads first sheet regardless of name

    # Read headers from row 1
    headers = [
        (ws.cell(row=1, column=c).value or "").strip().lower()
        for c in range(1, ws.max_column + 1)
    ]

    required = {"email", "firstname", "phone_number", "department_name", "role_name"}
    missing = required - set(headers)
    if missing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Missing required columns: {', '.join(sorted(missing))}"
        )

    def _cell(row_values: dict, key: str) -> str:
        return str(row_values.get(key) or "").strip()

    rows: List[BulkUserImportRow] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
        email = _cell(row_dict, "email")
        if not email:
            continue  # skip blank rows
        rows.append(BulkUserImportRow(
            email=email,
            firstname=_cell(row_dict, "firstname"),
            lastname=_cell(row_dict, "lastname") or None,
            phone_number=_cell(row_dict, "phone_number"),
            employee_id=_cell(row_dict, "employee_id") or None,
            department_name=_cell(row_dict, "department_name"),
            role_name=_cell(row_dict, "role_name"),
        ))

    if not rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No data rows found in the file"
        )

    service = OrgUserService(db)
    return service.bulk_import_users(
        organization_id=org_id,
        rows=rows,
        created_by=current_user.id
    )


@router.post("/", response_model=OrgUserWithRoles, status_code=status.HTTP_201_CREATED)
def create_org_user(
    org_id: UUID,
    user_data: OrgUserCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_org_admin)
):
    """
    Create a new user within the organization.
    Only org admins can create users.
    Auto-assigns to specified department and roles if provided.
    """
    service = OrgUserService(db)
    user = service.create_org_user(
        organization_id=org_id,
        user_data=user_data,
        created_by=current_user.id
    )
    # Convert to OrgUserWithRoles
    users_with_roles = service.list_org_users_with_roles(
        organization_id=org_id,
        search=user.email
    )
    return users_with_roles[0] if users_with_roles else user


@router.get("/", response_model=List[OrgUserWithRoles])
def list_org_users(
    org_id: UUID,
    department_id: Optional[UUID] = None,
    skip: int = 0,
    limit: Optional[int] = None,
    is_active: Optional[bool] = None,
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_org_member)
):
    """
    List users in the organization with their roles.
    Optionally filter by department, active status, or search term.
    """
    service = OrgUserService(db)
    return service.list_org_users_with_roles(
        organization_id=org_id,
        department_id=department_id,
        skip=skip,
        limit=limit,
        is_active=is_active,
        search=search
    )


@router.get("/{user_id}", response_model=OrgUserWithRoles)
def get_org_user(
    org_id: UUID,
    user_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_org_member)
):
    """
    Get user details with roles.
    Any organization member can view other members.
    """
    service = OrgUserService(db)
    return service.get_org_user_with_roles(user_id, org_id)


@router.put("/{user_id}", response_model=OrgUserWithRoles)
def update_org_user(
    org_id: UUID,
    user_id: UUID,
    user_data: OrgUserUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_org_admin)
):
    """
    Update user in organization.
    Only org admins can update users.
    """
    service = OrgUserService(db)
    service.update_org_user(
        user_id=user_id,
        organization_id=org_id,
        user_data=user_data,
        modified_by=current_user.id
    )
    # Return user with roles
    return service.get_org_user_with_roles(user_id, org_id)


@router.delete("/{user_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_org_user(
    org_id: UUID,
    user_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_org_admin)
):
    """
    Soft delete user in organization.
    Only org admins can delete users.
    """
    # Prevent self-deletion
    if user_id == current_user.id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Cannot delete your own account"
        )

    service = OrgUserService(db)
    service.delete_org_user(user_id, org_id)
    return None


@router.post("/{user_id}/roles", response_model=OrgUserRoleOut, status_code=status.HTTP_201_CREATED)
def assign_role_to_user(
    org_id: UUID,
    user_id: UUID,
    role_assignment: RoleAssignment,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_org_admin)
):
    """
    Assign a role to a user within the organization.
    Only org admins can assign roles.
    """
    service = OrgUserService(db)
    return service.assign_role_to_user(
        user_id=user_id,
        organization_id=org_id,
        org_role_id=role_assignment.org_role_id,
        department_id=role_assignment.department_id,
        assigned_by=current_user.id
    )


@router.delete("/{user_id}/roles/{role_id}", status_code=status.HTTP_204_NO_CONTENT)
def remove_role_from_user(
    org_id: UUID,
    user_id: UUID,
    role_id: UUID,
    department_id: Optional[UUID] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_org_admin)
):
    """
    Remove a role from a user.
    Only org admins can remove roles.
    """
    service = OrgUserService(db)
    service.remove_role_from_user(
        user_id=user_id,
        organization_id=org_id,
        org_role_id=role_id,
        department_id=department_id
    )
    return None


@router.get("/{user_id}/roles", response_model=List[OrgUserRoleOut])
def get_user_roles(
    org_id: UUID,
    user_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_org_member)
):
    """
    Get all roles assigned to a user.
    Any organization member can view user roles.
    """
    service = OrgUserService(db)
    return service.get_user_roles(user_id, org_id)


@router.get("/me", response_model=OrgUserWithRoles)
def get_current_org_user(
    org_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_org_member)
):
    """
    Get current user's information with roles.
    """
    service = OrgUserService(db)
    return service.get_org_user_with_roles(current_user.id, org_id)
