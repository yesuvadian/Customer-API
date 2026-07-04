
import os
import json
import uuid as _uuid
from io import BytesIO
from typing import List, Optional
from uuid import UUID
from fastapi.responses import JSONResponse
from fastapi import Response
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, extract

from database import get_db
from auth_utils import get_current_user
from models import CategoryDetails, CategoryMaster, Equipment, EquipmentTypeKitMapping, Module, OrgDepartment, User
from middleware.org_auth import check_org_permission
from schemas import (
    EquipmentCreate,
    EquipmentUpdate,
    EquipmentResponse,
    EquipmentChainRef,
    EquipmentRetireRequest,
    EquipmentReplaceRequest,
    EquipmentCountResponse,
)
from services.equipment_service import EquipmentService

UPLOADS_DIR = os.path.join(os.path.dirname(__file__), "..", "uploads", "analysis_reports")
os.makedirs(UPLOADS_DIR, exist_ok=True)

NAMEPLATE_FILES_DIR = os.path.join(os.path.dirname(__file__), "..", "uploads", "nameplate_files")
os.makedirs(NAMEPLATE_FILES_DIR, exist_ok=True)

router = APIRouter(
    prefix="/equipment",
    tags=["equipment"],
    dependencies=[Depends(get_current_user)],
)


# ═══════════════════════════════════════════════════════════════════════════════
# MODULE-LEVEL HELPERS  (not routes)
# ═══════════════════════════════════════════════════════════════════════════════

def _dt(val) -> Optional[str]:
    """Safely convert a datetime (or date) value to an ISO-8601 string."""
    if val is None:
        return None
    return val.isoformat() if hasattr(val, "isoformat") else str(val)


def _get_equipment_module_id(db: Session) -> int:
    """Resolve the Equipment module row; raises 500 if not seeded."""
    mod = db.query(Module).filter_by(path="equipment", is_active=True).first()
    if not mod:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Equipment module not configured. Run seed first.",
        )
    return mod.id


def _require_permission(db: Session, user: User, action: str) -> None:
    """Check that *user* has *action* on the Equipment module."""
    from models import OrgUserRole, OrgRole

    is_org_admin = (
        db.query(OrgUserRole)
        .join(OrgRole)
        .filter(
            OrgUserRole.user_id == user.id,
            OrgRole.is_org_admin == True,
            OrgUserRole.is_active == True,
            OrgRole.is_active == True,
        )
        .first()
    )
    if is_org_admin:
        return

    module_id = _get_equipment_module_id(db)
    if not check_org_permission(user.id, module_id, action, db):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=f"Permission '{action}' required on Equipment module",
        )


def _enforce_org_scope(user: User) -> UUID:
    """Return the user's organization_id or raise 403."""
    if not user.organization_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User must belong to an organization to access equipment",
        )
    return user.organization_id


def _np_get(nameplate_data: dict, *keys: str) -> Optional[str]:
    """Extract a value from nameplate_data sections by matching field keys."""
    if not nameplate_data:
        return None
    for section in nameplate_data.get("sections", []):
        for field in section.get("fields", []):
            if field.get("key") in keys:
                v = field.get("value")
                if v not in (None, "", "null"):
                    return str(v)
    return None


def _chain_ref(eq_obj) -> Optional[dict]:
    """Lightweight replacement-chain reference dict from an Equipment ORM object."""
    if eq_obj is None:
        return None
    return {
        "id": str(eq_obj.id) if eq_obj.id else None,
        "ueic": eq_obj.ueic,
        "status": eq_obj.status.value if eq_obj.status else None,
        "manufacturer": eq_obj.manufacturer,
        "model_number": eq_obj.model_number,
        "commissioned_date": _dt(eq_obj.commissioned_date),
        "retired_date":      _dt(eq_obj.retired_date),
    }


def _to_response(db: Session, eq: Equipment) -> dict:
    """Convert Equipment ORM object to response dict with computed fields."""
    return {
        "id": str(eq.id) if eq.id else None,
        "ueic": eq.ueic,
        "organization_id": str(eq.organization_id) if eq.organization_id else None,
        "department_id": str(eq.department_id) if eq.department_id else None,
        "equipment_type_id": eq.equipment_type_id,
        "equipment_type_name": eq.equipment_type.name if eq.equipment_type else None,
        "department_name": eq.department.name if eq.department else None,
        "voltage_class": eq.voltage_class,
        "bay_number": eq.bay_number,
        "serial_in_bay": eq.serial_in_bay,
        "nameplate_data": json.loads(json.dumps(eq.nameplate_data, default=str))
        if eq.nameplate_data else {},
        "status": eq.status.value if eq.status else None,
        "replaces_equipment_id": str(eq.replaces_equipment_id) if eq.replaces_equipment_id else None,
        "replaces_equipment": _chain_ref(eq.replaces_equipment),
        "replaced_by_id": str(eq.replaced_by_id) if eq.replaced_by_id else None,
        "replaced_by": _chain_ref(
            eq.replaced_by_equipment[0] if eq.replaced_by_equipment else None
        ),
        "replacement_reason_type": eq.replacement_reason_type,
        "commissioned_date": _dt(eq.commissioned_date),
        "retired_date":      _dt(eq.retired_date),
        "retirement_reason": eq.retirement_reason,
        "manufacturer": eq.manufacturer or _np_get(eq.nameplate_data, "manufacturer", "make"),
        "model_number": eq.model_number or _np_get(eq.nameplate_data, "model_number", "model"),
        "factory_serial_number": eq.factory_serial_number or _np_get(eq.nameplate_data, "factory_serial_number", "serial_number"),
        "year_of_manufacture": eq.year_of_manufacture or _np_get(eq.nameplate_data, "year_of_manufacture", "year"),
        "latitude": float(eq.latitude) if eq.latitude is not None else None,
        "longitude": float(eq.longitude) if eq.longitude is not None else None,
        "phase": eq.phase,
        "ct_ratio_actual": eq.ct_ratio_actual,
        "ct_ratio_current": eq.ct_ratio_current,
        "pt_ratio": eq.pt_ratio,
        "vector_group": eq.vector_group,
        "impedance_pct": (
            float(eq.impedance_pct)
            if eq.impedance_pct is not None
            and str(eq.impedance_pct).lower() != "nan"
            else None
        ),
        "created_by": str(eq.created_by) if eq.created_by else None,
        "modified_by": str(eq.modified_by) if eq.modified_by else None,
        "cts": _dt(eq.cts),
        "mts": _dt(eq.mts),
        "types_by_category": _types_by_category_for_equipment(db, eq),
    }


def _types_by_category_for_equipment(db, eq) -> dict:
    """Return types_by_category for this equipment's type with lifecycle flags."""
    from models import CategoryDetails, OrgTestTemplate

    if not eq.equipment_type_id:
        return {"test": [], "maintenance": [], "inspection": [], "repair_lifecycle": []}

    all_types = (
        db.query(CategoryDetails)
        .filter(
            CategoryDetails.category_master_id == eq.equipment_type_id,
            CategoryDetails.is_active.is_(True),
        )
        .order_by(CategoryDetails.name)
        .all()
    )
    buckets: dict = {"test": [], "maintenance": [], "inspection": [], "repair_lifecycle": []}
    for t in all_types:
        cat = t.category_type or "test"
        bucket = buckets.get(cat, buckets["test"])
        tpl = (
            db.query(OrgTestTemplate)
            .filter(OrgTestTemplate.test_type_id == t.id)
            .order_by(OrgTestTemplate.version.desc())
            .first()
        )
        tpl_data = (tpl.template_data or {}) if tpl else {}
        bucket.append(
            {
                "id": t.id,
                "name": t.name,
                "category_type": t.category_type,
                "enable_cumulative": bool(tpl_data.get("enable_cumulative", False)),
                "enable_calibration": bool(tpl_data.get("enable_calibration", False)),
            }
        )
    return buckets


def _get_departments_at_depth(db: Session, org_id, depth: int) -> list:
    """Return all departments at a specific tree depth for the given org."""
    from sqlalchemy import text

    sql = text(
        """
        WITH RECURSIVE dept_tree AS (
            SELECT id, name, parent_department_id, 0 AS depth
            FROM org_departments
            WHERE organization_id = :org_id
              AND is_active = true
              AND parent_department_id IS NULL

            UNION ALL

            SELECT d.id, d.name, d.parent_department_id, dt.depth + 1
            FROM org_departments d
            INNER JOIN dept_tree dt ON d.parent_department_id = dt.id
            WHERE d.is_active = true
              AND dt.depth < :max_depth
        )
        SELECT dt.id, dt.name
        FROM dept_tree dt
        WHERE dt.depth = :depth ORDER BY dt.name
        """
    )
    rows = db.execute(
        sql, {"org_id": str(org_id), "depth": depth, "max_depth": depth}
    ).fetchall()
    return [{"id": str(r[0]), "name": r[1], "aliases": []} for r in rows]


def _resolve_nameplate_file_field(db: Session, equipment: Equipment, field_key: str) -> dict:
    """Look up the template for this equipment type and return the field definition."""
    from models import CategoryDetails, OrgTestTemplate

    detail = (
        db.query(CategoryDetails)
        .filter(
            CategoryDetails.category_master_id == equipment.equipment_type_id,
            CategoryDetails.category_type == "nameplate",
        )
        .first()
    )
    if not detail:
        raise HTTPException(
            status_code=404,
            detail="No nameplate template found for this equipment type.",
        )

    tmpl = (
        db.query(OrgTestTemplate)
        .filter(
            OrgTestTemplate.test_type_id == detail.id,
            OrgTestTemplate.org_id == equipment.organization_id,
        )
        .first()
    ) or (
        db.query(OrgTestTemplate)
        .filter(
            OrgTestTemplate.test_type_id == detail.id,
            OrgTestTemplate.org_id == None,
        )
        .first()
    )
    if not tmpl:
        raise HTTPException(status_code=404, detail="Nameplate template not provisioned.")

    for section in (tmpl.template_data or {}).get("sections", []):
        for field in section.get("fields", []):
            if field.get("key") == field_key:
                if field.get("type") != "file":
                    raise HTTPException(
                        status_code=400,
                        detail=f"Field '{field_key}' is not a file-upload field in the template.",
                    )
                return field

    raise HTTPException(
        status_code=404,
        detail=f"Field '{field_key}' not found in nameplate template.",
    )


# ═══════════════════════════════════════════════════════════════════════════════
# BULK IMPORT — template download, validate, import
# ═══════════════════════════════════════════════════════════════════════════════

_FIXED_BULK_COLS = [
    ("bay_number",            "Bay Number",              "e.g. 04"),
    ("voltage_class",         "Voltage Class (kV)",      "400 / 220 / 110 / 66 / 33 / 11"),
    ("manufacturer",          "Manufacturer",            "e.g. ABB"),
    ("model_number",          "Model Number",            "e.g. TRAFO-X"),
    ("factory_serial_number", "Factory Serial Number",   "e.g. SN-2024-001"),
    ("year_of_manufacture",   "Year of Manufacture",     "e.g. 2015"),
    ("commissioned_date",     "Commissioned Date",       "YYYY-MM-DD"),
    ("phase",                 "Phase",                   "R / Y / B (leave blank for transformers)"),
    ("ct_ratio_actual",       "CT Ratio (Nameplate)",    "e.g. 800-400/1-1A"),
    ("ct_ratio_current",      "CT Ratio (Active Tap)",   "e.g. 800/1A"),
    ("pt_ratio",              "PT Ratio",                "e.g. 220kV/110V"),
    ("vector_group",          "Vector Group",            "e.g. YNyn0d11"),
    ("impedance_pct",         "Impedance (%)",           "e.g. 9.8"),
    ("latitude",              "Latitude",                "e.g. 12.9716"),
    ("longitude",             "Longitude",               "e.g. 77.5946"),
    ("scada_tag",             "SCADA Tag",               "e.g. SS_TR1_OIL_TEMP"),
]

_SKIP_FIELD_TYPES = {"readonly", "file", "table"}


def _get_nameplate_fields(db: Session, equipment_type_id: int, org_id) -> list:
    """Return editable nameplate template fields for a given equipment type."""
    from models import CategoryDetails, OrgTestTemplate
    detail = (
        db.query(CategoryDetails)
        .filter(
            CategoryDetails.category_master_id == equipment_type_id,
            CategoryDetails.category_type == "nameplate",
        )
        .first()
    )
    if not detail:
        return []
    tmpl = (
        db.query(OrgTestTemplate)
        .filter(OrgTestTemplate.test_type_id == detail.id,
                OrgTestTemplate.org_id == org_id)
        .first()
    ) or (
        db.query(OrgTestTemplate)
        .filter(OrgTestTemplate.test_type_id == detail.id,
                OrgTestTemplate.org_id.is_(None))
        .first()
    )
    if not tmpl:
        return []
    fields = []
    for section in (tmpl.template_data or {}).get("sections", []):
        for f in section.get("fields", []):
            if f.get("type") in _SKIP_FIELD_TYPES:
                continue
            fields.append(f)
    return fields


@router.get("/bulk-template")
def download_bulk_template(
    department_id: UUID = Query(...),
    equipment_type_id: int = Query(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Generate and return an Excel template pre-filled with the department and
    nameplate columns for the given equipment type.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Protection
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    org_id = _enforce_org_scope(current_user)

    dept = db.query(OrgDepartment).filter(OrgDepartment.id == department_id).first()
    if not dept:
        raise HTTPException(status_code=404, detail="Department not found")

    equip_type = db.query(CategoryMaster).filter(CategoryMaster.id == equipment_type_id).first()
    if not equip_type:
        raise HTTPException(status_code=404, detail="Equipment type not found")

    nameplate_fields = _get_nameplate_fields(db, equipment_type_id, org_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Equipment Import"

    # ── Styles ────────────────────────────────────────────────────────────────
    hdr_fill   = PatternFill("solid", fgColor="1E3A8A")
    hdr_font   = Font(bold=True, color="FFFFFF", size=11)
    locked_fill= PatternFill("solid", fgColor="E5E7EB")
    hint_fill  = PatternFill("solid", fgColor="F9FAFB")
    hint_font  = Font(italic=True, color="9CA3AF", size=10)
    np_fill    = PatternFill("solid", fgColor="EFF6FF")
    np_hdr_font= Font(bold=True, color="1D4ED8", size=11)
    center     = Alignment(horizontal="center", vertical="center")
    wrap       = Alignment(wrap_text=True, vertical="center")

    # ── Build column list ─────────────────────────────────────────────────────
    # Hidden metadata cols (col A, B) — not visible to user
    meta_cols = [
        ("__department_name__", dept.name),
        ("__equipment_type_id__", str(equipment_type_id)),
    ]
    # Locked display col (col C)
    display_cols = [("department", dept.name)]
    fixed_cols   = _FIXED_BULK_COLS
    np_cols      = [(f["key"], f.get("label", f["key"]), f.get("type", "text"), f.get("options", []))
                    for f in nameplate_fields]

    all_cols = meta_cols + display_cols + [(k, lbl, "") for k, lbl, _ in fixed_cols] + \
               [(k, lbl, t, opts) for k, lbl, t, opts in np_cols]

    total_cols = len(all_cols)

    # ── Row 1: Title banner ───────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1,
        value=f"Equipment Bulk Import — {dept.name} — {equip_type.name}")
    title_cell.font  = Font(bold=True, color="FFFFFF", size=13)
    title_cell.fill  = PatternFill("solid", fgColor="0F172A")
    title_cell.alignment = center
    ws.row_dimensions[1].height = 28

    # ── Row 2: Machine-readable keys (parsed by bulk import) ─────────────────
    # Row 3: Human-readable labels shown to user
    ws.row_dimensions[2].height = 0   # hidden — keys only, not for user
    ws.row_dimensions[3].height = 22  # visible label row
    for ci, col_def in enumerate(all_cols, start=1):
        key  = col_def[0]
        lbl  = col_def[1]
        is_meta = key.startswith("__")
        is_np   = ci > (len(meta_cols) + len(display_cols) + len(fixed_cols))

        # Row 2: always write the field key (machine-readable)
        key_cell = ws.cell(row=2, column=ci, value=key)
        key_cell.font      = Font(color="FFFFFF", size=1)  # invisible text
        key_cell.fill      = locked_fill if is_meta else (np_fill if is_np else hdr_fill)
        key_cell.alignment = center

        # Row 3: write human-readable label as the visible header
        lbl_cell = ws.cell(row=3, column=ci, value=lbl if not is_meta else "")
        lbl_cell.font      = np_hdr_font if is_np else hdr_font
        lbl_cell.fill      = np_fill if is_np else (locked_fill if is_meta else hdr_fill)
        lbl_cell.alignment = center

    # ── Row 4: Hint/example row ────────────────────────────────────────────────
    ws.row_dimensions[4].height = 18
    hint_values = (
        [dept.name, str(equipment_type_id), dept.name]
        + [hint for _, _, hint in fixed_cols]
        + [f.get("placeholder", f"Enter {f.get('label',f['key'])}") for f in nameplate_fields
           if f.get("type") not in _SKIP_FIELD_TYPES]
    )
    for ci, val in enumerate(hint_values, start=1):
        cell = ws.cell(row=4, column=ci, value=val)
        cell.font      = hint_font
        cell.fill      = hint_fill
        cell.alignment = wrap

    # ── Data validation: voltage_class dropdown ────────────────────────────────
    vc_col_idx = next(
        (ci for ci, c in enumerate(all_cols, 1) if c[0] == "voltage_class"), None
    )
    if vc_col_idx:
        vc_letter = get_column_letter(vc_col_idx)
        dv = DataValidation(
            type="list",
            formula1='"400,220,110,66,33,11"',
            allow_blank=True,
        )
        dv.sqref = f"{vc_letter}5:{vc_letter}1000"
        ws.add_data_validation(dv)

    # ── Data validation: phase dropdown ───────────────────────────────────────
    phase_col_idx = next(
        (ci for ci, c in enumerate(all_cols, 1) if c[0] == "phase"), None
    )
    if phase_col_idx:
        ph_letter = get_column_letter(phase_col_idx)
        dv2 = DataValidation(type="list", formula1='"R,Y,B"', allow_blank=True)
        dv2.sqref = f"{ph_letter}5:{ph_letter}1000"
        ws.add_data_validation(dv2)

    # ── Data validation: nameplate dropdowns ──────────────────────────────────
    np_start_ci = len(meta_cols) + len(display_cols) + len(fixed_cols) + 1
    for rel_i, (key, lbl, ftype, opts) in enumerate(np_cols):
        if ftype == "dropdown" and opts:
            col_letter = get_column_letter(np_start_ci + rel_i)
            formula    = '"' + ",".join(str(o) for o in opts[:30]) + '"'
            dv3 = DataValidation(type="list", formula1=formula, allow_blank=True)
            dv3.sqref = f"{col_letter}5:{col_letter}1000"
            ws.add_data_validation(dv3)

    # ── Column widths & hide metadata cols ────────────────────────────────────
    for ci, col_def in enumerate(all_cols, start=1):
        key = col_def[0]
        col_letter = get_column_letter(ci)
        if key.startswith("__"):
            ws.column_dimensions[col_letter].hidden = True
            ws.column_dimensions[col_letter].width  = 0
        elif key == "department":
            ws.column_dimensions[col_letter].width = 28
        else:
            ws.column_dimensions[col_letter].width = 22

    # ── Freeze panes at row 4, col D (first editable data col) ───────────────
    first_data_col = get_column_letter(len(meta_cols) + len(display_cols) + 1)
    ws.freeze_panes = f"{first_data_col}4"

    # ── Output ────────────────────────────────────────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"equipment_import_{dept.name.replace(' ', '_')}_{equip_type.name.replace(' ', '_')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _parse_bulk_excel(contents: bytes) -> tuple[list, list]:
    """
    Parse the bulk import Excel.
    Returns (meta, rows) where meta = {department_id, equipment_type_id}
    and rows = list of dicts with field values.
    """
    import openpyxl
    wb = openpyxl.load_workbook(filename=BytesIO(contents), data_only=True)
    ws = wb.active

    # Row 2 = machine keys (hidden), row 3 = labels, row 4 = hints, row 5+ = data
    headers = [ws.cell(row=2, column=ci).value for ci in range(1, ws.max_column + 1)]

    # Extract metadata from hidden cols (A, B) — values stored in hint row 4
    meta = {}
    for ci, h in enumerate(headers, start=1):
        if h == "__department_name__":
            meta["department_name"] = ws.cell(row=4, column=ci).value
        elif h == "__department_id__":
            meta["department_id"] = ws.cell(row=4, column=ci).value  # legacy: UUID stored here
        elif h == "department":
            meta["department_display"] = ws.cell(row=4, column=ci).value  # display name col
        elif h == "__equipment_type_id__":
            meta["equipment_type_id"] = ws.cell(row=4, column=ci).value

    from datetime import datetime as _dt_cls, date as _date_cls

    def _coerce(v):
        """Convert Excel cell values to JSON-safe types."""
        if isinstance(v, (_dt_cls, _date_cls)):
            return v.isoformat()
        return v

    rows = []
    for row_idx in range(4, ws.max_row + 1):
        row_vals = {h: _coerce(ws.cell(row=row_idx, column=ci).value)
                    for ci, h in enumerate(headers, start=1)
                    if h and not str(h).startswith("__")}
        # Skip completely empty rows
        if all(v is None or str(v).strip() == "" for v in row_vals.values()):
            continue
        row_vals["__row__"] = row_idx
        rows.append(row_vals)

    return meta, rows


def _validate_bulk_rows(rows: list, meta: dict, db: Session) -> list:
    """Validate each row and return list of result dicts."""
    import re
    from datetime import datetime

    valid_phase   = {"R", "Y", "B"}

    results = []
    for row in rows:
        errors = []
        row_num = row.get("__row__", "?")

        # bay_number — no validation needed (free text)

        # phase
        ph = row.get("phase")
        if ph and str(ph).strip().upper() not in valid_phase:
            errors.append(f"phase '{ph}' invalid — use: R/Y/B")
        if ph:
            row["phase"] = str(ph).strip().upper()

        # year_of_manufacture
        yom = row.get("year_of_manufacture")
        if yom is not None and str(yom).strip():
            try:
                yom_int = int(float(str(yom)))
                if not (1900 <= yom_int <= datetime.now().year + 1):
                    errors.append(f"year_of_manufacture '{yom}' out of range")
                else:
                    row["year_of_manufacture"] = yom_int
            except ValueError:
                errors.append(f"year_of_manufacture '{yom}' is not a valid year")

        # commissioned_date
        cd = row.get("commissioned_date")
        if cd and str(cd).strip():
            parsed_date = None
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
                try:
                    parsed_date = datetime.strptime(str(cd).strip(), fmt)
                    break
                except ValueError:
                    continue
            if not parsed_date and hasattr(cd, "year"):
                parsed_date = cd  # openpyxl already parsed it as datetime
            if parsed_date:
                row["commissioned_date"] = parsed_date.isoformat()
            else:
                errors.append(f"commissioned_date '{cd}' not parseable — use YYYY-MM-DD")

        # latitude / longitude
        for coord in ("latitude", "longitude"):
            val = row.get(coord)
            if val is not None and str(val).strip():
                try:
                    row[coord] = float(val)
                except (ValueError, TypeError):
                    errors.append(f"{coord} '{val}' is not a valid number")

        # impedance_pct
        imp = row.get("impedance_pct")
        if imp is not None and str(imp).strip():
            try:
                row["impedance_pct"] = float(imp)
            except (ValueError, TypeError):
                errors.append(f"impedance_pct '{imp}' is not a valid number")

        results.append({
            "row": row_num,
            "status": "error" if errors else "valid",
            "errors": errors,
            "data": {k: v for k, v in row.items() if not k.startswith("__")},
        })

    return results


@router.post("/bulk-validate")
async def bulk_validate(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Upload a filled bulk-import Excel and get per-row validation results.
    Returns {department_id, equipment_type_id, rows: [{row, status, errors, data}]}
    """
    _enforce_org_scope(current_user)
    contents = await file.read()
    try:
        meta, rows = _parse_bulk_excel(contents)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Could not parse Excel: {exc}")

    if not rows:
        raise HTTPException(status_code=400, detail="No data rows found in the file. Fill in equipment data starting from row 4.")

    results = _validate_bulk_rows(rows, meta, db)
    valid_count = sum(1 for r in results if r["status"] == "valid")
    error_count = len(results) - valid_count

    return {
        "department_name": meta.get("department_name"),
        "equipment_type_id": meta.get("equipment_type_id"),
        "total": len(results),
        "valid": valid_count,
        "errors": error_count,
        "rows": results,
    }


@router.post("/bulk-import")
async def bulk_import(
    file: UploadFile = File(...),
    skip_errors: bool = Query(default=True, description="Skip invalid rows and import valid ones"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Upload the filled Excel and import all valid rows as new equipment.
    Returns {imported, skipped, errors: [{row, errors}]}
    """
    org_id = _enforce_org_scope(current_user)
    _require_permission(db, current_user, "can_add")

    contents = await file.read()
    try:
        meta, rows = _parse_bulk_excel(contents)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Could not parse Excel: {exc}")

    if not rows:
        raise HTTPException(status_code=400, detail="No data rows found in the file")

    # Resolve equipment_type_id from metadata
    try:
        eq_type = int(str(meta.get("equipment_type_id", "")).strip())
    except (ValueError, TypeError):
        raise HTTPException(
            status_code=400,
            detail="Could not read equipment type from template. Please use the template downloaded from this system.",
        )

    # Resolve department — new templates use name, legacy templates use UUID
    dept = None
    dept_name = str(meta.get("department_name", "")).strip()
    if dept_name:
        dept = db.query(OrgDepartment).filter(
            OrgDepartment.organization_id == org_id,
            OrgDepartment.name == dept_name,
        ).first()
    if not dept and meta.get("department_id"):
        # Legacy template: try UUID lookup first
        try:
            dept = db.query(OrgDepartment).filter(
                OrgDepartment.id == UUID(str(meta["department_id"]).strip())
            ).first()
            if dept:
                dept_name = dept.name
        except (ValueError, TypeError):
            pass
    if not dept:
        # UUID missing or stale — resolve by display name embedded in template
        display_name = str(meta.get("department_display") or meta.get("department_name") or "").strip()
        if display_name:
            dept = db.query(OrgDepartment).filter(
                OrgDepartment.organization_id == org_id,
                OrgDepartment.name == display_name,
            ).first()
            if dept:
                dept_name = dept.name
    if not dept:
        raise HTTPException(
            status_code=400,
            detail="Department not found. Please download a fresh template from the Equipment Register page.",
        )

    validated = _validate_bulk_rows(rows, meta, db)

    imported   = 0
    skipped    = 0
    error_rows = []

    for result in validated:
        if result["status"] == "error":
            if skip_errors:
                skipped += 1
                error_rows.append({"row": result["row"], "errors": result["errors"]})
                continue
            else:
                raise HTTPException(
                    status_code=422,
                    detail=f"Row {result['row']} invalid: {'; '.join(result['errors'])}",
                )

        d = result["data"]

        # Separate fixed fields from nameplate data
        fixed_keys = {k for k, _, _ in _FIXED_BULK_COLS} | {"department"}
        nameplate_data = {k: v for k, v in d.items()
                          if k not in fixed_keys and v is not None and str(v).strip() != ""}

        from datetime import datetime as _dt
        cd_raw = d.get("commissioned_date")
        commissioned_date = None
        if cd_raw:
            try:
                commissioned_date = _dt.fromisoformat(str(cd_raw))
            except ValueError:
                pass

        try:
            equipment = EquipmentService.create_equipment(
                db=db,
                organization_id=org_id,
                department_id=dept.id,
                equipment_type_id=eq_type,
                voltage_class=str(d["voltage_class"]).strip() if d.get("voltage_class") else None,
                bay_number=str(d["bay_number"]).strip() if d.get("bay_number") else None,
                nameplate_data=nameplate_data or None,
                commissioned_date=commissioned_date,
                manufacturer=str(d["manufacturer"]).strip() if d.get("manufacturer") else None,
                model_number=str(d["model_number"]).strip() if d.get("model_number") else None,
                factory_serial_number=str(d["factory_serial_number"]).strip() if d.get("factory_serial_number") else None,
                year_of_manufacture=d.get("year_of_manufacture"),
                latitude=d.get("latitude"),
                longitude=d.get("longitude"),
                phase=d.get("phase"),
                ct_ratio_actual=str(d["ct_ratio_actual"]).strip() if d.get("ct_ratio_actual") else None,
                ct_ratio_current=str(d["ct_ratio_current"]).strip() if d.get("ct_ratio_current") else None,
                pt_ratio=str(d["pt_ratio"]).strip() if d.get("pt_ratio") else None,
                vector_group=str(d["vector_group"]).strip() if d.get("vector_group") else None,
                impedance_pct=d.get("impedance_pct"),
                scada_tag=str(d["scada_tag"]).strip() if d.get("scada_tag") else None,
                created_by=current_user.id,
            )
            db.flush()
            imported += 1

            # Non-fatal post-creation hooks
            try:
                from services.test_request_schedule_service import TestRequestScheduleService
                TestRequestScheduleService.instantiate_equipment_schedules(db, equipment, current_user.id)
            except Exception:
                pass

        except Exception as exc:
            db.rollback()
            skipped += 1
            error_rows.append({"row": result["row"], "errors": [str(exc)]})
            continue

    db.commit()

    return {
        "imported": imported,
        "skipped": skipped,
        "error_rows": error_rows,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# ① STATIC ROUTES — no UUID path segment
# ═══════════════════════════════════════════════════════════════════════════════

@router.post("/", response_model=EquipmentResponse, status_code=status.HTTP_201_CREATED)
def create_equipment(
    data: EquipmentCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Register a new equipment unit. UEIC is auto-generated."""
    org_id = _enforce_org_scope(current_user)
    _require_permission(db, current_user, "can_add")

    if data.organization_id and data.organization_id != org_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Cannot create equipment in another organization",
        )

    equipment = EquipmentService.create_equipment(
        db=db,
        organization_id=org_id,
        department_id=data.department_id,
        equipment_type_id=data.equipment_type_id,
        voltage_class=data.voltage_class,
        bay_number=data.bay_number,
        nameplate_data=data.nameplate_data,
        commissioned_date=data.commissioned_date,
        manufacturer=data.manufacturer,
        model_number=data.model_number,
        factory_serial_number=data.factory_serial_number,
        year_of_manufacture=data.year_of_manufacture,
        latitude=data.latitude,
        longitude=data.longitude,
        phase=data.phase,
        ct_ratio_actual=data.ct_ratio_actual,
        ct_ratio_current=data.ct_ratio_current,
        pt_ratio=data.pt_ratio,
        vector_group=data.vector_group,
        impedance_pct=data.impedance_pct,
        scada_tag=data.scada_tag,
        created_by=current_user.id,
    )
    db.commit()
    db.refresh(equipment)

    if data.precommission_request_id:
        try:
            from services.precommission_service import PreCommissionService
            PreCommissionService(db).link_equipment(
                data.precommission_request_id, equipment.id, current_user
            )
            equipment.precommission_request_id = data.precommission_request_id
            db.flush()
        except Exception as _pcr_exc:
            print(f"[WARN] PCR link failed (non-fatal): {_pcr_exc}")

    try:
        from services.test_request_schedule_service import TestRequestScheduleService
        TestRequestScheduleService.instantiate_equipment_schedules(db, equipment, current_user.id)
    except Exception as exc:
        import traceback
        traceback.print_exc()
        print(f"[WARN] instantiate_equipment_schedules failed: {exc}")

    try:
        from services.notification_service import NotificationService
        commissioned_by = (
            f"{current_user.firstname or ''} {current_user.lastname or ''}".strip()
            or current_user.email
        )
        NotificationService(db).notify_equipment_registered(
            equipment,
            commissioned_by=commissioned_by,
            organization_id=org_id,
            department_id=equipment.department_id,
        )
    except Exception as _n:
        print(f"[WARN] equipment_registered notification failed: {_n}")

    return _to_response(db, equipment)


@router.get("/", response_model=List[EquipmentResponse])
def list_equipment(
    # Basic filters
    department_id: Optional[UUID] = None,
    equipment_type_id: Optional[int] = None,
    status: Optional[str] = None,
    voltage_class: Optional[str] = None,
    manufacturer: Optional[str] = None,
    model_number: Optional[str] = None,
    search: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
    # Area filters
    tlss_division: Optional[str] = None,
    wm_circle: Optional[str] = None,
    transmission_zone: Optional[str] = None,
    substation_ids: Optional[str] = None,
    # Year filters
    commission_year: Optional[int] = None,
    commission_year_from: Optional[int] = None,
    commission_year_to: Optional[int] = None,
    failure_year: Optional[int] = None,
    failure_year_from: Optional[int] = None,
    failure_year_to: Optional[int] = None,
    replacement_year: Optional[int] = None,
    replacement_year_from: Optional[int] = None,
    replacement_year_to: Optional[int] = None,
    has_tests: bool = False,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """List equipment with optional filters. Scoped to user's organization."""
    org_id = _enforce_org_scope(current_user)
    _require_permission(db, current_user, "can_view")

    from utils.common_service import get_user_dept_scope, get_dept_subtree_ids
    substation_id_list: Optional[list] = None
    if department_id is None:
        is_admin, scoped_dept = get_user_dept_scope(db, current_user.id, org_id)
        if not is_admin and scoped_dept:
            department_id = scoped_dept

    items = EquipmentService.list_equipment(
        db=db,
        organization_id=org_id,
        department_id=department_id,
        equipment_type_id=equipment_type_id,
        status=status,
        voltage_class=voltage_class,
        manufacturer=manufacturer,
        model_number=model_number,
        substation_ids=substation_ids,
        tlss_division=tlss_division,
        wm_circle=wm_circle,
        transmission_zone=transmission_zone,
        search=search,
        skip=skip,
        limit=limit,
        commission_year=commission_year,
        commission_year_from=commission_year_from,
        commission_year_to=commission_year_to,
        failure_year=failure_year,
        failure_year_from=failure_year_from,
        failure_year_to=failure_year_to,
        replacement_year=replacement_year,
        replacement_year_from=replacement_year_from,
        replacement_year_to=replacement_year_to,
        has_tests=has_tests,
    )
    responses = []

    for i, eq in enumerate(items):
        try:
            responses.append(_to_response(db, eq))
        except Exception as e:
            print(f"\nBROKEN EQUIPMENT INDEX: {i}")
            print(f"BROKEN EQUIPMENT ID: {eq.id}")
            print(f"BROKEN UEIC: {eq.ueic}")
            print(f"ERROR: {e}\n")
            raise

    from fastapi.encoders import jsonable_encoder

    import json

    try:
        json.dumps(responses, default=str)
    except Exception as e:
        print("\nJSON SERIALIZATION FAILED")
        print(e)
        print(type(e))
        raise

    safe = []

    for i, item in enumerate(responses):
        try:
            json.dumps(item, default=str)
            safe.append(item)
        except Exception as e:
            print(f"\nBROKEN RESPONSE INDEX: {i}")
            print(f"BROKEN RESPONSE DATA: {item}")
            print(f"ERROR: {e}\n")

    return Response(
        content=json.dumps(safe, default=str),
        media_type="application/json"
    )


@router.get("/export/csv")
def export_equipment_csv(
    department_id: Optional[UUID] = None,
    equipment_type_id: Optional[int] = None,
    status: Optional[str] = None,
    voltage_class: Optional[str] = None,
    manufacturer: Optional[str] = None,
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export equipment list as UTF-8 CSV."""
    import csv
    import io
    from datetime import datetime

    org_id = _enforce_org_scope(current_user)
    _require_permission(db, current_user, "can_export")

    items = EquipmentService.list_equipment(
        db=db,
        organization_id=org_id,
        department_id=department_id,
        equipment_type_id=equipment_type_id,
        status=status,
        voltage_class=voltage_class,
        manufacturer=manufacturer,
        search=search,
        skip=0,
        limit=10_000,
    )

    columns = [
        ("UEIC",                    lambda eq: eq.ueic or ""),
        ("Equipment Type",          lambda eq: eq.equipment_type.name if eq.equipment_type else ""),
        ("Department",              lambda eq: eq.department.name if eq.department else ""),
        ("Status",                  lambda eq: eq.status.value if eq.status else ""),
        ("Voltage Class (kV)",      lambda eq: eq.voltage_class or ""),
        ("Bay Number",              lambda eq: eq.bay_number or ""),
        ("Manufacturer",            lambda eq: eq.manufacturer or ""),
        ("Model Number",            lambda eq: eq.model_number or ""),
        ("Factory Serial Number",   lambda eq: eq.factory_serial_number or ""),
        ("Year of Manufacture",     lambda eq: str(eq.year_of_manufacture) if eq.year_of_manufacture else ""),
        ("Commissioned Date",       lambda eq: _dt(eq.commissioned_date) or ""),
        ("Retired Date",            lambda eq: _dt(eq.retired_date) or ""),
        ("Retirement Reason",       lambda eq: eq.retirement_reason or ""),
        ("Created",                 lambda eq: _dt(eq.cts) or ""),
        ("Last Modified",           lambda eq: _dt(eq.mts) or ""),
    ]

    output = io.StringIO()
    writer = csv.writer(output, lineterminator="\r\n")
    writer.writerow([col[0] for col in columns])
    for eq in items:
        writer.writerow([col[1](eq) for col in columns])

    csv_bytes = output.getvalue().encode("utf-8-sig")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"equipment_export_{timestamp}.csv"

    return StreamingResponse(
        iter([csv_bytes]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ═══════════════════════════════════════════════════════════════════════════════
# ✅ FIXED: STATS / COUNTS with all filters
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/stats/counts", response_model=EquipmentCountResponse)
def get_equipment_counts(
    # Basic filters
    department_id: Optional[UUID] = None,
    equipment_type_id: Optional[int] = None,
    status: Optional[str] = None,
    voltage_class: Optional[str] = None,
    manufacturer: Optional[str] = None,
    model_number: Optional[str] = None,
    search: Optional[str] = None,
    # Area filters
    tlss_division: Optional[str] = None,
    wm_circle: Optional[str] = None,
    transmission_zone: Optional[str] = None,
    substation_ids: Optional[str] = None,
    # Year filters
    commission_year: Optional[int] = None,
    commission_year_from: Optional[int] = None,
    commission_year_to: Optional[int] = None,
    failure_year: Optional[int] = None,
    failure_year_from: Optional[int] = None,
    failure_year_to: Optional[int] = None,
    replacement_year: Optional[int] = None,
    replacement_year_from: Optional[int] = None,
    replacement_year_to: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Equipment counts by status — scoped to user's organization and optional filters.
    This now supports ALL the same filters as the list endpoint.
    """
    org_id = _enforce_org_scope(current_user)
    _require_permission(db, current_user, "can_view")
    
    # Start with base query
    query = db.query(Equipment.status, func.count(Equipment.id))
    
    # Organization scope
    query = query.filter(Equipment.organization_id == org_id)

    # Auto-scope to user's dept if no explicit filter provided
    if department_id is None:
        from utils.common_service import get_user_dept_scope
        is_admin, scoped_dept = get_user_dept_scope(db, current_user.id, org_id)
        if not is_admin and scoped_dept:
            department_id = scoped_dept

    # ── Department / substation filter ──────────────────────────────────────
    if department_id:
        from services.equipment_service import EquipmentService as ES
        dept_ids = ES._get_department_subtree_ids(db, department_id)
        query = query.filter(Equipment.department_id.in_(dept_ids))
    
    # ── Equipment type filter ──────────────────────────────────────────────
    if equipment_type_id:
        query = query.filter(Equipment.equipment_type_id == equipment_type_id)
    
    # ── Status filter ──────────────────────────────────────────────────────
    if status:
        query = query.filter(Equipment.status == status)
    
    # ── Voltage class filter ──────────────────────────────────────────────
    if voltage_class:
        query = query.filter(Equipment.voltage_class == voltage_class)
    
    # ── Manufacturer filter ────────────────────────────────────────────────
    if manufacturer:
        query = query.filter(Equipment.manufacturer.ilike(f"%{manufacturer}%"))
    
    # ── Model number filter ────────────────────────────────────────────────
    if model_number:
        query = query.filter(Equipment.model_number.ilike(f"%{model_number}%"))
    
    # ── Search filter ──────────────────────────────────────────────────────
    if search:
        query = query.filter(
            (Equipment.ueic.ilike(f"%{search}%")) |
            (Equipment.bay_number.ilike(f"%{search}%")) |
            (Equipment.manufacturer.ilike(f"%{search}%")) |
            (Equipment.model_number.ilike(f"%{search}%")) |
            (Equipment.factory_serial_number.ilike(f"%{search}%"))
        )
    
    # ── Substation IDs filter ─────────────────────────────────────────────
    if substation_ids:
        id_list = [s.strip() for s in substation_ids.split(",") if s.strip()]
        if id_list:
            try:
                parsed = [UUID(i) for i in id_list]
                query = query.filter(Equipment.department_id.in_(parsed))
            except ValueError:
                pass
    
    # ── Area filters (zone, circle, division) ─────────────────────────────
    from services.equipment_service import EquipmentService as ES
    
    if transmission_zone:
        dept_ids = ES._get_descendants_of_named(db, org_id, transmission_zone)
        if not dept_ids:
            return {"total": 0, "active": 0, "under_repair": 0, "retired": 0, "scrapped": 0}
        query = query.filter(Equipment.department_id.in_(dept_ids))
    
    if wm_circle:
        dept_ids = ES._get_descendants_of_named(db, org_id, wm_circle)
        if not dept_ids:
            return {"total": 0, "active": 0, "under_repair": 0, "retired": 0, "scrapped": 0}
        query = query.filter(Equipment.department_id.in_(dept_ids))
    
    if tlss_division:
        dept_ids = ES._get_descendants_of_named(db, org_id, tlss_division)
        if not dept_ids:
            return {"total": 0, "active": 0, "under_repair": 0, "retired": 0, "scrapped": 0}
        query = query.filter(Equipment.department_id.in_(dept_ids))
    
    # ── Commission year filters ───────────────────────────────────────────
    if commission_year:
        query = query.filter(
            extract('year', Equipment.commissioned_date) == commission_year
        )
    if commission_year_from:
        query = query.filter(
            extract('year', Equipment.commissioned_date) >= commission_year_from
        )
    if commission_year_to:
        query = query.filter(
            extract('year', Equipment.commissioned_date) <= commission_year_to
        )
    
    # ── Failure year filters ──────────────────────────────────────────────
    if failure_year:
        query = query.filter(
            extract('year', Equipment.retired_date) == failure_year
        )
    if failure_year_from:
        query = query.filter(
            extract('year', Equipment.retired_date) >= failure_year_from
        )
    if failure_year_to:
        query = query.filter(
            extract('year', Equipment.retired_date) <= failure_year_to
        )
    
    # ── Replacement year filters ──────────────────────────────────────────
    if replacement_year:
        query = query.filter(
            Equipment.replaces_equipment_id.isnot(None),
            extract('year', Equipment.commissioned_date) == replacement_year
        )
    if replacement_year_from:
        query = query.filter(
            Equipment.replaces_equipment_id.isnot(None),
            extract('year', Equipment.commissioned_date) >= replacement_year_from
        )
    if replacement_year_to:
        query = query.filter(
            Equipment.replaces_equipment_id.isnot(None),
            extract('year', Equipment.commissioned_date) <= replacement_year_to
        )
    
    # ── Execute and build response ────────────────────────────────────────
    rows = query.group_by(Equipment.status).all()
    from models import EquipmentStatus
    counts = {s.value: 0 for s in EquipmentStatus}
    for s, c in rows:
        counts[s.value] = c
    counts["total"] = sum(counts.values())
    return counts


@router.get("/stats/group-counts")
def get_equipment_group_counts(
    group_by: str = Query(..., description="Field to group by: equipment_type, manufacturer, voltage_class, model_type, bay_number, commission_year, failure_year, replacement_year"),
    department_id: Optional[UUID] = None,
    equipment_type_id: Optional[int] = None,
    status: Optional[str] = None,
    voltage_class: Optional[str] = None,
    manufacturer: Optional[str] = None,
    model_number: Optional[str] = None,
    search: Optional[str] = None,
    tlss_division: Optional[str] = None,
    wm_circle: Optional[str] = None,
    transmission_zone: Optional[str] = None,
    substation_ids: Optional[str] = None,
    commission_year: Optional[int] = None,
    commission_year_from: Optional[int] = None,
    commission_year_to: Optional[int] = None,
    failure_year: Optional[int] = None,
    failure_year_from: Optional[int] = None,
    failure_year_to: Optional[int] = None,
    replacement_year: Optional[int] = None,
    replacement_year_from: Optional[int] = None,
    replacement_year_to: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Return equipment counts grouped by a single field for the full filtered dataset."""
    org_id = _enforce_org_scope(current_user)
    _require_permission(db, current_user, "can_view")

    # Determine the group column expression
    valid_fields = {
        "equipment_type": None,  # handled via join
        "manufacturer": Equipment.manufacturer,
        "voltage_class": Equipment.voltage_class,
        "model_type": Equipment.model_number,
        "bay_number": Equipment.bay_number,
        "commission_year": extract("year", Equipment.commissioned_date),
        "failure_year": extract("year", Equipment.retired_date),
        "replacement_year": extract("year", Equipment.commissioned_date),
    }
    if group_by not in valid_fields:
        raise HTTPException(status_code=400, detail=f"Invalid group_by: {group_by}")

    from services.equipment_service import EquipmentService as ES

    if group_by == "equipment_type":
        group_col = CategoryMaster.name
        query = db.query(CategoryMaster.name, func.count(Equipment.id)).outerjoin(
            CategoryMaster, Equipment.equipment_type_id == CategoryMaster.id
        )
    else:
        group_col = valid_fields[group_by]
        query = db.query(group_col, func.count(Equipment.id))

    query = query.filter(Equipment.organization_id == org_id)

    if department_id:
        dept_ids = ES._get_department_subtree_ids(db, department_id)
        query = query.filter(Equipment.department_id.in_(dept_ids))
    if equipment_type_id:
        query = query.filter(Equipment.equipment_type_id == equipment_type_id)
    if status:
        query = query.filter(Equipment.status == status)
    if voltage_class:
        query = query.filter(Equipment.voltage_class == voltage_class)
    if manufacturer:
        query = query.filter(Equipment.manufacturer.ilike(f"%{manufacturer}%"))
    if model_number:
        query = query.filter(Equipment.model_number.ilike(f"%{model_number}%"))
    if search:
        query = query.filter(
            (Equipment.ueic.ilike(f"%{search}%")) |
            (Equipment.bay_number.ilike(f"%{search}%")) |
            (Equipment.manufacturer.ilike(f"%{search}%")) |
            (Equipment.model_number.ilike(f"%{search}%")) |
            (Equipment.factory_serial_number.ilike(f"%{search}%"))
        )
    if substation_ids:
        id_list = [s.strip() for s in substation_ids.split(",") if s.strip()]
        if id_list:
            try:
                parsed = [UUID(i) for i in id_list]
                query = query.filter(Equipment.department_id.in_(parsed))
            except ValueError:
                pass
    if transmission_zone:
        dept_ids = ES._get_descendants_of_named(db, org_id, transmission_zone)
        if not dept_ids:
            return {}
        query = query.filter(Equipment.department_id.in_(dept_ids))
    if wm_circle:
        dept_ids = ES._get_descendants_of_named(db, org_id, wm_circle)
        if not dept_ids:
            return {}
        query = query.filter(Equipment.department_id.in_(dept_ids))
    if tlss_division:
        dept_ids = ES._get_descendants_of_named(db, org_id, tlss_division)
        if not dept_ids:
            return {}
        query = query.filter(Equipment.department_id.in_(dept_ids))
    if group_by == "replacement_year":
        query = query.filter(Equipment.replaces_equipment_id.isnot(None))
    if commission_year:
        query = query.filter(extract("year", Equipment.commissioned_date) == commission_year)
    if commission_year_from:
        query = query.filter(extract("year", Equipment.commissioned_date) >= commission_year_from)
    if commission_year_to:
        query = query.filter(extract("year", Equipment.commissioned_date) <= commission_year_to)
    if failure_year:
        query = query.filter(extract("year", Equipment.retired_date) == failure_year)
    if failure_year_from:
        query = query.filter(extract("year", Equipment.retired_date) >= failure_year_from)
    if failure_year_to:
        query = query.filter(extract("year", Equipment.retired_date) <= failure_year_to)
    if replacement_year:
        query = query.filter(extract("year", Equipment.commissioned_date) == replacement_year)
    if replacement_year_from:
        query = query.filter(extract("year", Equipment.commissioned_date) >= replacement_year_from)
    if replacement_year_to:
        query = query.filter(extract("year", Equipment.commissioned_date) <= replacement_year_to)

    rows = query.group_by(group_col).all()
    result = {}
    for key, count in rows:
        label = str(int(key)) if key is not None and isinstance(key, float) else (str(key) if key is not None else "Unknown")
        result[label] = count
    return result


# ═══════════════════════════════════════════════════════════════════════════════
# ② OTHER STATIC ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/types-by-category/{equipment_type_id}")
def get_types_by_category_for_type(
    equipment_type_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Return types_by_category for a CategoryMaster (equipment type) ID."""
    class _FakeEq:
        pass
    fake = _FakeEq()
    fake.equipment_type_id = equipment_type_id
    return _types_by_category_for_equipment(db, fake)


@router.get("/types-by-test-type/{test_type_id}")
def get_types_by_test_type(
    test_type_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Return types_by_category for the equipment category that owns test_type_id."""
    from models import CategoryDetails
    detail = db.query(CategoryDetails).filter(CategoryDetails.id == test_type_id).first()
    if not detail or not detail.category_master_id:
        return {"test": [], "maintenance": [], "inspection": [], "repair_lifecycle": []}
    
    class _FakeEq:
        pass
    fake = _FakeEq()
    fake.equipment_type_id = detail.category_master_id
    return _types_by_category_for_equipment(db, fake)


@router.get("/by-ueic/{ueic}", response_model=EquipmentResponse)
def get_equipment_by_ueic(
    ueic: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    org_id = _enforce_org_scope(current_user)
    _require_permission(db, current_user, "can_view")
    equipment = EquipmentService.get_equipment_by_ueic(db, ueic)
    if not equipment:
        raise HTTPException(status_code=404, detail="Equipment not found")
    if equipment.organization_id != org_id:
        raise HTTPException(status_code=404, detail="Equipment not found")
    return _to_response(db, equipment)


@router.get("/manufacturers")
def get_manufacturers(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Return distinct manufacturer names present in the org's equipment."""
    org_id = _enforce_org_scope(current_user)
    _require_permission(db, current_user, "can_view")
    from sqlalchemy import distinct
    rows = (
        db.query(distinct(Equipment.manufacturer))
        .filter(
            Equipment.organization_id == org_id,
            Equipment.manufacturer.isnot(None),
            Equipment.manufacturer != "",
        )
        .order_by(Equipment.manufacturer)
        .all()
    )
    return [{"id": r[0], "name": r[0]} for r in rows]


@router.get("/models")
def get_models(
    manufacturer: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Return distinct model numbers, optionally filtered by manufacturer."""
    org_id = _enforce_org_scope(current_user)
    _require_permission(db, current_user, "can_view")
    from sqlalchemy import distinct
    query = db.query(distinct(Equipment.model_number)).filter(
        Equipment.organization_id == org_id,
        Equipment.model_number.isnot(None),
        Equipment.model_number != "",
    )
    if manufacturer:
        query = query.filter(Equipment.manufacturer.ilike(f"%{manufacturer}%"))
    rows = query.order_by(Equipment.model_number).all()
    return [{"model_number": r[0]} for r in rows]


@router.get("/substations")
def get_substations(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Return leaf departments (substations) for the org."""
    org_id = _enforce_org_scope(current_user)
    _require_permission(db, current_user, "can_view")
    from sqlalchemy import text
    sql = text(
        """
        WITH RECURSIVE dept_tree AS (
            SELECT id, name, code, parent_department_id, 0 AS depth
            FROM org_departments
            WHERE organization_id = :org_id
              AND is_active = true
              AND parent_department_id IS NULL

            UNION ALL

            SELECT d.id, d.name, d.code, d.parent_department_id, dt.depth + 1
            FROM org_departments d
            INNER JOIN dept_tree dt ON d.parent_department_id = dt.id
            WHERE d.is_active = true
        ),
        leaf_depts AS (
            SELECT dt.id, dt.name, dt.code
            FROM dept_tree dt
            WHERE NOT EXISTS (
                SELECT 1 FROM org_departments child
                WHERE child.parent_department_id = dt.id
                  AND child.is_active = true
            )
        )
        SELECT id, name, code FROM leaf_depts ORDER BY name
        """
    )
    rows = db.execute(sql, {"org_id": str(org_id)}).fetchall()
    return [{"id": str(r[0]), "name": r[1], "code": r[2]} for r in rows]


@router.get("/divisions")
def get_divisions(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Return depth-2 departments (TLSS Divisions)."""
    org_id = _enforce_org_scope(current_user)
    _require_permission(db, current_user, "can_view")
    return _get_departments_at_depth(db, org_id, depth=2)


@router.get("/circles")
def get_circles(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Return depth-1 departments (W&M Circles)."""
    org_id = _enforce_org_scope(current_user)
    _require_permission(db, current_user, "can_view")
    return _get_departments_at_depth(db, org_id, depth=1)


@router.get("/zones")
def get_zones(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Return depth-0 departments (Transmission Zones / root)."""
    org_id = _enforce_org_scope(current_user)
    _require_permission(db, current_user, "can_view")
    return _get_departments_at_depth(db, org_id, depth=0)

@router.get("/department-hierarchy-with-counts")
def get_department_hierarchy_with_counts(
    parent_id: Optional[UUID] = None,
    # ── same filters as /stats/counts ──────────────────────────────────
    equipment_type_id: Optional[int] = None,
    status: Optional[str] = None,
    voltage_class: Optional[str] = None,
    manufacturer: Optional[str] = None,
    model_number: Optional[str] = None,
    search: Optional[str] = None,
    commission_year: Optional[int] = None,
    commission_year_from: Optional[int] = None,
    commission_year_to: Optional[int] = None,
    failure_year: Optional[int] = None,
    failure_year_from: Optional[int] = None,
    failure_year_to: Optional[int] = None,
    replacement_year: Optional[int] = None,
    replacement_year_from: Optional[int] = None,
    replacement_year_to: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Department hierarchy where each node also carries equipment_count —
    the count of equipment in that node's full subtree, with the same
    filters applied as the main list/stats endpoints."""
    org_id = _enforce_org_scope(current_user)
    _require_permission(db, current_user, "can_view")

    from sqlalchemy import text
    from utils.common_service import get_user_dept_scope

    # When fetching root cards (no parent_id), scope to user's own dept directly
    scoped_root_id = None
    if parent_id is None:
        is_admin, user_dept_id = get_user_dept_scope(db, current_user.id, org_id)
        if not is_admin and user_dept_id:
            scoped_root_id = user_dept_id

    sql = text(
        """
        SELECT id, name,
               EXISTS (
                   SELECT 1 FROM org_departments child
                   WHERE child.parent_department_id = od.id
                     AND child.is_active = true
               ) AS has_children
        FROM org_departments od
        WHERE organization_id = :org_id
          AND is_active = true
          AND (
                CASE
                  WHEN :scoped_root_id IS NOT NULL THEN od.id = :scoped_root_id
                  WHEN :parent_id IS NULL THEN parent_department_id IS NULL
                  ELSE parent_department_id = :parent_id
                END
              )
        ORDER BY name
        """
    )
    rows = db.execute(
        sql, {
            "org_id": str(org_id),
            "parent_id": str(parent_id) if parent_id else None,
            "scoped_root_id": str(scoped_root_id) if scoped_root_id else None,
        }
    ).fetchall()

    result = []
    for r in rows:
        dept_id = r[0]
        dept_ids = EquipmentService._get_department_subtree_ids(db, dept_id)

        query = db.query(func.count(Equipment.id)).filter(
            Equipment.organization_id == org_id,
            Equipment.department_id.in_(dept_ids),
        )

        if equipment_type_id:
            query = query.filter(Equipment.equipment_type_id == equipment_type_id)
        if status:
            query = query.filter(Equipment.status == status)
        if voltage_class:
            query = query.filter(Equipment.voltage_class == voltage_class)
        if manufacturer:
            query = query.filter(Equipment.manufacturer.ilike(f"%{manufacturer}%"))
        if model_number:
            query = query.filter(Equipment.model_number.ilike(f"%{model_number}%"))
        if search:
            query = query.filter(
                (Equipment.ueic.ilike(f"%{search}%")) |
                (Equipment.bay_number.ilike(f"%{search}%")) |
                (Equipment.manufacturer.ilike(f"%{search}%")) |
                (Equipment.model_number.ilike(f"%{search}%")) |
                (Equipment.factory_serial_number.ilike(f"%{search}%"))
            )
        if commission_year:
            query = query.filter(extract('year', Equipment.commissioned_date) == commission_year)
        if commission_year_from:
            query = query.filter(extract('year', Equipment.commissioned_date) >= commission_year_from)
        if commission_year_to:
            query = query.filter(extract('year', Equipment.commissioned_date) <= commission_year_to)
        if failure_year:
            query = query.filter(extract('year', Equipment.retired_date) == failure_year)
        if failure_year_from:
            query = query.filter(extract('year', Equipment.retired_date) >= failure_year_from)
        if failure_year_to:
            query = query.filter(extract('year', Equipment.retired_date) <= failure_year_to)
        if replacement_year:
            query = query.filter(
                Equipment.replaces_equipment_id.isnot(None),
                extract('year', Equipment.commissioned_date) == replacement_year,
            )
        if replacement_year_from:
            query = query.filter(
                Equipment.replaces_equipment_id.isnot(None),
                extract('year', Equipment.commissioned_date) >= replacement_year_from,
            )
        if replacement_year_to:
            query = query.filter(
                Equipment.replaces_equipment_id.isnot(None),
                extract('year', Equipment.commissioned_date) <= replacement_year_to,
            )

        count = query.scalar() or 0

        result.append({
            "id": str(dept_id),
            "name": r[1],
            "has_children": bool(r[2]),
            "equipment_count": count,
        })

    return result


@router.get("/testing-kits/import-template")
def download_testing_kit_import_template(
    dept_name: Optional[str] = Query(None, description="Pre-fill department name in sample row"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Return a CSV template for bulk-importing testing kits.

    The file has three sections:
      1. Header row
      2. One sample data row (delete before uploading)
      3. A reference block listing all valid kit_type values from the DB
    """
    import csv
    import io

    headers = [
        "kit_type",
        "manufacturer",
        "model_number",
        "factory_serial_number",
        "year_of_manufacture",
        "last_calibration_date",
        "calibration_due_date",
        "calibration_authority",
        "certificate_ref",
        "measurement_range",
        "accuracy_class",
        "is_portable",
        "rated_voltage",
        "department_name",
        "storage_location",
        "notes",
        "status",
    ]
    sample = [
        "Relay Test Kit",          # must match one of the valid kit types below
        "Omicron",
        "CMC 356",
        "SN-00123",                # unique serial number
        "2022",
        "2024-01-15",              # YYYY-MM-DD
        "2025-01-15",              # YYYY-MM-DD, must be after last_calibration_date
        "NABL Lab",
        "CERT-2024-001",
        "0-400 V / 0-10 A",
        "Class 0.5",
        "true",                    # true or false
        "415V",
        dept_name or "— select from reference below —",
        "Cabinet A2",
        "Primary relay test kit",
        "active",                  # active | under_repair | retired
    ]

    # Fetch valid kit types live from the DB
    kit_subtypes = (
        db.query(CategoryDetails.name)
        .join(CategoryMaster, CategoryDetails.category_master_id == CategoryMaster.id)
        .filter(
            CategoryMaster.name == "Testing Kit",
            CategoryMaster.is_active == True,
            CategoryDetails.is_active == True,
        )
        .order_by(CategoryDetails.name)
        .all()
    )
    valid_kit_types = [row[0] for row in kit_subtypes]

    # Fetch valid station names for this org
    org_id = _enforce_org_scope(current_user)
    departments = (
        db.query(OrgDepartment.name)
        .filter(OrgDepartment.organization_id == org_id, OrgDepartment.is_active == True)
        .order_by(OrgDepartment.name)
        .all()
    )
    valid_depts = [row[0] for row in departments]

    buf = io.StringIO()
    writer = csv.writer(buf)

    # Section 1 & 2: header + sample
    writer.writerow(headers)
    writer.writerow(sample)

    # Section 3: reference block
    writer.writerow([])
    writer.writerow(["# ── REFERENCE: Valid kit_type values (copy exactly) ──"])
    writer.writerow(["# kit_type"])
    for name in valid_kit_types:
        writer.writerow([f"# {name}"])
    writer.writerow([])
    writer.writerow(["# ── REFERENCE: Valid department_name (station) values (copy exactly) ──"])
    writer.writerow(["# department_name"])
    for name in valid_depts:
        writer.writerow([f"# {name}"])
    writer.writerow([])
    writer.writerow(["# ── REFERENCE: Valid status values ──"])
    for s in ["# active", "# under_repair", "# retired"]:
        writer.writerow([s])
    writer.writerow([])
    writer.writerow(["# ── REFERENCE: Date format (any of these work) ──"])
    writer.writerow(["# YYYY-MM-DD  e.g. 2025-06-15"])
    writer.writerow(["# DD-MM-YYYY  e.g. 15-06-2025"])
    writer.writerow(["# DD/MM/YYYY  e.g. 15/06/2025"])
    writer.writerow([])
    writer.writerow(["# DELETE this reference section and the sample row before uploading."])

    csv_bytes = buf.getvalue().encode("utf-8-sig")

    return StreamingResponse(
        iter([csv_bytes]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=testing_kit_import_template.csv"},
    )


@router.post("/testing-kits/bulk-import")
def bulk_import_testing_kits(
    file: UploadFile = File(...),
    org_id: UUID = Query(...),
    skip_errors: bool = Query(False, description="Skip invalid rows instead of aborting"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Bulk-import testing kits from a CSV file.

    Accepts the CSV produced by /testing-kits/import-template.
    Reference/comment rows (starting with #) and blank rows are ignored.
    Returns { imported, skipped, errors: [{row, field, reason}] }.
    """
    import csv
    import io
    import uuid as _uuid
    from datetime import datetime as _dt

    _require_permission(db, current_user, "can_import")

    content = file.file.read().decode("utf-8-sig")
    def _is_comment_or_blank(line: str) -> bool:
        s = line.strip()
        return not s or s.startswith("#") or s.startswith('"#')

    reader = csv.DictReader(
        (line for line in io.StringIO(content) if not _is_comment_or_blank(line))
    )

    # Fetch kit subtype lookup: name (lower) → CategoryDetails id
    kit_master = db.query(CategoryMaster).filter(
        CategoryMaster.name == "Testing Kit", CategoryMaster.is_active == True
    ).first()
    if not kit_master:
        raise HTTPException(status_code=400, detail="Testing Kit master category not found")

    kit_subtypes = {
        cd.name.lower(): cd
        for cd in db.query(CategoryDetails).filter(
            CategoryDetails.category_master_id == kit_master.id,
            CategoryDetails.is_active == True,
        ).all()
    }

    # Department lookup: name (lower) → OrgDepartment
    depts = {
        d.name.lower(): d
        for d in db.query(OrgDepartment).filter(OrgDepartment.organization_id == org_id).all()
    }

    # Existing serial numbers (uniqueness check)
    existing_serials = {
        r[0].lower()
        for r in db.query(Equipment.factory_serial_number).filter(
            Equipment.organization_id == org_id,
            Equipment.equipment_type_id == kit_master.id,
            Equipment.factory_serial_number.isnot(None),
        ).all()
        if r[0]
    }

    imported, skipped = 0, 0
    errors = []

    for row_num, row in enumerate(reader, start=2):  # row 1 = headers
        # Skip blank rows and any comment row that slipped through (e.g. quoted "#..." values)
        if not any(v and v.strip() for v in row.values()):
            continue
        first_val = next(iter(row.values()), "") or ""
        if first_val.strip().startswith("#"):
            continue

        row_errors = []

        kit_type_name = (row.get("kit_type") or "").strip()
        manufacturer = (row.get("manufacturer") or "").strip() or None
        model_number = (row.get("model_number") or "").strip() or None
        serial = (row.get("factory_serial_number") or "").strip()
        year_mfg = (row.get("year_of_manufacture") or "").strip() or None
        last_cal = (row.get("last_calibration_date") or "").strip() or None
        cal_due = (row.get("calibration_due_date") or "").strip() or None
        cal_auth = (row.get("calibration_authority") or "").strip() or None
        cert_ref = (row.get("certificate_ref") or "").strip() or None
        meas_range = (row.get("measurement_range") or "").strip() or None
        accuracy = (row.get("accuracy_class") or "").strip() or None
        portable_raw = (row.get("is_portable") or "true").strip().lower()
        rated_voltage = (row.get("rated_voltage") or "").strip() or None
        dept_name = (row.get("department_name") or "").strip()
        storage_loc = (row.get("storage_location") or "").strip() or None
        notes = (row.get("notes") or "").strip() or None
        status_raw = (row.get("status") or "active").strip().lower()

        # Validate kit_type
        kit_detail = kit_subtypes.get(kit_type_name.lower())
        if not kit_detail:
            row_errors.append(f"kit_type '{kit_type_name}' not found — valid values: {', '.join(kit_subtypes.keys())}")

        # Validate serial
        if not serial:
            row_errors.append("factory_serial_number is required")
        elif serial.lower() in existing_serials:
            row_errors.append(f"factory_serial_number '{serial}' already exists in this org")

        # Validate dates
        def _parse_date(val, field):
            if not val:
                return None
            for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
                try:
                    return _dt.strptime(val, fmt)
                except ValueError:
                    continue
            row_errors.append(f"{field}: unrecognised date '{val}' — use YYYY-MM-DD or DD-MM-YYYY")
            return None

        last_cal_dt = _parse_date(last_cal, "last_calibration_date")
        cal_due_dt = _parse_date(cal_due, "calibration_due_date")
        if last_cal_dt and cal_due_dt and cal_due_dt <= last_cal_dt:
            row_errors.append("calibration_due_date must be after last_calibration_date")

        # Validate status
        valid_statuses = ("active", "under_repair", "retired")
        if status_raw not in valid_statuses:
            row_errors.append(f"status must be one of: {', '.join(valid_statuses)}")

        # Validate portable
        is_portable = portable_raw in ("true", "1", "yes")

        # Resolve department — optional, kit can be registered without a station
        dept = depts.get(dept_name.lower()) if dept_name else None
        if dept_name and not dept:
            row_errors.append(f"department_name '{dept_name}' not found — must match an exact station name or leave blank")
        dept_id = dept.id if dept else None

        if row_errors:
            skipped += 1
            errors.append({"row": row_num, "kit_type": kit_type_name, "serial": serial, "errors": row_errors})
            if not skip_errors:
                raise HTTPException(
                    status_code=422,
                    detail={"message": f"Row {row_num} invalid", "errors": row_errors},
                )
            continue

        nameplate = {
            "kit_subtype": kit_type_name,
            **({"last_calibration_date": last_cal} if last_cal else {}),
            **({"calibration_due_date": cal_due} if cal_due else {}),
            **({"calibration_authority": cal_auth} if cal_auth else {}),
            **({"certificate_ref": cert_ref} if cert_ref else {}),
            **({"measurement_range": meas_range} if meas_range else {}),
            **({"accuracy_class": accuracy} if accuracy else {}),
            "is_portable": is_portable,
            **({"rated_voltage": rated_voltage} if rated_voltage else {}),
            **({"storage_location": storage_loc} if storage_loc else {}),
        }

        equipment = EquipmentService.create_equipment(
            db=db,
            organization_id=org_id,
            department_id=dept_id,
            equipment_type_id=kit_master.id,
            bay_number=kit_type_name,
            voltage_class="0LV",
            nameplate_data=nameplate,
            manufacturer=manufacturer,
            model_number=model_number,
            factory_serial_number=serial,
            year_of_manufacture=int(year_mfg) if year_mfg and year_mfg.isdigit() else None,
            created_by=current_user.id,
        )
        existing_serials.add(serial.lower())
        imported += 1

    db.commit()
    return {"imported": imported, "skipped": skipped, "errors": errors}


@router.get("/testing-kits")
def get_testing_kits(
    org_id: UUID = Query(..., description="Organization ID"),
    department_id: Optional[UUID] = Query(None, description="Station dept to check first"),
    equipment_type_id: Optional[int] = Query(None, description="Filter kits required for this equipment type"),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Return testing kits (equipment with type 'Testing Kit') available at a station."""
    kit_master = db.query(CategoryMaster).filter(
        CategoryMaster.name == "Testing Kit",
        CategoryMaster.is_active == True,
    ).first()
    if not kit_master:
        return {"at_station": [], "nearby": [], "all": []}

    required_kit_type_ids: Optional[list] = None
    if equipment_type_id:
        mappings = db.query(EquipmentTypeKitMapping).filter(
            EquipmentTypeKitMapping.equipment_type_id == equipment_type_id,
        ).all()
        required_kit_type_ids = [m.kit_type_id for m in mappings]
        if not required_kit_type_ids:
            return {"at_station": [], "nearby": [], "required_mappings": []}

    def _kit_row(eq: Equipment, location_label: str) -> dict:
        _np = eq.nameplate_data or {}
        _kit_type = (
            _np.get("kit_subtype")
            or eq.bay_number
            or (eq.equipment_type.name if eq.equipment_type else "Testing Kit")
        )
        return {
            "id": str(eq.id),
            "ueic": eq.ueic,
            "kit_type": _kit_type,
            "manufacturer": eq.manufacturer,
            "model_number": eq.model_number,
            "factory_serial_number": eq.factory_serial_number,
            "status": eq.status.value if eq.status else "active",
            "department_id": str(eq.department_id),
            "department_name": eq.department.name if eq.department else "",
            "location_label": location_label,
            "nameplate_data": _np,
        }

    def _query_kits(dept_ids: list) -> list:
        q = db.query(Equipment).filter(
            Equipment.organization_id == org_id,
            Equipment.equipment_type_id == kit_master.id,
            Equipment.status == "active",
            Equipment.department_id.in_(dept_ids),
        )
        return q.all()

    if not department_id:
        all_kits = db.query(Equipment).filter(
            Equipment.organization_id == org_id,
            Equipment.equipment_type_id == kit_master.id,
            Equipment.status == "active",
        ).all()
        return {"all": [_kit_row(k, k.department.name if k.department else "") for k in all_kits]}

    # Recursively collect all descendant dept IDs so zone-level drill-down works
    def _all_descendant_ids(root_id) -> list:
        from sqlalchemy import text as sa_text
        rows = db.execute(sa_text("""
            WITH RECURSIVE subtree AS (
                SELECT id FROM org_departments WHERE id = :root_id
                UNION ALL
                SELECT d.id FROM org_departments d
                INNER JOIN subtree s ON d.parent_department_id = s.id
                WHERE d.is_active = true
            )
            SELECT id FROM subtree
        """), {"root_id": str(root_id)}).fetchall()
        return [r[0] for r in rows]

    dept_ids = _all_descendant_ids(department_id)
    all_kits = _query_kits(dept_ids)

    return {"all": [_kit_row(k, k.department.name if k.department else "") for k in all_kits]}


# ═══════════════════════════════════════════════════════════════════════════════
# Department hierarchy with per-node equipment counts
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/department-hierarchy-with-counts")
def get_department_hierarchy_with_counts(
    parent_id: Optional[str] = None,
    status: Optional[str] = None,
    equipment_type_id: Optional[int] = None,
    search: Optional[str] = None,
    voltage_class: Optional[str] = None,
    manufacturer: Optional[str] = None,
    model_number: Optional[str] = None,
    commission_year: Optional[int] = None,
    commission_year_from: Optional[int] = None,
    commission_year_to: Optional[int] = None,
    failure_year: Optional[int] = None,
    failure_year_from: Optional[int] = None,
    failure_year_to: Optional[int] = None,
    replacement_year: Optional[int] = None,
    replacement_year_from: Optional[int] = None,
    replacement_year_to: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Return child departments of parent_id (or roots) with equipment counts respecting active filters."""
    from sqlalchemy import extract as sa_extract, func as sa_func

    org_id = _enforce_org_scope(current_user)

    # Fetch departments at the requested level
    parent_uuid = None
    if parent_id:
        try:
            parent_uuid = _uuid.UUID(parent_id)
        except ValueError:
            raise HTTPException(status_code=422, detail="Invalid parent_id UUID")

    depts = (
        db.query(OrgDepartment)
        .filter(
            OrgDepartment.organization_id == org_id,
            OrgDepartment.is_active == True,
            OrgDepartment.parent_department_id == parent_uuid,
        )
        .order_by(OrgDepartment.name)
        .all()
    )

    result = []
    for dept in depts:
        # Check whether this dept has children
        has_children = (
            db.query(OrgDepartment.id)
            .filter(
                OrgDepartment.parent_department_id == dept.id,
                OrgDepartment.is_active == True,
            )
            .first()
        ) is not None

        # Count equipment in the full subtree of this dept (applying all active filters)
        from services.equipment_service import EquipmentService
        subtree_ids = EquipmentService._get_department_subtree_ids(db, dept.id)

        q = db.query(sa_func.count(Equipment.id)).filter(
            Equipment.organization_id == org_id,
            Equipment.department_id.in_(subtree_ids),
        )
        if status:
            q = q.filter(Equipment.status == status)
        if equipment_type_id:
            q = q.filter(Equipment.equipment_type_id == equipment_type_id)
        if voltage_class:
            q = q.filter(Equipment.voltage_class == voltage_class)
        if manufacturer:
            q = q.filter(Equipment.manufacturer.ilike(f"%{manufacturer}%"))
        if model_number:
            q = q.filter(Equipment.model_number.ilike(f"%{model_number}%"))
        if search:
            q = q.filter(
                (Equipment.ueic.ilike(f"%{search}%")) |
                (Equipment.bay_number.ilike(f"%{search}%")) |
                (Equipment.manufacturer.ilike(f"%{search}%")) |
                (Equipment.model_number.ilike(f"%{search}%")) |
                (Equipment.factory_serial_number.ilike(f"%{search}%"))
            )
        if commission_year:
            q = q.filter(sa_extract('year', Equipment.commissioned_date) == commission_year)
        if commission_year_from:
            q = q.filter(sa_extract('year', Equipment.commissioned_date) >= commission_year_from)
        if commission_year_to:
            q = q.filter(sa_extract('year', Equipment.commissioned_date) <= commission_year_to)
        if failure_year:
            q = q.filter(sa_extract('year', Equipment.retired_date) == failure_year)
        if failure_year_from:
            q = q.filter(sa_extract('year', Equipment.retired_date) >= failure_year_from)
        if failure_year_to:
            q = q.filter(sa_extract('year', Equipment.retired_date) <= failure_year_to)
        if replacement_year:
            q = q.filter(sa_extract('year', Equipment.commissioned_date) == replacement_year)
        if replacement_year_from:
            q = q.filter(sa_extract('year', Equipment.commissioned_date) >= replacement_year_from)
        if replacement_year_to:
            q = q.filter(sa_extract('year', Equipment.commissioned_date) <= replacement_year_to)

        count = q.scalar() or 0

        result.append({
            "id": str(dept.id),
            "name": dept.name,
            "department_name": dept.name,
            "parent_department_id": str(dept.parent_department_id) if dept.parent_department_id else None,
            "equipment_count": count,
            "has_children": has_children,
            "organization_id": str(dept.organization_id),
        })

    # Sort by count descending
    result.sort(key=lambda x: x["equipment_count"], reverse=True)
    return result


# ═══════════════════════════════════════════════════════════════════════════════
# ③ DYNAMIC ROUTES — all contain /{equipment_id: UUID}
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/{equipment_id}", response_model=EquipmentResponse)
def get_equipment(
    equipment_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    org_id = _enforce_org_scope(current_user)
    _require_permission(db, current_user, "can_view")
    equipment = EquipmentService.get_equipment(db, equipment_id)
    if not equipment:
        raise HTTPException(status_code=404, detail="Equipment not found")
    if equipment.organization_id != org_id:
        raise HTTPException(status_code=404, detail="Equipment not found")
    return _to_response(db, equipment)


@router.put("/{equipment_id}", response_model=EquipmentResponse)
def update_equipment(
    equipment_id: UUID,
    data: EquipmentUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    org_id = _enforce_org_scope(current_user)
    _require_permission(db, current_user, "can_edit")
    existing = EquipmentService.get_equipment(db, equipment_id)
    if not existing or existing.organization_id != org_id:
        raise HTTPException(status_code=404, detail="Equipment not found")
    equipment = EquipmentService.update_equipment(
        db=db,
        equipment_id=equipment_id,
        modified_by=current_user.id,
        **data.model_dump(exclude_unset=True),
    )
    db.commit()
    db.refresh(equipment)
    return _to_response(db, equipment)


@router.post("/{equipment_id}/retire", response_model=EquipmentResponse)
def retire_equipment(
    equipment_id: UUID,
    data: EquipmentRetireRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Retire equipment (soft-delete). UEIC and historical data remain."""
    org_id = _enforce_org_scope(current_user)
    _require_permission(db, current_user, "can_edit")
    existing = EquipmentService.get_equipment(db, equipment_id)
    if not existing or existing.organization_id != org_id:
        raise HTTPException(status_code=404, detail="Equipment not found")
    equipment = EquipmentService.retire_equipment(
        db=db,
        equipment_id=equipment_id,
        reason=data.reason,
        modified_by=current_user.id,
    )
    db.commit()
    db.refresh(equipment)

    try:
        from services.notification_service import NotificationService
        retired_by = (
            f"{current_user.firstname or ''} {current_user.lastname or ''}".strip()
            or current_user.email
        )
        NotificationService(db).notify_equipment_retired(
            equipment,
            retired_by=retired_by,
            reason=data.reason or "",
            organization_id=org_id,
            department_id=equipment.department_id,
        )
    except Exception as _n:
        print(f"[WARN] equipment_retired notification failed: {_n}")

    return _to_response(db, equipment)


@router.post("/{equipment_id}/replace", status_code=status.HTTP_201_CREATED)
async def replace_equipment(
    equipment_id: UUID,
    reason: str = Form(..., description="Reason for replacement"),
    reason_type: str = Form(
        "other",
        description="'recommendation_compliance' or 'other'",
    ),
    recommendation_id: Optional[str] = Form(None),
    nameplate_data: Optional[str] = Form(None, description="JSON-encoded nameplate fields"),
    manufacturer: Optional[str] = Form(None),
    model_number: Optional[str] = Form(None),
    factory_serial_number: Optional[str] = Form(None),
    year_of_manufacture: Optional[int] = Form(None),
    commissioned_date: Optional[str] = Form(None, description="ISO date string"),
    phase: Optional[str] = Form(None, description="R, Y, or B"),
    ct_ratio_actual: Optional[str] = Form(None),
    ct_ratio_current: Optional[str] = Form(None),
    pt_ratio: Optional[str] = Form(None),
    vector_group: Optional[str] = Form(None),
    impedance_pct: Optional[float] = Form(None),
    voltage_class: Optional[str] = Form(None),
    bay_number: Optional[str] = Form(None),
    latitude: Optional[float] = Form(None),
    longitude: Optional[float] = Form(None),
    precommission_request_id: Optional[str] = Form(None),
    analysis_report: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Retire old equipment and register a replacement."""
    import json
    from datetime import datetime as _dt_cls

    org_id = _enforce_org_scope(current_user)
    _require_permission(db, current_user, "can_add")

    existing = EquipmentService.get_equipment(db, equipment_id)
    if not existing or existing.organization_id != org_id:
        raise HTTPException(status_code=404, detail="Equipment not found")

    if reason_type == "other" and analysis_report is None:
        raise HTTPException(
            status_code=400,
            detail="An analysis report PDF is required when replacement reason is not linked to a recommendation.",
        )
    if reason_type == "recommendation_compliance" and not recommendation_id:
        raise HTTPException(
            status_code=400,
            detail="recommendation_id is required when reason_type is 'recommendation_compliance'.",
        )

    report_path: Optional[str] = None
    if analysis_report is not None:
        file_ext = os.path.splitext(analysis_report.filename or "report.pdf")[1] or ".pdf"
        safe_name = f"{_uuid.uuid4()}{file_ext}"
        dest = os.path.join(UPLOADS_DIR, safe_name)
        content = await analysis_report.read()
        with open(dest, "wb") as f:
            f.write(content)
        report_path = f"uploads/analysis_reports/{safe_name}"

    parsed_nameplate: Optional[dict] = None
    if nameplate_data:
        try:
            parsed_nameplate = json.loads(nameplate_data)
        except Exception:
            raise HTTPException(status_code=400, detail="nameplate_data must be valid JSON.")

    parsed_date = None
    if commissioned_date:
        try:
            parsed_date = _dt_cls.fromisoformat(commissioned_date)
        except Exception:
            raise HTTPException(status_code=400, detail="commissioned_date must be a valid ISO date.")

    rec_uuid: Optional[UUID] = None
    if recommendation_id:
        try:
            rec_uuid = UUID(recommendation_id)
        except Exception:
            raise HTTPException(status_code=400, detail="recommendation_id must be a valid UUID.")

    old, new = EquipmentService.replace_equipment(
        db=db,
        old_equipment_id=equipment_id,
        reason=reason,
        created_by=current_user.id,
        reason_type=reason_type,
        recommendation_id=rec_uuid,
        analysis_report_path=report_path,
        nameplate_data=parsed_nameplate,
        commissioned_date=parsed_date,
        manufacturer=manufacturer,
        model_number=model_number,
        factory_serial_number=factory_serial_number,
        year_of_manufacture=year_of_manufacture,
        phase=phase,
        ct_ratio_actual=ct_ratio_actual,
        ct_ratio_current=ct_ratio_current,
        pt_ratio=pt_ratio,
        vector_group=vector_group,
        impedance_pct=impedance_pct,
        voltage_class=voltage_class,
        bay_number=bay_number,
        latitude=latitude,
        longitude=longitude,
    )
    db.commit()
    db.refresh(old)
    db.refresh(new)

    if precommission_request_id:
        try:
            _pcr_uuid = UUID(precommission_request_id)
        except Exception:
            raise HTTPException(status_code=400, detail="precommission_request_id must be a valid UUID.")
        try:
            from services.precommission_service import PreCommissionService
            PreCommissionService(db).link_equipment(_pcr_uuid, new.id, current_user)
            new.precommission_request_id = _pcr_uuid
            db.commit()
            db.refresh(new)
        except Exception as _pcr_exc:
            print(f"[WARN] PCR link failed on replace (non-fatal): {_pcr_exc}")

    try:
        from services.test_request_schedule_service import TestRequestScheduleService
        TestRequestScheduleService.instantiate_equipment_schedules(db, new, current_user.id)
    except Exception as exc:
        import traceback
        traceback.print_exc()
        print(f"[WARN] instantiate_equipment_schedules failed for replacement: {exc}")

    try:
        from services.notification_service import NotificationService
        eq_type_name = (
            existing.equipment_type.name
            if existing.equipment_type
            else str(existing.equipment_type_id)
        )
        dept_name = existing.department.name if existing.department else "-"
        NotificationService(db).fire(
            event_type="equipment_replacement",
            context={
                "old_ueic":       old.ueic,
                "new_ueic":       new.ueic,
                "equipment_type": eq_type_name,
                "department":     dept_name,
                "reason_type":    reason_type,
                "reason":         reason,
                "replaced_by":    (
                    f"{current_user.firstname or ''} {current_user.lastname or ''}".strip()
                    or current_user.email
                ),
                "replaced_on":    _dt(old.retired_date) or "-",
            },
            organization_id=org_id,
            department_id=existing.department_id,
            source_id=new.id,
            source_type="equipment",
            severity="info",
        )
    except Exception:
        pass

    report_url = f"/equipment/{new.id}/replacement-report"
    return {
        "retired_equipment": _to_response(db, old),
        "new_equipment":     _to_response(db, new),
        "report_url":        report_url,
    }


@router.get("/{equipment_id}/replacement-report")
def download_replacement_report(
    equipment_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    org_id = _enforce_org_scope(current_user)
    _require_permission(db, current_user, "can_view")
    new_eq = EquipmentService.get_equipment(db, equipment_id)
    if not new_eq or new_eq.organization_id != org_id:
        raise HTTPException(status_code=404, detail="Equipment not found")
    if not new_eq.replaces_equipment_id:
        raise HTTPException(
            status_code=400,
            detail="This equipment is not a replacement unit. No replacement report available.",
        )
    from services.equipment_replacement_pdf_service import EquipmentReplacementPDFService
    try:
        buf: BytesIO = EquipmentReplacementPDFService(db).generate_pdf(
            old_equipment_id=new_eq.replaces_equipment_id,
            new_equipment_id=new_eq.id,
        )
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    filename = f"Replacement_Report_{new_eq.ueic}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{equipment_id}/applicable-tests")
def get_applicable_tests(
    equipment_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Test types applicable to this equipment's type."""
    org_id = _enforce_org_scope(current_user)
    _require_permission(db, current_user, "can_view")
    existing = EquipmentService.get_equipment(db, equipment_id)
    if not existing or existing.organization_id != org_id:
        raise HTTPException(status_code=404, detail="Equipment not found")
    tests = EquipmentService.get_applicable_tests(db, equipment_id)
    from models import OrgTestTemplate
    def _template_flags(test_type_id: int) -> dict:
        tpl = (
            db.query(OrgTestTemplate)
            .filter(OrgTestTemplate.test_type_id == test_type_id)
            .order_by(OrgTestTemplate.version.desc())
            .first()
        )
        data = (tpl.template_data or {}) if tpl else {}
        return {
            "enable_cumulative": bool(data.get("enable_cumulative", False)),
            "enable_calibration": bool(data.get("enable_calibration", False)),
        }
    return [
        {
            "id": t.id,
            "name": t.name,
            "description": t.description,
            "category_type": t.category_type,
            "is_active": t.is_active,
            **_template_flags(t.id),
        }
        for t in tests
    ]


@router.get("/{equipment_id}/history")
def get_equipment_history(
    equipment_id: UUID,
    skip: int = 0,
    limit: int = 50,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Testing and failure history for one equipment unit."""
    from models import TestingRequest
    from sqlalchemy.orm import joinedload
    org_id = _enforce_org_scope(current_user)
    _require_permission(db, current_user, "can_view")
    equipment = EquipmentService.get_equipment(db, equipment_id)
    if not equipment or equipment.organization_id != org_id:
        raise HTTPException(status_code=404, detail="Equipment not found")
    rows = (
        db.query(TestingRequest)
        .options(
            joinedload(TestingRequest.originator),
            joinedload(TestingRequest.test_results),
        )
        .filter(TestingRequest.equipment_id == equipment_id)
        .order_by(TestingRequest.cts.desc())
        .offset(skip)
        .limit(limit)
        .all()
    )
    def _fmt(r):
        result = r.test_results[0] if r.test_results else None
        originator = r.originator
        return {
            "id": str(r.id),
            "request_number": r.request_number,
            "title": r.title,
            "request_category": r.request_category.value if r.request_category else None,
            "status": r.status.value if r.status else None,
            "priority": r.priority,
            "is_direct_submission": r.is_direct_submission,
            "overall_result": result.overall_result if result else None,
            "submitted_by": (
                f"{originator.firstname or ''} {originator.lastname or ''}".strip()
                if originator
                else None
            ),
            "cts": _dt(r.cts),
        }
    return {
        "equipment_id": str(equipment_id),
        "ueic": equipment.ueic,
        "total": len(rows),
        "records": [_fmt(r) for r in rows],
    }


@router.get("/{equipment_id}/location-hierarchy")
def get_equipment_location_hierarchy(
    equipment_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Full department hierarchy for an equipment's location."""
    org_id = _enforce_org_scope(current_user)
    _require_permission(db, current_user, "can_view")
    equipment = EquipmentService.get_equipment(db, equipment_id)
    if not equipment:
        raise HTTPException(status_code=404, detail="Equipment not found")
    if equipment.organization_id != org_id:
        raise HTTPException(status_code=404, detail="Equipment not found")
    ancestry = EquipmentService._get_department_ancestry_names(db, equipment.department_id)
    return {
        "equipment_id": str(equipment.id),
        "ueic": equipment.ueic,
        "department_id": str(equipment.department_id),
        "department_name": equipment.department.name if equipment.department else None,
        "hierarchy": ancestry,
    }


@router.post("/{equipment_id}/nameplate-files/{field_key}", status_code=200)
async def upload_nameplate_file(
    equipment_id: UUID,
    field_key: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Upload a file for a 'file'-type field in the nameplate template."""
    from datetime import datetime as _dt_cls
    org_id = _enforce_org_scope(current_user)
    _require_permission(db, current_user, "can_edit")
    equipment = EquipmentService.get_equipment(db, equipment_id)
    if not equipment or equipment.organization_id != org_id:
        raise HTTPException(status_code=404, detail="Equipment not found.")
    field_def = _resolve_nameplate_file_field(db, equipment, field_key)
    accepted = set(field_def.get("accept", ["image/jpeg", "application/pdf"]))
    content_type = file.content_type or ""
    if content_type not in accepted:
        raise HTTPException(
            status_code=400,
            detail=f"File type '{content_type}' not allowed. Accepted: {sorted(accepted)}",
        )
    max_bytes = field_def.get("max_size_kb", 10240) * 1024
    content = await file.read()
    if len(content) > max_bytes:
        raise HTTPException(
            status_code=400,
            detail=f"File exceeds maximum size of {field_def.get('max_size_kb', 10240)} KB.",
        )
    eq_dir = os.path.join(NAMEPLATE_FILES_DIR, str(equipment_id))
    os.makedirs(eq_dir, exist_ok=True)
    ext = os.path.splitext(file.filename or "upload")[1] or (
        ".jpg" if "jpeg" in content_type else ".pdf"
    )
    stored_name = f"{field_key}_{_uuid.uuid4()}{ext}"
    dest = os.path.join(eq_dir, stored_name)
    with open(dest, "wb") as fh:
        fh.write(content)
    relative_path = f"uploads/nameplate_files/{equipment_id}/{stored_name}"
    nameplate = dict(equipment.nameplate_data or {})
    nameplate[field_key] = {
        "original_filename": file.filename,
        "path": relative_path,
        "size_bytes": len(content),
        "mime_type": content_type,
        "uploaded_at": _dt_cls.utcnow().isoformat(),
        "uploaded_by": str(current_user.id),
    }
    from sqlalchemy.orm.attributes import flag_modified
    equipment.nameplate_data = nameplate
    flag_modified(equipment, "nameplate_data")
    equipment.modified_by = current_user.id
    db.commit()
    return {
        "field_key": field_key,
        "original_filename": file.filename,
        "path": relative_path,
        "size_bytes": len(content),
        "mime_type": content_type,
    }


@router.get("/{equipment_id}/nameplate-files/{field_key}")
def download_nameplate_file(
    equipment_id: UUID,
    field_key: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Serve the uploaded file for a nameplate field."""
    org_id = _enforce_org_scope(current_user)
    _require_permission(db, current_user, "can_view")
    equipment = EquipmentService.get_equipment(db, equipment_id)
    if not equipment or equipment.organization_id != org_id:
        raise HTTPException(status_code=404, detail="Equipment not found.")
    nameplate = equipment.nameplate_data or {}
    file_meta = nameplate.get(field_key)
    if not isinstance(file_meta, dict) or "path" not in file_meta:
        raise HTTPException(
            status_code=404,
            detail=f"No file uploaded for field '{field_key}'.",
        )
    abs_path = os.path.join(os.path.dirname(__file__), "..", file_meta["path"])
    if not os.path.isfile(abs_path):
        raise HTTPException(status_code=404, detail="File not found on server.")
    import mimetypes
    def _stream():
        with open(abs_path, "rb") as fh:
            yield from iter(lambda: fh.read(65536), b"")
    mime = (
        file_meta.get("mime_type")
        or mimetypes.guess_type(abs_path)[0]
        or "application/octet-stream"
    )
    filename = file_meta.get("original_filename", os.path.basename(abs_path))
    return StreamingResponse(
        _stream(),
        media_type=mime,
        headers={"Content-Disposition": f'inline; filename="{filename}"'},
    )


@router.post("/sync-schedules", status_code=200)
def sync_equipment_schedules(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Backfill operational test schedules for all active equipment in the org.
    Calls instantiate_equipment_schedules for each equipment that doesn't yet
    have operational schedules — safe to call multiple times (skips existing).
    """
    org_id = _enforce_org_scope(current_user)
    from services.test_request_schedule_service import TestRequestScheduleService

    equipments = (
        db.query(Equipment)
        .filter(Equipment.organization_id == org_id, Equipment.status == "active")
        .all()
    )

    synced = 0
    errors = []
    for eq in equipments:
        try:
            TestRequestScheduleService.instantiate_equipment_schedules(db, eq, current_user.id)
            synced += 1
        except Exception as e:
            errors.append({"equipment_id": str(eq.id), "error": str(e)})

    return {"synced": synced, "errors": errors}
