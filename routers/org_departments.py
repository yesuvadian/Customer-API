"""
Organization department management endpoints.
Provides CRUD operations for departments within organizations.
"""

import io
import uuid as _uuid
from typing import List, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi import APIRouter, Depends, HTTPException, status, Body, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from uuid import UUID

from database import get_db
from middleware.org_auth import require_org_member, require_org_admin, require_org_admin_or_dept_admin
from models import OrgDepartment, User
from schemas import (
    OrgDepartmentCreate,
    OrgDepartmentUpdate,
    OrgDepartmentOut,
    User as UserSchema
)
from services.org_department_service import OrgDepartmentService


# ── Dept bulk-import response schema ─────────────────────────────────────────
class BulkDeptImportResult(BaseModel):
    total: int
    created: int
    skipped: int
    failed: int
    errors: List[str]


router = APIRouter(
    prefix="/organizations/{org_id}/departments",
    tags=["org-departments"]
)


@router.post("/", response_model=OrgDepartmentOut, status_code=status.HTTP_201_CREATED)
def create_department(
    org_id: UUID,
    dept_data: OrgDepartmentCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_org_admin_or_dept_admin)
):
    """
    Create a department within the organization.
    Only org admins or department admins can create departments.
    """
    # Ensure department is created in the correct org
    if dept_data.organization_id != org_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Department organization_id must match the URL org_id"
        )

    service = OrgDepartmentService(db)
    return service.create_department(dept_data, created_by=current_user.id)


@router.get("/", response_model=List[OrgDepartmentOut])
def list_departments(
    org_id: UUID,
    skip: int = 0,
    limit: Optional[int] = None,
    is_active: Optional[bool] = None,
    parent_department_id: Optional[UUID] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_org_member)
):
    """
    List all departments in the organization.
    Any organization member can list departments.
    """
    service = OrgDepartmentService(db)
    return service.list_departments(
        organization_id=org_id,
        skip=skip,
        limit=limit,
        is_active=is_active,
        parent_department_id=parent_department_id
    )


@router.get("/{dept_id}", response_model=OrgDepartmentOut)
def get_department(
    org_id: UUID,
    dept_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_org_member)
):
    """
    Get department details.
    Any organization member can view department details.
    """
    service = OrgDepartmentService(db)
    dept = service.get_department(dept_id)

    # Verify department belongs to the organization
    if dept.organization_id != org_id:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Department not found in this organization"
        )

    return dept


@router.put("/{dept_id}", response_model=OrgDepartmentOut)
def update_department(
    org_id: UUID,
    dept_id: UUID,
    dept_data: OrgDepartmentUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_org_admin_or_dept_admin)
):
    """
    Update department.
    Only org admins or department admins can update departments.
    """
    service = OrgDepartmentService(db)
    dept = service.get_department(dept_id)

    # Verify department belongs to the organization
    if dept.organization_id != org_id:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Department not found in this organization"
        )

    return service.update_department(dept_id, dept_data, modified_by=current_user.id)


@router.delete("/{dept_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_department(
    org_id: UUID,
    dept_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_org_admin)
):
    """
    Soft delete department.
    Only org admins can delete departments.
    """
    service = OrgDepartmentService(db)
    dept = service.get_department(dept_id)

    # Verify department belongs to the organization
    if dept.organization_id != org_id:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Department not found in this organization"
        )

    service.delete_department(dept_id)
    return None


@router.post("/{dept_id}/users", response_model=dict)
def assign_users_to_department(
    org_id: UUID,
    dept_id: UUID,
    user_ids: List[UUID] = Body(..., embed=True),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_org_admin_or_dept_admin)
):
    """
    Assign multiple users to a department.
    Only org admins or department admins can assign users.
    """
    service = OrgDepartmentService(db)
    dept = service.get_department(dept_id)

    # Verify department belongs to the organization
    if dept.organization_id != org_id:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Department not found in this organization"
        )

    count = service.assign_users_to_department(dept_id, user_ids)
    return {
        "message": f"Successfully assigned {count} user(s) to department",
        "count": count
    }


@router.get("/{dept_id}/ancestors")
def get_department_ancestors(
    org_id: UUID,
    dept_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_org_member)
):
    """Return the ancestor chain from root down to dept_id as an ordered list."""
    from models import OrgDepartment
    ancestors = []
    current_id = dept_id
    seen = set()
    while current_id:
        if current_id in seen:
            break
        seen.add(current_id)
        dept = db.query(OrgDepartment).filter(
            OrgDepartment.id == current_id,
            OrgDepartment.organization_id == org_id,
        ).first()
        if not dept:
            break
        ancestors.append({
            "id": str(dept.id),
            "name": dept.name,
            "parent_department_id": str(dept.parent_department_id) if dept.parent_department_id else None,
        })
        current_id = dept.parent_department_id
    ancestors.reverse()
    # Assign level based on depth position (1 = root)
    for i, a in enumerate(ancestors):
        a["level"] = i + 1
    return ancestors


@router.get("/{dept_id}/users", response_model=List[UserSchema])
def get_department_users(
    org_id: UUID,
    dept_id: UUID,
    skip: int = 0,
    limit: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_org_member)
):
    """
    Get all users in a department.
    Any organization member can view department users.
    """
    service = OrgDepartmentService(db)
    dept = service.get_department(dept_id)

    # Verify department belongs to the organization
    if dept.organization_id != org_id:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Department not found in this organization"
        )

    return service.get_department_users(dept_id, skip=skip, limit=limit)


# ── Bulk Import ───────────────────────────────────────────────────────────────

HIERARCHY_LEVELS = ["Zone", "Circle", "Division", "Sub Division", "Section", "Substation"]


def _generate_code(name: str) -> str:
    return name.upper().replace(" ", "_")[:50]


@router.get("/bulk-import/schema")
def download_dept_schema(
    org_id: UUID,
    export_existing: bool = False,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_org_admin),
):
    """
    Download Excel template for department bulk import.
    Pass ?export_existing=true to pre-fill the sheet with the current hierarchy
    so you can review, extend, or share it.
    """
    wb = openpyxl.Workbook()

    # ── Sheet 1: Template ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Departments"

    header_fill  = PatternFill("solid", start_color="0F2B6B")
    header_font  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    example_fill = PatternFill("solid", start_color="EFF4FF")
    existing_fill = PatternFill("solid", start_color="F0FFF4")   # light green for existing rows

    for col_idx, level in enumerate(HIERARCHY_LEVELS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=level)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = 22

    ws.row_dimensions[1].height = 22

    if export_existing:
        # ── Export current hierarchy as full paths ────────────────────────────
        all_depts = (
            db.query(OrgDepartment)
            .filter(OrgDepartment.organization_id == org_id, OrgDepartment.is_active == True)
            .all()
        )
        # Build id→dept lookup and walk leaf nodes up to root
        id_map = {d.id: d for d in all_depts}

        def ancestors(dept):
            path, node = [], dept
            while node:
                path.append(node.name)
                node = id_map.get(node.parent_department_id)
            return list(reversed(path))

        # Collect every node as a full path (leaf nodes give complete rows;
        # intermediate-only nodes would be redundant since leaves include them)
        leaf_ids = set(d.id for d in all_depts) - set(
            d.parent_department_id for d in all_depts if d.parent_department_id
        )
        # Also include intermediate nodes that have no children (standalone nodes)
        paths = sorted(ancestors(id_map[lid]) for lid in leaf_ids)

        if paths:
            for row_num, path in enumerate(paths, start=2):
                for col_idx, name in enumerate(path, start=1):
                    cell = ws.cell(row=row_num, column=col_idx, value=name)
                    cell.fill = existing_fill
                    cell.font = Font(name="Arial", size=10)
        else:
            # No departments yet — fall through to blank template with example row
            example_row = ["NORTHERN ZONE", "BANGALORE CIRCLE", "BANGALORE NORTH DIVISION",
                           "HEBBAL SUB DIVISION", "HEBBAL SECTION", "HEBBAL SUBSTATION"]
            for col_idx, val in enumerate(example_row, start=1):
                cell = ws.cell(row=2, column=col_idx, value=val)
                cell.fill = example_fill
                cell.font = Font(name="Arial", size=10, italic=True)
    else:
        # ── Blank template with one example row ──────────────────────────────
        example_row = ["NORTHERN ZONE", "BANGALORE CIRCLE", "BANGALORE NORTH DIVISION",
                       "HEBBAL SUB DIVISION", "HEBBAL SECTION", "HEBBAL SUBSTATION"]
        for col_idx, val in enumerate(example_row, start=1):
            cell = ws.cell(row=2, column=col_idx, value=val)
            cell.fill = example_fill
            cell.font = Font(name="Arial", size=10, italic=True)

    # ── Sheet 2: Instructions ─────────────────────────────────────────────────
    wi = wb.create_sheet("Instructions")
    instructions = [
        ("Cogniwatt — Department Bulk Import Guide", True),
        ("", False),
        ("COLUMNS", True),
        ("Zone → Circle → Division → Sub Division → Section → Substation", False),
        ("Each row represents one complete path from Zone (root) to Substation (leaf).", False),
        ("", False),
        ("RULES", True),
        ("• Every row MUST have at least a Zone value.", False),
        ("• Intermediate levels (Circle, Division …) are optional per row.", False),
        ("• Existing nodes at the same level with the same name will be reused (safe to re-upload).", False),
        ("• Codes are auto-generated (UPPERCASE_WITH_UNDERSCORES).", False),
        ("• Do not modify column headers.", False),
        ("• Maximum 5 000 rows per upload.", False),
        ("", False),
        ("GREEN ROWS (export_existing mode): these are your current departments.", False),
        ("You can add new rows below them or alongside — existing rows are safely skipped on re-import.", False),
        ("", False),
        ("LIGHT BLUE ROW (blank template): example row — delete it before uploading.", False),
    ]
    wi.column_dimensions["A"].width = 90
    for r, (text, bold) in enumerate(instructions, start=1):
        cell = wi.cell(row=r, column=1, value=text)
        cell.font = Font(bold=bold, name="Arial", size=11 if bold else 10)

    filename = "departments_export.xlsx" if export_existing else "department_import_template.xlsx"
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/bulk-import", response_model=BulkDeptImportResult, status_code=200)
async def bulk_import_departments(
    org_id: UUID,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_org_admin),
):
    """Parse Excel and create 6-level department hierarchy."""
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Please upload a valid Excel file (.xlsx or .xls)")

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Cannot read Excel file — it may be corrupt")

    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    if len(rows) > 5000:
        raise HTTPException(status_code=400, detail="Maximum 5 000 rows per upload")

    # dept_map: (level_index, parent_id_or_None, name_lower) → OrgDepartment.id
    dept_map: dict = {}
    total = skipped = failed = 0
    nodes_created = 0
    errors: list[str] = []

    for row_num, row in enumerate(rows, start=2):
        if not any(row):
            continue

        total += 1
        parent_id = None
        row_ok = True
        row_new_nodes = 0

        for level_idx, level_name in enumerate(HIERARCHY_LEVELS):
            raw = row[level_idx] if level_idx < len(row) else None
            name = str(raw).strip() if raw else None
            if not name:
                break  # remaining levels are empty — stop descending

            map_key = (level_idx, str(parent_id), name.lower())
            if map_key in dept_map:
                parent_id = dept_map[map_key]
                continue

            # Check DB for existing node
            existing = db.query(OrgDepartment).filter(
                OrgDepartment.organization_id == org_id,
                OrgDepartment.parent_department_id == parent_id,
                OrgDepartment.name == name,
                OrgDepartment.is_active == True,
            ).first()

            if existing:
                dept_map[map_key] = existing.id
                parent_id = existing.id
                continue

            # Create new node
            try:
                new_dept = OrgDepartment(
                    id=_uuid.uuid4(),
                    organization_id=org_id,
                    name=name,
                    code=_generate_code(name),
                    parent_department_id=parent_id,
                    is_active=True,
                )
                db.add(new_dept)
                db.flush()
                dept_map[map_key] = new_dept.id
                parent_id = new_dept.id
                row_new_nodes += 1
                nodes_created += 1
            except Exception as exc:
                db.rollback()
                errors.append(f"Row {row_num}, {level_name} '{name}': {exc}")
                row_ok = False
                failed += 1
                break

        if not row_ok:
            continue
        if row_new_nodes == 0:
            skipped += 1

    try:
        db.commit()
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Database error during commit: {exc}")

    imported_rows = total - failed - skipped
    return BulkDeptImportResult(total=total, created=imported_rows, skipped=skipped, failed=failed, errors=errors[:50])
