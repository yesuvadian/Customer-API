"""
Data Import Router — upload PDF/Excel → extract records → review → submit as closed TRs.

Endpoints:
  GET    /data-import/categories              — list categories with test types
  POST   /data-import/extract                 — upload file, get extracted records
                                                 (records with no serial are also
                                                 persisted as PendingDataImport rows)
  POST   /data-import/preview-form            — build pre-filled form data for one record
  POST   /data-import/submit                  — batch-create TRs from reviewed records

  GET    /data-import/pending                 — list pending (no-serial) import rows
  PATCH  /data-import/pending/{id}             — edit serial/equipment/form_data on a pending row
  POST   /data-import/pending/{id}/submit      — submit one pending row as a TR (needs serial/equipment)
  DELETE /data-import/pending/{id}             — discard a pending row
"""

from __future__ import annotations

import io
import traceback
from typing import Any, List, Optional
from uuid import UUID

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from auth_utils import get_current_user
from database import get_db
from models import (
    CategoryDetails,
    Equipment,
    OrgDepartment,
    Organization,
    PendingDataImport,
    TestingRequest,
    TestingRequestStatus,
    User,
)
from services.import_extractor_service import (
    IMPORT_CATEGORIES,
    build_form_data,
    extract_records,
    get_template_key,
    get_test_types_for_category,
    resolve_equipment,
)

router = APIRouter(
    prefix="/data-import",
    tags=["data-import"],
    dependencies=[Depends(get_current_user)],
)


# ── Schemas ───────────────────────────────────────────────────────────────────

class ExtractedRecord(BaseModel):
    index: int
    report: dict                        # raw extractor output
    serial_number: Optional[str] = None
    test_date: Optional[str] = None
    sub_station: Optional[str] = None
    equipment_found: bool = False
    equipment_id: Optional[str] = None
    equipment_ueic: Optional[str] = None
    form_data: Optional[dict] = None    # pre-built form values
    warnings: List[str] = []
    # NEW: set when this record had no serial and was persisted as a
    # PendingDataImport row instead of being purely transient.
    pending_id: Optional[str] = None


class PreviewFormRequest(BaseModel):
    report: dict
    test_type_name: str
    equipment_id: Optional[UUID] = None


class SubmitRecord(BaseModel):
    report: dict
    form_data: dict
    test_type_name: str
    equipment_id: Optional[UUID] = None
    organization_id: Optional[UUID] = None
    department_id: Optional[UUID] = None
    overall_result: str = "pass"
    remarks: Optional[str] = None


class SubmitRequest(BaseModel):
    records: List[SubmitRecord]
    category: str = "test"


class SubmitResultItem(BaseModel):
    index: int
    success: bool
    tr_id: Optional[str] = None
    request_number: Optional[str] = None
    serial_number: Optional[str] = None
    error: Optional[str] = None


# NEW: schemas for the pending-import review flow
class PendingRecordResponse(BaseModel):
    id: str
    source_filename: Optional[str] = None
    test_type_name: str
    template_key: Optional[str] = None
    report: dict
    form_data: Optional[dict] = None
    serial_number: Optional[str] = None
    test_date: Optional[str] = None
    sub_station: Optional[str] = None
    equipment_id: Optional[str] = None
    equipment_ueic: Optional[str] = None
    status: str
    warnings: List[str] = []
    cts: Optional[str] = None

    class Config:
        from_attributes = True


class PendingRecordUpdate(BaseModel):
    serial_number: Optional[str] = None
    equipment_id: Optional[UUID] = None
    form_data: Optional[dict] = None


class PendingSubmitRequest(BaseModel):
    equipment_id: Optional[UUID] = None
    organization_id: Optional[UUID] = None
    department_id: Optional[UUID] = None
    overall_result: str = "pass"
    remarks: Optional[str] = None


# ── Shared TR-creation helper ──────────────────────────────────────────────────
# Factored out of submit_import() so /submit and /pending/{id}/submit share
# identical TR-creation logic instead of duplicating it.

def _create_tr_from_record(
    db: Session,
    current_user: User,
    report: dict,
    form_data: dict,
    test_type_name: str,
    equipment: Equipment,
    organization_id: Optional[UUID],
    department_id: Optional[UUID],
    overall_result: str,
    remarks: Optional[str],
    request_category: str = "test",
) -> tuple[Optional[dict], Optional[str]]:
    """
    Create a closed TestingRequest from a resolved record + equipment.
    Returns (result_dict, error_message) — exactly one will be non-None.
    Mirrors the exact steps submit_import() used to perform inline.
    """
    from services.testing_request_service import TestingRequestService
    from services.testing_service import TestingService
    from services.recommendation_service import RecommendationService
    from sqlalchemy.exc import InvalidRequestError
    from sqlalchemy import text
    from datetime import datetime, timezone

    serial = report.get("serial_number", "").strip()
    test_date_str = report.get("test_date")
    sub_station = report.get("sub_station", "")

    test_type = db.query(CategoryDetails).filter(
        CategoryDetails.name == test_type_name
    ).first()
    if not test_type:
        return None, f"Test type '{test_type_name}' not found in database."

    template_key = get_template_key(test_type_name)
    if not template_key:
        return None, f"No template mapped for test type '{test_type_name}'."

    org_id = organization_id or equipment.organization_id
    dept_id = department_id
    if not dept_id and equipment.department:
        dept_id = equipment.department.id
    if not org_id:
        org = db.query(Organization).first()
        org_id = org.id if org else None

    tested_at = None
    if test_date_str:
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d-%b-%Y"):
            try:
                tested_at = datetime.strptime(test_date_str, fmt).replace(tzinfo=timezone.utc)
                break
            except ValueError:
                continue
    if not tested_at:
        tested_at = datetime.now(timezone.utc)

    import_tag = f"IMP-{tested_at.strftime('%Y%m%d')}-{serial}-{template_key[:8].upper()}"
    existing = (
        db.query(TestingRequest)
        .filter(TestingRequest.notes.contains(import_tag))
        .first()
    )
    if existing:
        return {
            "tr_id": str(existing.id),
            "request_number": existing.request_number,
            "serial_number": serial,
            "note": "Already imported (skipped duplicate).",
        }, None

    tr_svc = TestingRequestService(db)
    tst_svc = TestingService(db)
    rec_svc = RecommendationService(db)

    remarks_final = remarks or f"Imported from file. Sub-station: {sub_station}."

    tr = tr_svc.create_request(
        data={
            "title": (
                f"{test_type_name} — "
                f"{equipment.factory_serial_number or equipment.ueic}"
                f" ({tested_at.strftime('%Y-%m-%d')})"
            ),
            "description": f"Imported from file. Serial: {serial}. Sub-station: {sub_station}.",
            "equipment_id": equipment.id,
            "equipment_type_id": equipment.equipment_type_id,
            "test_type_id": test_type.id,
            "organization_id": org_id,
            "department_id": dept_id,
            "assigned_tester_id": current_user.id,
            "requested_date": tested_at,
            "scheduled_start_date": tested_at,
            "request_category": request_category,
            "notes": f"[{import_tag}] Data import by {current_user.email}.",
        },
        originator_id=current_user.id,
    )

    tr.status = TestingRequestStatus.in_progress
    tr.completed_at = tested_at
    db.flush()

    try:
        tst_svc.create_structured_result(
            request_id=tr.id,
            template_key=template_key,
            test_data=form_data,
            overall_result=overall_result,
            remarks=remarks_final,
            tester_id=current_user.id,
        )
    except InvalidRequestError:
        db.expire_all()

    db.execute(
        text(
            "UPDATE public.test_results "
            "SET tested_at = :d WHERE testing_request_id = :rid"
        ),
        {"d": tested_at, "rid": tr.id},
    )
    db.flush()

    recommendation = rec_svc.create_recommendation(
        testing_request_id=tr.id,
        recommendation_type=overall_result,
        summary=f"{overall_result.capitalize()} — {remarks_final}",
        submitted_by=current_user.id,
        next_action="none",
    )

    tr.status = TestingRequestStatus.closed
    recommendation.approval_status = "approved"
    recommendation.approved_by = current_user.id
    recommendation.approved_at = tested_at
    db.commit()

    return {
        "tr_id": str(tr.id),
        "request_number": tr.request_number,
        "serial_number": serial,
    }, None


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.get("/categories")
def list_categories(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Return all import categories with their available test types.

    Filters to only test types that have an active OrgTestTemplate for the org
    (matched via test_type_id). Falls back to global system templates if the
    org has no overrides. If no templates have test_type_id set, shows all.
    """
    from models import OrgTestTemplate
    org_id = current_user.organization_id

    def _active_type_ids(query):
        """Return set of test_type_ids from query, filtering by is_active if column exists."""
        try:
            rows = query.filter(OrgTestTemplate.is_active.is_(True), OrgTestTemplate.test_type_id.isnot(None)).all()
        except Exception:
            rows = query.filter(OrgTestTemplate.test_type_id.isnot(None)).all()
        return {row[0] for row in rows}

    org_ids = _active_type_ids(
        db.query(OrgTestTemplate.test_type_id).filter(OrgTestTemplate.org_id == org_id)
    )
    if org_ids:
        active_type_ids = org_ids
    else:
        active_type_ids = _active_type_ids(
            db.query(OrgTestTemplate.test_type_id).filter(
                OrgTestTemplate.org_id.is_(None),
                OrgTestTemplate.is_system.is_(True),
            )
        )

    result = []
    for key, cfg in IMPORT_CATEGORIES.items():
        test_types = get_test_types_for_category(key, db, active_type_ids=active_type_ids or None)
        result.append({
            "key": key,
            "label": cfg["label"],
            "request_category": cfg["request_category"],
            "test_types": test_types,
        })
    return result


@router.get("/schema")
def download_schema(
    test_type_name: str = Query(...),
    _: User = Depends(get_current_user),
):
    """Download a blank Excel import template for any test type."""
    from test_templates import get_template_for_test_type
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    tpl = get_template_for_test_type(test_type_name)

    # Fixed identity columns always present (needed for equipment matching)
    fixed_cols = [
        ("sub_station",   "Sub Station",   "Name of substation / location"),
        ("serial_number", "Serial Number", "Equipment serial number (used to match equipment)"),
        ("test_date",     "Date of Test",  "YYYY-MM-DD format"),
    ]

    # Dynamic scalar columns + collect table field definitions
    dyn_cols: list[tuple[str, str, str]] = []
    table_fields: list[dict] = []
    if tpl:
        for sec in tpl.get("sections", []):
            sec_title = sec.get("title", "")
            for f in sec.get("fields", []):
                ftype = f.get("type", "text")
                if ftype == "table":
                    table_fields.append(f)
                    continue
                if ftype in ("calculated", "readonly"):
                    continue
                if f.get("import_skip"):
                    continue
                key   = f.get("key", "")
                label = f.get("label", key)
                hint  = f.get("unit", "") or f.get("hint", "") or sec_title
                if key:
                    dyn_cols.append((key, label, hint))

    all_cols = fixed_cols + dyn_cols

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Import Data"

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    hint_fill   = PatternFill("solid", fgColor="EFF6FF")
    table_fill  = PatternFill("solid", fgColor="1E3A5F")
    row_fill    = PatternFill("solid", fgColor="F8FAFF")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    hint_font   = Font(italic=True, color="6B7280", size=9)
    key_font    = Font(color="FFFFFF", size=1)

    for ci, (key, label, hint) in enumerate(all_cols, start=1):
        # Row 1: visible header label
        h = ws.cell(row=1, column=ci, value=label)
        h.font      = header_font
        h.fill      = header_fill
        h.alignment = Alignment(horizontal="center", wrap_text=True)

        # Row 2: machine key (hidden — white on white)
        k = ws.cell(row=2, column=ci, value=key)
        k.font = key_font

        # Row 3: hint
        hnt = ws.cell(row=3, column=ci, value=hint)
        hnt.font      = hint_font
        hnt.fill      = hint_fill
        hnt.alignment = Alignment(wrap_text=True)

        # Column width
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(18, len(label) + 4)

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 0   # hide key row
    ws.row_dimensions[3].height = 20
    ws.freeze_panes = "A4"

    # ── One sheet per table field ──────────────────────────────────────────────
    for tf in table_fields:
        tbl_key   = tf.get("key", "table")
        tbl_label = tf.get("label", tbl_key)

        # Only editable, non-calculated columns
        editable_cols = [
            c for c in tf.get("columns", [])
            if c.get("type") not in ("calculated",) and not c.get("read_only")
        ]
        if not editable_cols:
            continue

        # Sheet name: strip Excel-invalid chars, truncate to 31 chars
        import re as _re
        sheet_name = _re.sub(r'[:\\/?*\[\]]', '', tbl_label)[:31]
        tws = wb.create_sheet(title=sheet_name)

        # Row 1: visible label  Row 2: machine key (hidden)  Row 3: unit hint
        for ci, col in enumerate(editable_cols, start=1):
            col_key   = col.get("key", f"col{ci}")
            col_label = col.get("label", col_key)
            col_hint  = col.get("unit", "") or col.get("placeholder", "")

            h = tws.cell(row=1, column=ci, value=col_label)
            h.font      = header_font
            h.fill      = table_fill
            h.alignment = Alignment(horizontal="center", wrap_text=True)

            k = tws.cell(row=2, column=ci, value=col_key)
            k.font = key_font

            hnt = tws.cell(row=3, column=ci, value=col_hint)
            hnt.font      = hint_font
            hnt.fill      = hint_fill
            hnt.alignment = Alignment(wrap_text=True)

            tws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(18, len(col_label) + 4)

        # Row 1 in col 0 (A): store table field key as a hidden marker
        meta_cell = tws.cell(row=2, column=len(editable_cols) + 1, value=f"__table_key__={tbl_key}")
        meta_cell.font = key_font

        tws.row_dimensions[1].height = 28
        tws.row_dimensions[2].height = 0
        tws.row_dimensions[3].height = 18
        tws.freeze_panes = "A4"

        # Pre-fill default rows (starting at row 4)
        default_rows = tf.get("default_rows", [])
        for ri, drow in enumerate(default_rows, start=4):
            for ci, col in enumerate(editable_cols, start=1):
                val = drow.get(col.get("key", ""), "")
                cell = tws.cell(row=ri, column=ci, value=val if val != "" else None)
                if ri % 2 == 0:
                    cell.fill = row_fill
        # Add 5 blank rows after defaults for extra data
        start_blank = max(4, 4 + len(default_rows))
        for ri in range(start_blank, start_blank + 5):
            for ci in range(1, len(editable_cols) + 1):
                cell = tws.cell(row=ri, column=ci, value=None)
                if ri % 2 == 0:
                    cell.fill = row_fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = test_type_name.replace(" ", "_").replace("/", "-")
    filename = f"import_schema_{safe_name}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/extract", status_code=status.HTTP_200_OK)
async def extract_file(
    file: UploadFile = File(...),
    test_type_name: str = Form(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Upload a PDF or Excel file and extract all records from it.

    Returns a list of ExtractedRecord objects — one per report found in the file.
    Each record includes the raw report dict, auto-matched equipment info,
    pre-built form_data, and any extraction warnings.

    NEW: records with no serial_number are ALSO persisted as PendingDataImport
    rows (status="pending") so they survive beyond this request/response and
    can be reviewed/completed later via /data-import/pending. Records that DO
    have a serial are returned exactly as before — no DB row is created for
    them here; they still flow through the existing /submit endpoint.
    """
    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")

    records_raw, global_warnings = extract_records(
        file_bytes=file_bytes,
        filename=file.filename or "upload",
        test_type_name=test_type_name,
    )

    if global_warnings and not records_raw:
        raise HTTPException(status_code=422, detail="; ".join(global_warnings))

    template_key = get_template_key(test_type_name)

    results: list[dict] = []
    for idx, report in enumerate(records_raw):
        serial = report.get("serial_number") or ""
        rec_warnings: list[str] = list(global_warnings)

        eq = resolve_equipment(serial or None, None, db)
        eq_found = eq is not None

        if serial and not eq_found:
            rec_warnings.append(
                f"Equipment with serial '{serial}' not found in asset register. "
                "Please select equipment manually before submitting."
            )
        elif not serial:
            rec_warnings.append(
                "No serial number could be extracted from this report. "
                "It has been saved for manual review — add a serial or "
                "select equipment before submitting."
            )

        # Build form_data using the template builder (mirrors seed scripts)
        try:
            form_data = build_form_data(report, test_type_name, eq)
        except Exception:
            form_data = report
            rec_warnings.append("Form template could not be applied — raw values shown.")

        pending_id = None
        if not serial:
            # NEW: persist so this extraction isn't lost once the response
            # is returned. Does not touch TestingRequest/Equipment at all.
            pending_row = PendingDataImport(
                organization_id=current_user.organization_id,
                source_filename=file.filename,
                test_type_name=test_type_name,
                template_key=template_key,
                report=report,
                form_data=form_data,
                serial_number=serial or None,
                test_date=report.get("test_date") or None,
                sub_station=report.get("sub_station") or None,
                equipment_id=eq.id if eq else None,
                status="pending",
                warnings=rec_warnings,
                created_by=current_user.id,
            )
            db.add(pending_row)
            db.commit()
            db.refresh(pending_row)
            pending_id = str(pending_row.id)

        results.append(
            ExtractedRecord(
                index=idx,
                report=report,
                serial_number=serial or None,
                test_date=report.get("test_date") or None,
                sub_station=report.get("sub_station") or None,
                equipment_found=eq_found,
                equipment_id=str(eq.id) if eq else None,
                equipment_ueic=eq.ueic if eq else None,
                form_data=form_data,
                warnings=rec_warnings,
                pending_id=pending_id,
            ).model_dump()
        )

    return {
        "total": len(results),
        "filename": file.filename,
        "test_type_name": test_type_name,
        "records": results,
    }


@router.post("/preview-form")
def preview_form(
    body: PreviewFormRequest,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Build (or rebuild) the pre-filled form_data for a single report dict."""
    eq = None
    if body.equipment_id:
        eq = db.query(Equipment).filter(Equipment.id == body.equipment_id).first()
        if not eq:
            raise HTTPException(status_code=404, detail="Equipment not found.")

    if not eq:
        serial = body.report.get("serial_number")
        eq = resolve_equipment(serial, None, db)

    try:
        form_data = build_form_data(body.report, body.test_type_name, eq)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Form build failed: {e}")

    template_key = get_template_key(body.test_type_name)
    equipment_info = None
    if eq:
        dept = getattr(eq, "department", None)
        equipment_info = {
            "id": str(eq.id),
            "ueic": eq.ueic,
            "serial_number": eq.factory_serial_number,
            "manufacturer": getattr(eq, "manufacturer", None),
            "department_id": str(dept.id) if dept else None,
            "department_name": dept.name if dept else None,
            "organization_id": str(eq.organization_id) if eq.organization_id else None,
        }

    return {
        "form_data": form_data,
        "template_key": template_key,
        "equipment_info": equipment_info,
    }


@router.post("/submit")
def submit_import(
    body: SubmitRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Batch-submit reviewed import records as closed TestingRequests.

    Each record goes through the full TR lifecycle:
      draft → in_progress → test result created → recommendation → closed

    Failures on individual records are isolated and reported; other records continue.
    """
    results: list[dict] = []

    for idx, rec in enumerate(body.records):
        serial = rec.report.get("serial_number", "").strip()
        try:
            eq = resolve_equipment(serial or None, rec.equipment_id, db)
            if not eq:
                results.append(SubmitResultItem(
                    index=idx, success=False, serial_number=serial,
                    error=f"Equipment '{serial}' not found. Assign equipment_id manually."
                ).model_dump())
                continue

            form_data = rec.form_data
            if not form_data:
                form_data = build_form_data(rec.report, rec.test_type_name, eq)

            outcome, err = _create_tr_from_record(
                db=db,
                current_user=current_user,
                report=rec.report,
                form_data=form_data,
                test_type_name=rec.test_type_name,
                equipment=eq,
                organization_id=rec.organization_id,
                department_id=rec.department_id,
                overall_result=rec.overall_result,
                remarks=rec.remarks,
                request_category=body.category,
            )

            if err:
                results.append(SubmitResultItem(
                    index=idx, success=False, serial_number=serial, error=err
                ).model_dump())
                continue

            results.append(SubmitResultItem(
                index=idx,
                success=True,
                tr_id=outcome["tr_id"],
                request_number=outcome["request_number"],
                serial_number=outcome["serial_number"],
                error=outcome.get("note"),
            ).model_dump())

        except Exception as exc:
            db.rollback()
            results.append(SubmitResultItem(
                index=idx,
                success=False,
                serial_number=serial,
                error=str(exc),
            ).model_dump())

    total_ok = sum(1 for r in results if r["success"])
    return {
        "submitted": len(body.records),
        "created": total_ok,
        "failed": len(body.records) - total_ok,
        "results": results,
    }


# ── Pending import review endpoints ────────────────────────────────────────────

def _pending_to_response(row: PendingDataImport, db: Session) -> dict:
    eq_ueic = None
    if row.equipment_id:
        eq = db.query(Equipment).filter(Equipment.id == row.equipment_id).first()
        eq_ueic = eq.ueic if eq else None
    return PendingRecordResponse(
        id=str(row.id),
        source_filename=row.source_filename,
        test_type_name=row.test_type_name,
        template_key=row.template_key,
        report=row.report,
        form_data=row.form_data,
        serial_number=row.serial_number,
        test_date=row.test_date,
        sub_station=row.sub_station,
        equipment_id=str(row.equipment_id) if row.equipment_id else None,
        equipment_ueic=eq_ueic,
        status=row.status,
        warnings=row.warnings or [],
        cts=row.cts.isoformat() if row.cts else None,
    ).model_dump()


@router.get("/pending")
def list_pending(
    status_filter: str = "pending",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """List pending (unlinked / no-serial) import records for review."""
    query = db.query(PendingDataImport)
    if current_user.organization_id:
        query = query.filter(
            PendingDataImport.organization_id == current_user.organization_id
        )
    if status_filter:
        query = query.filter(PendingDataImport.status == status_filter)

    rows = query.order_by(PendingDataImport.cts.desc()).all()
    return {
        "total": len(rows),
        "records": [_pending_to_response(r, db) for r in rows],
    }


@router.patch("/pending/{pending_id}")
def update_pending(
    pending_id: UUID,
    body: PendingRecordUpdate,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Edit serial number / equipment / form_data on a pending row before submitting."""
    row = db.query(PendingDataImport).filter(PendingDataImport.id == pending_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Pending record not found.")
    if row.status != "pending":
        raise HTTPException(status_code=400, detail=f"Record is already '{row.status}'.")

    if body.serial_number is not None:
        row.serial_number = body.serial_number
        row.report = {**row.report, "serial_number": body.serial_number}

    if body.equipment_id is not None:
        eq = db.query(Equipment).filter(Equipment.id == body.equipment_id).first()
        if not eq:
            raise HTTPException(status_code=404, detail="Equipment not found.")
        row.equipment_id = eq.id
        if not body.serial_number and eq.factory_serial_number:
            row.serial_number = eq.factory_serial_number
            row.report = {**row.report, "serial_number": eq.factory_serial_number}

    if body.form_data is not None:
        row.form_data = body.form_data

    db.commit()
    db.refresh(row)
    return _pending_to_response(row, db)


@router.post("/pending/{pending_id}/submit")
def submit_pending(
    pending_id: UUID,
    body: PendingSubmitRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Submit one pending record as a closed TestingRequest, using the exact
    same TR-creation path as /data-import/submit. Requires a serial number
    (on the row or supplied here) that resolves to an Equipment, or an
    explicit equipment_id.
    """
    row = db.query(PendingDataImport).filter(PendingDataImport.id == pending_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Pending record not found.")
    if row.status != "pending":
        raise HTTPException(status_code=400, detail=f"Record is already '{row.status}'.")

    eq = resolve_equipment(row.serial_number or None, body.equipment_id or row.equipment_id, db)
    if not eq:
        raise HTTPException(
            status_code=422,
            detail="No matching equipment. Update the record with a valid serial "
                   "or equipment_id before submitting.",
        )

    form_data = row.form_data
    if not form_data:
        form_data = build_form_data(row.report, row.test_type_name, eq)

    outcome, err = _create_tr_from_record(
        db=db,
        current_user=current_user,
        report=row.report,
        form_data=form_data,
        test_type_name=row.test_type_name,
        equipment=eq,
        organization_id=body.organization_id,
        department_id=body.department_id,
        overall_result=body.overall_result,
        remarks=body.remarks,
    )

    if err:
        raise HTTPException(status_code=422, detail=err)

    row.status = "submitted"
    row.equipment_id = eq.id
    db.commit()

    return outcome


@router.delete("/pending/{pending_id}")
def discard_pending(
    pending_id: UUID,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Discard a pending record without submitting it."""
    row = db.query(PendingDataImport).filter(PendingDataImport.id == pending_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Pending record not found.")
    row.status = "discarded"
    db.commit()
    return {"id": str(row.id), "status": "discarded"}