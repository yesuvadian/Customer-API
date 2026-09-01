"""
Comparison View API Router — "AI-Assisted Comparison View"
================================================================
Multi-test-type comparison dashboard for ANY equipment type and ANY of its
test types. Equipment Type is selected first; the test types offered are
whichever templates are actually assigned to that equipment type (via
OrgTestTemplate -> CategoryDetails -> CategoryMaster), sliced further by
department (with hierarchical drill-down) and year range.

Each selected test type's counts are always kept separate — never pooled
into one blended figure, since different templates use different
thresholds and severity scales (a "Critical" reading on one test type is
not equivalent to a "Critical" reading on another).

  GET /comparison-view/filters               <- all equipment types with >=1 test template
  GET /comparison-view/test-types             <- test types (+ year bounds) for one equipment type
  GET /comparison-view/dashboard              <- dept cards + one panel per selected template_key
  GET /comparison-view/export                 <- the current dashboard view as an .xlsx workbook
  GET /comparison-view/equipment              <- equipment list for one (template_key, year) slice
  GET /comparison-view/equipment/{id}/tests   <- test instances for that equipment, same slice

Step 4 of the drill-down ("Test Details" with parameter-level data) is not
duplicated here: it reuses the existing GET /analytics/test-results/{id}
and /raw endpoints, linked from the tests list above.
"""

import io
import re
import uuid
import logging
from datetime import date, datetime, timezone
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, Response
from sqlalchemy import func, or_
from sqlalchemy.orm import Session

from database import get_vendor_db
from auth_utils import get_current_user
from models import (
    CategoryDetails,
    CategoryMaster,
    Equipment,
    HierarchyAnalytics,
    OrgDepartment,
    OrgTestTemplate,
    TestAnalytics,
    TestRequestSchedule,
)
from routers.analytics import _collect_department_ids
from utils.common_service import get_user_dept_scope

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/comparison-view", tags=["Comparison View"])

# TestAnalytics.condition_summary ("Good/Fair/Poor") -> this dashboard's Normal/Warning/Critical buckets.
_CONDITION_TO_BUCKET = {"Good": "normal", "Fair": "warning", "Poor": "critical"}


def _org_id(user) -> Optional[uuid.UUID]:
    return user.get("organization_id") if isinstance(user, dict) else getattr(user, "organization_id", None)


def _dept_scope(user, org_id, db: Session) -> tuple[bool, Optional[uuid.UUID]]:
    """(is_admin, user_dept_id) for the requesting user.

    Same rule already used by /testing_requests/department_hierarchy and the
    equipment endpoints, applied here too so a dept-scoped (e.g. substation
    -level) user is confined to their own subtree by default — not just once
    they explicitly drill into a department_id, which otherwise leaves the
    dashboard defaulting to every top-level zone in the org.
    """
    user_id = user.get("id") if isinstance(user, dict) else getattr(user, "id", None)
    if user_id is None:
        return True, None
    return get_user_dept_scope(db, user_id, org_id)


def _overdue_counts_for_scope(template_keys: list[str], root_dept_ids: list[uuid.UUID], db: Session) -> dict:
    """Overdue-for-retest count per root department (subtree rollup) AND per
    template_key within it, scoped to `template_keys`. Reuses the same
    "next_run_date in the past" definition of overdue already established in
    TestScheduleDashboardService._compute_kpis, but scoped to just these
    template keys instead of every schedule for the equipment.

    Returns {dept_id_str: {template_key: overdue_count}} — per-type, never
    pre-summed, so callers can both show the breakdown and derive the pooled
    total by summing it themselves (one source of truth, no separate figure
    that could drift out of sync with the breakdown)."""
    if not root_dept_ids or not template_keys:
        return {}

    tk_to_ttid = dict(
        db.query(OrgTestTemplate.template_key, OrgTestTemplate.test_type_id)
        .filter(
            OrgTestTemplate.template_key.in_(template_keys),
            OrgTestTemplate.test_type_id.isnot(None),
        )
        .all()
    )
    ttid_to_tk = {v: k for k, v in tk_to_ttid.items()}
    test_type_ids = list(ttid_to_tk.keys())
    if not test_type_ids:
        return {}

    today = date.today()
    counts: dict = {}
    for root_id in root_dept_ids:
        subtree_ids = _collect_department_ids(root_id, db)
        rows = (
            db.query(TestRequestSchedule.next_run_date, TestRequestSchedule.test_type_id)
            .join(Equipment, Equipment.id == TestRequestSchedule.equipment_id)
            .filter(
                TestRequestSchedule.test_type_id.in_(test_type_ids),
                TestRequestSchedule.is_active == True,  # noqa: E712
                TestRequestSchedule.is_deleted == False,  # noqa: E712
                Equipment.department_id.in_(subtree_ids),
            )
            .all()
        )
        per_tk: dict = {}
        for next_run, ttid in rows:
            if next_run is None:
                continue
            next_date = next_run.date() if isinstance(next_run, datetime) else next_run
            if (next_date - today).days <= 0:
                tk = ttid_to_tk.get(ttid)
                if tk:
                    per_tk[tk] = per_tk.get(tk, 0) + 1
        counts[str(root_id)] = per_tk
    return counts


@router.get("/filters", summary="All equipment types that have at least one test template assigned")
def get_filters(
    db: Session = Depends(get_vendor_db),
    user: dict = Depends(get_current_user),
):
    org_id = _org_id(user)

    type_rows = (
        db.query(CategoryMaster.id, CategoryMaster.name)
        .join(CategoryDetails, CategoryDetails.category_master_id == CategoryMaster.id)
        .join(OrgTestTemplate, OrgTestTemplate.test_type_id == CategoryDetails.id)
        .filter(or_(OrgTestTemplate.org_id == org_id, OrgTestTemplate.org_id.is_(None)))
        .distinct()
        .all()
    )
    type_ids = [r[0] for r in type_rows]

    # Order by how much real equipment each type actually has, so the pill
    # list (and whichever type the frontend auto-selects first) leads with
    # fleet types that have data, not administrative/audit pseudo-categories
    # like "Annual Audit Categories" that happen to sort first alphabetically.
    counts: dict = {}
    if type_ids:
        count_q = db.query(Equipment.equipment_type_id, func.count(Equipment.id)).filter(
            Equipment.equipment_type_id.in_(type_ids),
            Equipment.status != "retired",
        )
        if org_id:
            count_q = count_q.filter(Equipment.organization_id == org_id)
        counts = dict(count_q.group_by(Equipment.equipment_type_id).all())

    equipment_types = sorted(
        [{"id": r[0], "name": r[1], "equipment_count": counts.get(r[0], 0)} for r in type_rows],
        key=lambda t: (-t["equipment_count"], t["name"]),
    )

    return {"equipment_types": equipment_types}


@router.get("/test-types", summary="Test types (+ year bounds) available for one equipment type")
def get_test_types(
    equipment_type_id: int = Query(...),
    db: Session = Depends(get_vendor_db),
    user: dict = Depends(get_current_user),
):
    org_id = _org_id(user)

    # Test types assigned to this equipment type, real display names sourced from
    # each template's own stored name (template_data['name']) — never hardcoded
    # on the frontend. Org-specific override wins over the global default.
    tmpl_rows = (
        db.query(OrgTestTemplate.template_key, OrgTestTemplate.org_id, OrgTestTemplate.template_data)
        .join(CategoryDetails, CategoryDetails.id == OrgTestTemplate.test_type_id)
        .filter(
            CategoryDetails.category_master_id == equipment_type_id,
            or_(OrgTestTemplate.org_id == org_id, OrgTestTemplate.org_id.is_(None)),
        )
        .all()
    )
    label_by_key: dict = {}
    for key, row_org_id, data in tmpl_rows:
        name = (data or {}).get("name") or key
        is_org_specific = org_id is not None and row_org_id == org_id
        if key not in label_by_key or is_org_specific:
            label_by_key[key] = name
    template_types = [
        {"key": k, "label": v}
        for k, v in sorted(label_by_key.items(), key=lambda kv: kv[1])
    ]
    template_keys = list(label_by_key.keys())

    year_min = year_max = None
    if template_keys:
        year_bounds_q = db.query(
            func.min(func.extract("year", TestAnalytics.tested_at)),
            func.max(func.extract("year", TestAnalytics.tested_at)),
        ).filter(
            TestAnalytics.template_key.in_(template_keys),
            TestAnalytics.tested_at.isnot(None),
        )
        if org_id:
            year_bounds_q = year_bounds_q.filter(TestAnalytics.organization_id == org_id)
        year_bounds = year_bounds_q.first()
        year_min = int(year_bounds[0]) if year_bounds and year_bounds[0] is not None else None
        year_max = int(year_bounds[1]) if year_bounds and year_bounds[1] is not None else None

    return {
        "equipment_type_id": equipment_type_id,
        "template_types": template_types,
        "year_min": year_min,
        "year_max": year_max,
    }


@router.get("/dashboard", summary="Department cards + one panel per selected test type")
def get_dashboard(
    department_id: Optional[uuid.UUID] = Query(None),
    equipment_type_id: Optional[int] = Query(None),
    template_keys: list[str] = Query(default=[]),
    year_from: Optional[int] = Query(None),
    year_to: Optional[int] = Query(None),
    db: Session = Depends(get_vendor_db),
    user: dict = Depends(get_current_user),
):
    org_id = _org_id(user)
    if not template_keys:
        template_keys = []
    is_admin, user_dept_id = _dept_scope(user, org_id, db)
    # An explicit department_id (the user drilled into a card) always wins;
    # otherwise a dept-scoped user is confined to their own subtree so the
    # dashboard never defaults to org-wide data for them.
    effective_department_id = department_id or (None if is_admin else user_dept_id)
    dept_ids = _collect_department_ids(effective_department_id, db) if effective_department_id else None

    # ── Equipment in scope ──────────────────────────────────────────────────
    eq_q = db.query(Equipment.id, Equipment.department_id).filter(Equipment.status != "retired")
    if org_id:
        eq_q = eq_q.filter(Equipment.organization_id == org_id)
    if equipment_type_id:
        eq_q = eq_q.filter(Equipment.equipment_type_id == equipment_type_id)
    if dept_ids:
        eq_q = eq_q.filter(Equipment.department_id.in_(dept_ids))
    eq_dept_map = {row[0]: row[1] for row in eq_q.all()}
    eq_ids = list(eq_dept_map.keys())

    # ── TestAnalytics rows in scope (one query, aggregated in Python below) ──
    # Departments are still listed below even when this comes back empty (no
    # equipment of this type, or no test data yet for these template_keys) —
    # an empty result set here must never hide the department carousel, or
    # drill-down/navigation becomes impossible for a fresh selection.
    ta_rows = []
    if eq_ids and template_keys:
        ta_q = db.query(TestAnalytics).filter(
            TestAnalytics.equipment_id.in_(eq_ids),
            TestAnalytics.template_key.in_(template_keys),
            TestAnalytics.tested_at.isnot(None),
        )
        if org_id:
            ta_q = ta_q.filter(TestAnalytics.organization_id == org_id)
        if year_from:
            ta_q = ta_q.filter(func.extract("year", TestAnalytics.tested_at) >= year_from)
        if year_to:
            ta_q = ta_q.filter(func.extract("year", TestAnalytics.tested_at) <= year_to)
        ta_rows = ta_q.all()

    # ── Per-template-key panels — never merged across template_keys ─────────
    panels = []
    for tk in template_keys:
        rows = [r for r in ta_rows if r.template_key == tk]
        total = len(rows)
        bucket_counts = {"normal": 0, "warning": 0, "critical": 0}
        tests_by_year: dict = {}
        condition_by_year: dict = {}
        condition_by_month: dict = {}
        year_scores: dict = {}
        for r in rows:
            yr = str(r.tested_at.year)
            mk = f"{r.tested_at.year:04d}-{r.tested_at.month:02d}"
            bucket = _CONDITION_TO_BUCKET.get(r.condition_summary)
            if bucket:
                bucket_counts[bucket] += 1
            tests_by_year[yr] = tests_by_year.get(yr, 0) + 1
            yc = condition_by_year.setdefault(yr, {"Good": 0, "Fair": 0, "Poor": 0})
            if r.condition_summary in yc:
                yc[r.condition_summary] += 1
            # Month-level breakdown — lets the "Tests by year" bar be stacked
            # by month instead of one flat average color, so a critical month
            # is visible even when most of the year was fine.
            mc = condition_by_month.setdefault(mk, {"Good": 0, "Fair": 0, "Poor": 0})
            if r.condition_summary in mc:
                mc[r.condition_summary] += 1
            if r.health_score is not None:
                year_scores.setdefault(yr, []).append(float(r.health_score))

        def _pct(n, _total=total):
            return round(n / _total * 100, 1) if _total else 0.0

        yoy_trend = [
            {"year": yr, "avg_health_score": round(sum(v) / len(v), 1)}
            for yr, v in sorted(year_scores.items())
        ]

        panels.append({
            "template_key": tk,
            "total_tests": total,
            "normal":   {"count": bucket_counts["normal"],   "pct": _pct(bucket_counts["normal"])},
            "warning":  {"count": bucket_counts["warning"],  "pct": _pct(bucket_counts["warning"])},
            "critical": {"count": bucket_counts["critical"], "pct": _pct(bucket_counts["critical"])},
            "tests_by_year": dict(sorted(tests_by_year.items())),
            "condition_by_year": dict(sorted(condition_by_year.items())),
            "condition_by_month": dict(sorted(condition_by_month.items())),
            "yoy_trend": yoy_trend,
        })

    # ── Department cards (immediate children — subtree rollup each) ─────────
    if department_id:
        child_depts = (
            db.query(OrgDepartment)
            .filter(OrgDepartment.parent_department_id == department_id)
            .all()
        )
    elif not is_admin and user_dept_id:
        # Dept-scoped user's root view — just their own department (mirrors
        # the root_id behavior in /testing_requests/department_hierarchy),
        # never the org's full list of top-level zones they can't access.
        child_depts = (
            db.query(OrgDepartment)
            .filter(OrgDepartment.id == user_dept_id)
            .all()
        )
    elif org_id:
        child_depts = (
            db.query(OrgDepartment)
            .filter(
                OrgDepartment.parent_department_id.is_(None),
                OrgDepartment.organization_id == org_id,
            )
            .all()
        )
    else:
        child_depts = []

    ha_map = {}
    if child_depts:
        ha_rows = (
            db.query(HierarchyAnalytics)
            .filter(HierarchyAnalytics.department_id.in_([d.id for d in child_depts]))
            .all()
        )
        ha_map = {h.department_id: h for h in ha_rows}

    overdue_by_dept = _overdue_counts_for_scope(template_keys, [d.id for d in child_depts], db)

    # Latest in-scope TestAnalytics row per (equipment, template_key) — kept
    # separate per type (not just per equipment) so a department's Critical
    # count can be broken down by test type without one type's more-recent
    # test masking another type's still-current Poor result.
    latest_by_eq_tk: dict = {}
    for r in ta_rows:
        key = (r.equipment_id, r.template_key)
        cur = latest_by_eq_tk.get(key)
        if cur is None or (r.tested_at and cur.tested_at and r.tested_at > cur.tested_at):
            latest_by_eq_tk[key] = r

    dept_cards = []
    for d in child_depts:
        subtree_ids = _collect_department_ids(d.id, db)
        subtree_eq_ids = {eq_id for eq_id, dep_id in eq_dept_map.items() if dep_id in subtree_ids}
        overdue_for_dept = overdue_by_dept.get(str(d.id), {})

        # Per-type breakdown first — the pooled top-level numbers are always
        # the sum of this, never a separately computed figure, so they can
        # never drift out of sync with what the breakdown says they should be.
        breakdown = []
        for tk in template_keys:
            tk_total = sum(
                1 for r in ta_rows
                if r.equipment_id in subtree_eq_ids and r.template_key == tk
            )
            tk_critical = sum(
                1 for (eq_id, rtk), r in latest_by_eq_tk.items()
                if rtk == tk and eq_id in subtree_eq_ids and r.condition_summary == "Poor"
            )
            breakdown.append({
                "template_key": tk,
                "total_tests": tk_total,
                "critical_count": tk_critical,
                "overdue_count": overdue_for_dept.get(tk, 0),
            })

        ha = ha_map.get(d.id)
        dept_cards.append({
            "department_id": str(d.id),
            "department_name": d.name,
            "level_type": ha.level_type if ha else None,
            "critical_count": sum(b["critical_count"] for b in breakdown),
            "overdue_count": sum(b["overdue_count"] for b in breakdown),
            "total_tests": sum(b["total_tests"] for b in breakdown),
            "breakdown": breakdown,
        })

    dept_cards.sort(key=lambda c: c["total_tests"], reverse=True)

    return {
        "department_id": str(department_id) if department_id else None,
        "dept_cards": dept_cards,
        "panels": panels,
    }


# ── Excel styling — mirrors services/reporting_service.py's _render_excel ───
_HDR_FILL_COLOR = "1565C0"
_ALT_FILL_COLOR = "EBF2FB"


def _write_export_sheet(ws, title: str, headers: list[str], rows: list[list]) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    hdr_fill  = PatternFill("solid", fgColor=_HDR_FILL_COLOR)
    hdr_font  = Font(bold=True, color="FFFFFF", size=10)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin      = Side(style="thin", color="CCCCCC")
    bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill  = PatternFill("solid", fgColor=_ALT_FILL_COLOR)

    cols = max(len(headers), 1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(bold=True, size=13, color=_HDR_FILL_COLOR)
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
    ws.cell(row=2, column=1, value=f"Generated: {ts}")
    ws.row_dimensions[2].height = 16

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = hdr_align
        c.border = bdr
    ws.row_dimensions[3].height = 20

    if not rows:
        ws.cell(row=4, column=1, value="No data for the selected filters.")
    for ri, row in enumerate(rows, 4):
        fill = alt_fill if ri % 2 == 0 else None
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            if fill:
                cell.fill = fill
            cell.border = bdr

    for ci, h in enumerate(headers, 1):
        col_vals = [str(r[ci - 1]) if ci - 1 < len(r) and r[ci - 1] is not None else "" for r in rows]
        max_len = max(len(h), max((len(v) for v in col_vals), default=0))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 50)

    ws.freeze_panes = "A4"


@router.get("/export", summary="Export the current comparison view as an Excel workbook")
def export_dashboard(
    department_id: Optional[uuid.UUID] = Query(None),
    equipment_type_id: Optional[int] = Query(None),
    template_keys: list[str] = Query(default=[]),
    year_from: Optional[int] = Query(None),
    year_to: Optional[int] = Query(None),
    db: Session = Depends(get_vendor_db),
    user: dict = Depends(get_current_user),
):
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed. Run: pip install openpyxl")

    # Reuse the exact same aggregation the on-screen dashboard renders — the
    # export can never drift from what's actually displayed.
    data = get_dashboard(
        department_id=department_id, equipment_type_id=equipment_type_id,
        template_keys=template_keys, year_from=year_from, year_to=year_to,
        db=db, user=user,
    )

    label_by_key: dict = {}
    if equipment_type_id:
        tt = get_test_types(equipment_type_id=equipment_type_id, db=db, user=user)
        label_by_key = {t["key"]: t["label"] for t in tt["template_types"]}

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Summary"
    summary_rows = [
        [
            label_by_key.get(p["template_key"], p["template_key"]),
            p["total_tests"],
            p["normal"]["count"], f'{p["normal"]["pct"]}%',
            p["warning"]["count"], f'{p["warning"]["pct"]}%',
            p["critical"]["count"], f'{p["critical"]["pct"]}%',
        ]
        for p in data["panels"]
    ]
    _write_export_sheet(
        ws, "Comparison Summary",
        ["Test Type", "Total Tests", "Normal", "Normal %", "Warning", "Warning %", "Critical", "Critical %"],
        summary_rows,
    )

    ws_dept = wb.create_sheet("Departments")
    dept_rows = [
        [d["department_name"], d["critical_count"], d["overdue_count"], d["total_tests"]]
        for d in data["dept_cards"]
    ]
    _write_export_sheet(ws_dept, "Departments", ["Department", "Critical", "Overdue", "Total Tests"], dept_rows)

    used_sheet_names = {"Summary", "Departments"}
    for p in data["panels"]:
        name = label_by_key.get(p["template_key"], p["template_key"])
        safe = re.sub(r'[\/\\\?\*\[\]\:]', '-', name)[:28] or p["template_key"][:28]
        unique = safe
        suffix = 2
        while unique in used_sheet_names:
            unique = f"{safe[:26]}-{suffix}"
            suffix += 1
        used_sheet_names.add(unique)

        ws_tk = wb.create_sheet(unique)
        yoy_by_year = {pt["year"]: pt["avg_health_score"] for pt in p["yoy_trend"]}
        years = sorted(p["tests_by_year"].keys())
        rows = []
        for yr in years:
            cond = p["condition_by_year"].get(yr, {})
            rows.append([
                yr, p["tests_by_year"].get(yr, 0),
                cond.get("Good", 0), cond.get("Fair", 0), cond.get("Poor", 0),
                yoy_by_year.get(yr),
            ])
        _write_export_sheet(
            ws_tk, name,
            ["Year", "Total Tests", "Normal", "Warning", "Critical", "Avg Health Score"],
            rows,
        )

    buf = io.BytesIO()
    wb.save(buf)
    filename = f"comparison_view_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/equipment", summary="Equipment list for one (template_key, year) slice")
def get_equipment_for_slice(
    template_key: str = Query(...),
    year: int = Query(...),
    department_id: Optional[uuid.UUID] = Query(None),
    equipment_type_id: Optional[int] = Query(None),
    db: Session = Depends(get_vendor_db),
    user: dict = Depends(get_current_user),
):
    org_id = _org_id(user)
    is_admin, user_dept_id = _dept_scope(user, org_id, db)
    effective_department_id = department_id or (None if is_admin else user_dept_id)
    dept_ids = _collect_department_ids(effective_department_id, db) if effective_department_id else None

    eq_q = db.query(Equipment).filter(Equipment.status != "retired")
    if org_id:
        eq_q = eq_q.filter(Equipment.organization_id == org_id)
    if equipment_type_id:
        eq_q = eq_q.filter(Equipment.equipment_type_id == equipment_type_id)
    if dept_ids:
        eq_q = eq_q.filter(Equipment.department_id.in_(dept_ids))
    eq_list = eq_q.all()
    if not eq_list:
        return {"equipment": []}
    eq_map = {e.id: e for e in eq_list}

    ta_rows = (
        db.query(TestAnalytics)
        .filter(
            TestAnalytics.equipment_id.in_(list(eq_map.keys())),
            TestAnalytics.template_key == template_key,
            func.extract("year", TestAnalytics.tested_at) == year,
        )
        .all()
    )
    latest_by_eq: dict = {}
    count_by_eq: dict = {}
    for r in ta_rows:
        cur = latest_by_eq.get(r.equipment_id)
        if cur is None or (r.tested_at and cur.tested_at and r.tested_at > cur.tested_at):
            latest_by_eq[r.equipment_id] = r
        count_by_eq[r.equipment_id] = count_by_eq.get(r.equipment_id, 0) + 1

    dept_ids_for_names = list({e.department_id for e in eq_map.values() if e.department_id})
    dept_name_map = {
        d.id: d.name
        for d in db.query(OrgDepartment).filter(OrgDepartment.id.in_(dept_ids_for_names)).all()
    }

    equipment = []
    for eq_id, r in latest_by_eq.items():
        eq = eq_map.get(eq_id)
        if not eq:
            continue
        equipment.append({
            "equipment_id": str(eq_id),
            "ueic": eq.ueic,
            "department_id": str(eq.department_id) if eq.department_id else None,
            "department_name": dept_name_map.get(eq.department_id),
            "health_score": float(r.health_score) if r.health_score is not None else None,
            "condition_summary": r.condition_summary,
            "tested_at": r.tested_at.isoformat() if r.tested_at else None,
            "test_result_id": str(r.test_result_id),
            "test_count": count_by_eq.get(eq_id, 0),
        })
    equipment.sort(key=lambda x: x["health_score"] if x["health_score"] is not None else 999.0)

    # Raw per-test points (not deduped to latest-per-equipment) — lets the
    # popup plot every test by its actual date, so a same-year recovery or
    # decline is visible instead of only showing each equipment's latest.
    tests = [
        {
            "equipment_id": str(r.equipment_id),
            "tested_at": r.tested_at.isoformat() if r.tested_at else None,
            "health_score": float(r.health_score) if r.health_score is not None else None,
            "condition_summary": r.condition_summary,
        }
        for r in ta_rows
    ]
    return {"equipment": equipment, "tests": tests}


@router.get("/equipment/{equipment_id}/tests", summary="Test instances for one equipment, scoped to a template_key/year")
def get_equipment_tests_for_slice(
    equipment_id: uuid.UUID,
    template_key: str = Query(...),
    year: Optional[int] = Query(None),
    db: Session = Depends(get_vendor_db),
    user: dict = Depends(get_current_user),
):
    q = db.query(TestAnalytics).filter(
        TestAnalytics.equipment_id == equipment_id,
        TestAnalytics.template_key == template_key,
    )
    if year:
        q = q.filter(func.extract("year", TestAnalytics.tested_at) == year)
    rows = q.order_by(TestAnalytics.tested_at.desc()).all()

    return {
        "equipment_id": str(equipment_id),
        "template_key": template_key,
        "tests": [
            {
                "test_result_id": str(r.test_result_id),
                "tested_at": r.tested_at.isoformat() if r.tested_at else None,
                "health_score": float(r.health_score) if r.health_score is not None else None,
                "condition_summary": r.condition_summary,
                "risk_level": r.risk_level,
                # Step 4 of the drill-down ("Test Details") reuses these existing
                # endpoints unchanged — that's already where per-parameter data
                # (BDV, Water Content, Methane, …) lives via ParameterAnalytics.
                "links": {
                    "detail": f"/analytics/test-results/{r.test_result_id}",
                    "raw": f"/analytics/test-results/{r.test_result_id}/raw",
                },
            }
            for r in rows
        ],
    }
